"""B-10 · Validación de datos del receptor.

Los errores de este informe generan requerimientos del SAT y problemas ante el IMSS, y son
**invisibles en los informes de importes**: todos los demás pueden cuadrar con un NSS mal
capturado.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.informes import b10_validacion_receptor as b10, excel, validadores
from app.informes.base import ContextoInforme
from app.models.configuracion_fiscal import ConfiguracionEmpresa, ParamFiscal
from app.models.enums import OrigenValor, ZonaSalarial
from app.services import configuracion_fiscal as cfg
from tests import factories
from tests.helpers_nomina import insertar_nomina

# Cifras 2026 reales (las de `config/fiscal/param_fiscal.yaml`), para que los topes que
# calculan estas pruebas sean los que producirá la instalación de verdad. Las mismas que usa
# `tests/test_informe_b03.py`.
_UMA_DIARIA = "117.31"
_SM_ZLFN = "440.87"
_SM_GENERAL = "315.04"

_CONFIRMADO_EN = datetime(2026, 8, 6, 12, 0, 0)
_FUENTE = "INEGI, boletín UMA 2026 (fixture de prueba)"


def _p(**kw: object) -> b10.Parametros:
    base = {"fecha_desde": date(2026, 6, 1), "fecha_hasta": date(2026, 7, 31)}
    base.update(kw)
    return b10.Parametros(**base)  # type: ignore[arg-type]


async def _empresa(db: AsyncSession) -> int:
    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    return empresa.empresa_id


async def _sembrar_param(
    db: AsyncSession, clave: str, valor: str, *, confirmado: bool, desde: date = date(2026, 2, 1)
) -> None:
    """Siembra `param_fiscal` **por la puerta de escritura del servicio**, no con un `INSERT`
    directo: así la prueba ejercita el camino real, incluida la lista blanca de claves y el
    invariante de confirmación —que es justo lo que varias de estas pruebas verifican—. Mismo
    helper que `tests/test_informe_b03.py`."""
    await cfg.guardar_param_fiscal(
        db, clave=clave, valor=Decimal(valor), vigencia_desde=desde, origen=OrigenValor.SEMILLA, fuente=_FUENTE
    )
    await db.commit()
    if not confirmado:
        return
    fila = await db.get(ParamFiscal, (clave, desde))
    assert fila is not None
    fila.confirmado_por = "uid-prueba"
    fila.confirmado_en = _CONFIRMADO_EN
    await db.commit()


async def _zona(db: AsyncSession, empresa_id: int, zona: ZonaSalarial) -> None:
    db.add(ConfiguracionEmpresa(empresa_id=empresa_id, zona_salarial=zona))
    await db.commit()


async def _empleado_con_sbc(db: AsyncSession, eid: int, *, uuid: str, sbc: str) -> None:
    """Un empleado bien capturado salvo por el SBC, que es lo que cada prueba de esta sección
    mueve. Sin defectos de estructura, para que la única variable sea el SBC."""
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid=uuid,
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        tipo_regimen="02",
        sbc=sbc,
        sdi=sbc,
        fecha_pago=date(2026, 6, 30),
        fecha_final_pago=date(2026, 6, 30),
    )


def _banderas(resultado: object, clave: str) -> list[object]:
    return [b for b in resultado.banderas if b.clave == clave]  # type: ignore[attr-defined]


def _claves(resultado: object) -> set[str]:
    titulos = [c.titulo for c in resultado.columnas]  # type: ignore[attr-defined]
    i = titulos.index("Validación")
    return {f[i] for f in resultado.filas}  # type: ignore[attr-defined]


async def test_una_fila_por_hallazgo(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="11111111-1111-1111-1111-111111111111",
        rfc_receptor="XAXX010101000",
        curp="VECJ880326HDFLNS09",
        nss="",
        tipo_regimen="02",
        puesto="",
        departamento="",
    )

    resultado = await b10.consultar(db, eid, _p())
    claves = _claves(resultado)
    assert "NSS_FALTANTE" in claves
    assert "PUESTO_VACIO" in claves
    assert "DEPARTAMENTO_VACIO" in claves
    # Una fila por hallazgo, no una por empleado.
    assert len(resultado.filas) >= 3


async def test_datos_correctos_no_generan_hallazgos(db: AsyncSession) -> None:
    """Si un empleado bien capturado produjera hallazgos, el informe sería ruido.

    `antiguedad="P339W"` (no `P330W`): se verificó con `(fecha_final_pago -
    fecha_inicio_rel_laboral).days` que la antigüedad real entre 2020-01-01 y 2026-06-30 es
    de 2372 días (339 semanas exactas serían 2373, 1 día de diferencia, dentro de la
    tolerancia de dos semanas). `P330W` (2310 días) queda a 62 días de esa fecha —fuera de
    tolerancia— y habría hecho fallar esta prueba con un falso `ANTIGUEDAD_INCONSISTENTE`
    sobre un dato que en realidad está bien capturado. Es el mismo tipo de verificación que
    el dígito verificador del NSS: no se copia un valor de ejemplo sin comprobar que cumple.
    """
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="22222222-2222-2222-2222-222222222222",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        tipo_regimen="02",
        puesto="Auxiliar",
        departamento="Administración",
        sbc="500.00",
        sdi="600.00",
        fecha_inicio_rel_laboral=date(2020, 1, 1),
        antiguedad="P339W",
        banco="002",
        cuenta_bancaria="1234567890",
        percepciones=[("001", "001", "Sueldo", "7500.00", "0.00")],
        dias="15.000",
    )

    resultado = await b10.consultar(db, eid, _p())
    assert resultado.filas == [], _claves(resultado)


async def test_rfc_curp_inconsistente(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="33333333-3333-3333-3333-333333333333",
        rfc_receptor="VECJ880326XXX",
        curp="AAAA880326HDFLNS09",
        nss="12345678903",
    )
    assert "RFC_CURP_INCONSISTENTE" in _claves(await b10.consultar(db, eid, _p()))


async def test_curp_duplicada_y_rfc_duplicado(db: AsyncSession) -> None:
    """Validaciones de conjunto: no se ven mirando un CFDI aislado."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="44444444-4444-4444-4444-444444444441",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
    )
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="44444444-4444-4444-4444-444444444442",
        rfc_receptor="AAAA880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
    )

    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "CURP_DUPLICADA" in claves
    assert "NSS_DUPLICADO" in claves


async def test_datos_cambiantes_entre_periodos(db: AsyncSession) -> None:
    """Un mismo RFC con distinto NSS entre quincenas: error de captura que solo se ve
    comparando periodos.

    La prueba usaba antes dos CURP distintas, pero ese caso concreto lo reporta `RFC_DUPLICADO`
    y `DATOS_CAMBIANTES` ya no lo cubre (revisión final: dos claves de severidad alta para un
    solo defecto producían dos filas en un informe cuyo grano es "una fila = algo que
    corregir"). Ver `test_rfc_duplicado_es_la_unica_clave_del_rfc_con_varias_curp`."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="55555555-5555-5555-5555-555555555551",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        fecha_pago=date(2026, 6, 30),
        fecha_final_pago=date(2026, 6, 30),
    )
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="55555555-5555-5555-5555-555555555552",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="98765432105",
        fecha_pago=date(2026, 7, 15),
        fecha_final_pago=date(2026, 7, 15),
    )

    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "DATOS_CAMBIANTES" in claves


async def test_rfc_duplicado_es_la_unica_clave_del_rfc_con_varias_curp(db: AsyncSession) -> None:
    """`RFC_DUPLICADO`: un mismo RFC con dos CURP distintas en el rango.

    Doble propósito. (1) Cubre `RFC_DUPLICADO`, que no tenía ninguna prueba: se podía borrar del
    módulo sin que la suite se enterara. (2) Fija la decisión de la revisión final sobre el
    reporte doble: este hecho lo reporta **solo** `RFC_DUPLICADO`, no también `DATOS_CAMBIANTES`.
    """
    eid = await _empresa(db)
    for sufijo, curp, fecha in (
        ("1", "VECJ880326HDFLNS09", date(2026, 6, 30)),
        ("2", "VECJ880326HDFLNS08", date(2026, 7, 15)),
    ):
        await insertar_nomina(
            db,
            empresa_id=eid,
            uuid=f"5555555a-5555-5555-5555-55555555555{sufijo}",
            rfc_receptor="VECJ880326XXX",
            curp=curp,
            nss="12345678903",
            fecha_pago=fecha,
            fecha_final_pago=fecha,
        )

    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "RFC_DUPLICADO" in claves
    assert "DATOS_CAMBIANTES" not in claves, "el mismo defecto no debe generar dos filas"


async def test_rfc_estructura(db: AsyncSession) -> None:
    """El receptor de un CFDI de nómina es siempre persona física: un RFC de persona moral
    (3 letras iniciales) es un error de captura."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000001-0000-0000-0000-000000000001",
        rfc_receptor="EKU9003173C9",
        curp="EKUX900317HDFLNS01",
        nss="12345678903",
    )
    assert "RFC_ESTRUCTURA" in _claves(await b10.consultar(db, eid, _p()))


async def test_curp_estructura(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000002-0000-0000-0000-000000000002",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326XDFLNS09",  # ni H ni M en la posición del sexo
        nss="12345678903",
    )
    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "CURP_ESTRUCTURA" in claves
    # Con la estructura general mal, las posiciones 12-13 no son fiables: no se marca también
    # CURP_ENTIDAD (regla de aplicabilidad del docstring del módulo).
    assert "CURP_ENTIDAD" not in claves


async def test_curp_entidad(db: AsyncSession) -> None:
    """Estructura correcta pero `ZZ` no es una clave de entidad federativa del RENAPO."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000003-0000-0000-0000-000000000003",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HZZLNS09",
        nss="12345678903",
    )
    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "CURP_ENTIDAD" in claves
    assert "CURP_ESTRUCTURA" not in claves


async def test_nss_longitud(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000004-0000-0000-0000-000000000004",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="123456789",  # 9 caracteres
    )
    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "NSS_LONGITUD" in claves
    # Luhn sobre una longitud equivocada no es significativo y no se evalúa.
    assert "NSS_DIGITO_VERIFICADOR" not in claves


async def test_nss_digito_verificador(db: AsyncSession) -> None:
    """11 dígitos exactos, pero el dígito verificador de Luhn no cuadra (`...01` en vez de
    `...03`, verificado en `tests/test_informes_validadores.py`)."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000005-0000-0000-0000-000000000005",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678901",
    )
    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "NSS_DIGITO_VERIFICADOR" in claves
    assert "NSS_LONGITUD" not in claves


async def test_sbc_cero(db: AsyncSession) -> None:
    """Un SBC en cero bajo `tipo_regimen='02'` (obligatorio IMSS) es imposible: sin base de
    cotización no hay cuota que enterar."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000006-0000-0000-0000-000000000006",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        tipo_regimen="02",
        sbc="0.00",
        sdi="600.00",
    )
    assert "SBC_CERO" in _claves(await b10.consultar(db, eid, _p()))


async def test_sdi_cero(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000007-0000-0000-0000-000000000007",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        sbc="500.00",
        sdi="0.00",
    )
    assert "SDI_CERO" in _claves(await b10.consultar(db, eid, _p()))


async def test_sdi_menor_sbc(db: AsyncSession) -> None:
    """B-10.R1: severidad **media**, no alta. Un SDI algo menor que el SBC es teóricamente
    normal (bases distintas); por debajo del 80% ya es sospechoso de captura."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000008-0000-0000-0000-000000000008",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        sbc="1000.00",
        sdi="700.00",  # 70% del SBC, por debajo del umbral de 80%
        dias="15.000",
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
    )
    resultado = await b10.consultar(db, eid, _p())
    titulos = [c.titulo for c in resultado.columnas]
    fila = next(f for f in resultado.filas if f[titulos.index("Validación")] == "SDI_MENOR_SBC")
    assert fila[titulos.index("Severidad")] == "media"


async def test_sdi_no_menor_al_80_por_ciento_del_sbc_no_dispara(db: AsyncSession) -> None:
    """La otra mitad de la prueba anterior: sin ella, una comparación invertida (o un umbral
    puesto al revés) pasaría igual y no protegería nada."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a0000009-0000-0000-0000-000000000009",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        sbc="1000.00",
        sdi="900.00",  # 90% del SBC: menor que el SBC, pero dentro de lo normal
        dias="15.000",
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
    )
    assert "SDI_MENOR_SBC" not in _claves(await b10.consultar(db, eid, _p()))


async def test_antiguedad_inconsistente(db: AsyncSession) -> None:
    """`@Antigüedad` declarada contra el cálculo desde `fecha_inicio_rel_laboral`, con
    tolerancia de dos semanas para absorber la aproximación de la conversión ISO 8601.

    `P52W` (364 días) contra un alta de 2020-01-01 y un cierre de 2026-06-30 (2372 días reales):
    más de 5 años de diferencia, muy fuera de la tolerancia."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a000000a-0000-0000-0000-00000000000a",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        fecha_final_pago=date(2026, 6, 30),
        fecha_inicio_rel_laboral=date(2020, 1, 1),
        antiguedad="P52W",
    )
    assert "ANTIGUEDAD_INCONSISTENTE" in _claves(await b10.consultar(db, eid, _p()))


async def test_cuenta_con_longitud_de_clabe_no_dispara_cuenta_invalida(db: AsyncSession) -> None:
    """La mitad **negativa** de `CUENTA_INVALIDA`: 18 caracteres es una CLABE interbancaria y no
    debe marcarse. Sin ella, una comprobación que se disparara con cualquier cuenta pasaría igual.

    El nombre dice solo esto a propósito: la versión anterior se llamaba
    `..._es_invalida_y_la_correcta_no` y anunciaba las dos direcciones aseverando una sola. La mitad
    positiva (12 caracteres → sí dispara) vive en
    `test_banco_sin_cuenta_y_cuenta_invalida`."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="a000000b-0000-0000-0000-00000000000b",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        banco="002",
        cuenta_bancaria="1" * 18,
    )
    assert "CUENTA_INVALIDA" not in _claves(await b10.consultar(db, eid, _p()))


async def test_conteo_de_validaciones_ejecutadas(db: AsyncSession) -> None:
    """**La bandera que hace auditable al propio informe** (§13 del diseño, el mismo
    razonamiento que `identidades_b00.verificar` y su `cotejos`).

    Una validación que no corre no puede fallar, así que `filas == []` no distingue "los datos
    están bien" de "no se validó nada" — y una prueba que asevera eso pasa **más fácil** cuando
    alguien borra una comprobación. El revisor de la fase 2 lo comprobó por mutación: borró siete
    de las 21 validaciones y la suite siguió verde. Con este número aseverado, borrar una la
    rompe.

    El fixture tiene **todos** los campos poblados, así que se ejecutan las 15 validaciones por
    empleado que no son mutuamente excluyentes, más las 3 de conjunto y 1 entre periodos del
    único RFC del rango.

    **Y desde la fase 3 el número es variable por diseño**, que es lo que esta prueba y su
    gemela `test_el_conteo_sube_cuando_la_configuracion_fiscal_esta_completa` fijan juntas.
    Aquí **no** hay configuración fiscal —el estado de la instalación real hoy—, así que las dos
    validaciones de SBC no corren y **no se cuentan**. Contarlas igual sería peor que no tener el
    conteo: la bandera existe para hacer auditable al informe, y un número que afirma haber
    ejecutado comprobaciones que no corrieron es exactamente la clase de afirmación falsa que
    esta bandera vino a impedir.
    """
    eid = await _empresa(db)
    await _empleado_completo(db, eid, uuid="a000000c-0000-0000-0000-00000000000c")

    resultado = await b10.consultar(db, eid, _p())
    bandera = next(b for b in resultado.banderas if b.clave == "VALIDACIONES_EJECUTADAS")
    esperadas = b10.VALIDACIONES_POR_EMPLEADO_COMPLETO + b10.VALIDACIONES_DE_CONJUNTO + b10.VALIDACIONES_ENTRE_PERIODOS_POR_RFC
    assert f"Se ejecutaron {esperadas} validaciones" in bandera.mensaje, bandera.mensaje
    assert bandera.severidad == "baja" and bandera.ambito == "informe"
    # Lo no contado, contrastado contra lo que dicen las banderas de la misma hoja: las dos
    # validaciones de SBC no corrieron y hay una bandera por cada dato que falta.
    assert {b.clave for b in resultado.banderas} >= {"FALTA_UMA", "FALTA_ZONA_SALARIAL"}
    # Y el fixture no produce hallazgos: es el mismo camino limpio de
    # `test_datos_correctos_no_generan_hallazgos`, pero ahora con el conteo que lo hace auditable.
    assert resultado.filas == [], _claves(resultado)


async def _empleado_completo(db: AsyncSession, eid: int, *, uuid: str) -> None:
    """La fotografía con **todos** los campos poblados y ningún defecto, que es la premisa de
    las dos pruebas del conteo. `antiguedad="P339W"`: ver
    `test_datos_correctos_no_generan_hallazgos` para por qué ese valor y no `P330W`."""
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid=uuid,
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        tipo_regimen="02",
        puesto="Auxiliar",
        departamento="Administración",
        sbc="500.00",
        sdi="600.00",
        fecha_inicio_rel_laboral=date(2020, 1, 1),
        antiguedad="P339W",
        banco="002",
        cuenta_bancaria="1234567890",
        percepciones=[("001", "001", "Sueldo", "7500.00", "0.00")],
        dias="15.000",
    )


async def test_el_conteo_sube_cuando_la_configuracion_fiscal_esta_completa(db: AsyncSession) -> None:
    """La gemela de la anterior: **el mismo empleado**, ahora con la UMA, la zona salarial y el
    mínimo de esa zona confirmados. Las dos validaciones de SBC sí corren, así que el conteo sube
    en `VALIDACIONES_QUE_EXIGEN_CONFIGURACION` y no queda ninguna bandera de configuración.

    Las dos pruebas juntas son lo que hace honesto al conteo. Con una sola, un conteo fijo —o uno
    que sumara las dos validaciones sin haberlas evaluado— pasaría igual.

    El SBC de 500 está entre los dos límites (mínimo general 315.04, tope 25 × 117.31 = 2932.75),
    así que ninguna de las dos dispara: lo que se mide aquí es que **se ejecutaron**, no que
    encontraran algo."""
    eid = await _empresa(db)
    await _zona(db, eid, ZonaSalarial.GENERAL)
    await _sembrar_param(db, "UMA_DIARIA", _UMA_DIARIA, confirmado=True)
    await _sembrar_param(db, "SALARIO_MINIMO_GENERAL", _SM_GENERAL, confirmado=True, desde=date(2026, 1, 1))
    await _empleado_completo(db, eid, uuid="a000000d-0000-0000-0000-00000000000d")

    resultado = await b10.consultar(db, eid, _p())
    bandera = next(b for b in resultado.banderas if b.clave == "VALIDACIONES_EJECUTADAS")
    esperadas = (
        b10.VALIDACIONES_POR_EMPLEADO_COMPLETO
        + b10.VALIDACIONES_QUE_EXIGEN_CONFIGURACION
        + b10.VALIDACIONES_DE_CONJUNTO
        + b10.VALIDACIONES_ENTRE_PERIODOS_POR_RFC
    )
    assert f"Se ejecutaron {esperadas} validaciones" in bandera.mensaje, bandera.mensaje
    # Nada que configurar: ninguna de las cuatro banderas de degradación.
    assert not [
        b for b in resultado.banderas
        if b.clave in ("FALTA_UMA", "UMA_SIN_CONFIRMAR", "FALTA_ZONA_SALARIAL", "FALTA_SALARIO_MINIMO")
    ], [b.clave for b in resultado.banderas]
    assert resultado.filas == [], _claves(resultado)


async def test_ninguna_celda_de_datos_lleva_curp_ni_nss_completos(db: AsyncSession) -> None:
    """**El hallazgo Critical de la revisión final.** Con `enmascarar_datos_personales=True` las
    columnas `CURP` y `NSS` salían enmascaradas (`****NS09`) pero la columna "Descripción del
    hallazgo" —que no es sensible ni puede serlo sin volverse ilegible— interpolaba el dato crudo:
    6 de 7 filas filtraban la CURP y el NSS completos de cada empleado con algún defecto.

    Escenario real: un usuario con rol `CONSULTA` —el que por diseño solo puede generar
    enmascarado— pide B-10 y lo manda por correo a quien va a corregir las capturas.

    Se recorre el libro **ya escrito**, no el `ResultadoInforme`, porque el enmascaramiento lo
    aplica el motor: es el archivo que circula por correo lo que hay que auditar. Y se recorren
    TODAS las celdas de texto, no solo las declaradas sensibles — verificar el mecanismo no es
    verificar el resultado.

    **Dos comprobaciones, no una, porque la de patrón sola tiene un punto ciego.**
    `dato_personal_en_texto` busca la *estructura* de una CURP o de un NSS, así que no puede ver
    una CURP **mal formada** interpolada en un mensaje — y el mensaje de `CURP_ESTRUCTURA` es
    justamente el que solo puede llevar una CURP mal formada. Un dato personal defectuoso sigue
    siendo un dato personal (una CURP con un carácter equivocado identifica igual a la persona),
    así que se agrega la comprobación por **valor**. Esto se descubrió por mutación: reintroducir
    el dato crudo en el mensaje de `CURP_ESTRUCTURA` dejaba pasar la versión de esta prueba que
    solo hacía la comprobación de patrón.

    **Los dos huecos que tenía esta red, ya cerrados** (no había fuga viva por ninguno cuando se
    cerraron: esto es blindaje de regresión, no un incidente):

    1. **La comprobación por valor era por fila.** Comparaba cada celda contra la CURP y el NSS de
       *ese* empleado, así que una CURP **mal formada de otro** empleado interpolada en un mensaje
       se escapaba de las dos comprobaciones a la vez: del patrón por estar mal formada, y del
       valor por compararse contra la fila equivocada. Ahora se recoge el conjunto de CURP y NSS de
       **todos** los empleados del universo una sola vez y se busca cualquiera de ellos en
       cualquier celda (ver `test_una_curp_de_otro_empleado_en_una_celda_es_fuga`).
    2. **Solo se recorría la hoja `Datos`.** Las hojas `Banderas`, `Parámetros` y `Diccionario`
       viajan en el **mismo archivo**, así que un mensaje de bandera que interpolara un dato
       personal salía igual de la empresa — y es donde el riesgo es más real, porque los mensajes
       de bandera son texto libre. Ahora se auditan las cuatro (ver
       `test_un_dato_personal_en_la_hoja_banderas_es_fuga`).

    Las dos correcciones viven en `validadores.fugas_de_datos_personales_en_libro`, compartida con
    `scripts/verificar_informes.py`: tenerlas en un solo sitio es lo que evita que las dos copias
    de la red vuelvan a divergir con un hueco distinto cada una.
    """
    eid = await _empresa(db)
    # Cuatro CFDI que entre ellos disparan **todas** las validaciones cuyo mensaje interpolaba el
    # dato: estructura de CURP y de NSS, entidad, inconsistencia RFC/CURP, los tres duplicados y
    # `DATOS_CAMBIANTES`. Los valores son inventados (regla 12).
    valores_por_empleado = [
        # (uuid, rfc, curp, nss, fecha)
        ("c0000001-0000-0000-0000-000000000001", "VECJ880326XXX", "VECJ880326HDFLNS09", "12345678901", date(2026, 6, 30)),
        ("c0000002-0000-0000-0000-000000000002", "AAAA880326XXX", "VECJ880326HDFLNS09", "12345678901", date(2026, 7, 15)),
        ("c0000003-0000-0000-0000-000000000003", "AAAA880326XXX", "AAAA880326HZZLNS08", "98765432101", date(2026, 7, 31)),
        # CURP mal formada (13 caracteres) y NSS de longitud equivocada: dispara
        # `CURP_ESTRUCTURA`, `RFC_CURP_INCONSISTENTE` y `NSS_LONGITUD`, los tres mensajes que la
        # comprobación de patrón por sí sola no puede auditar.
        ("c0000004-0000-0000-0000-000000000004", "BBBB880326XXX", "MALA880326HDF", "123456789", date(2026, 7, 31)),
    ]
    for uuid_cfdi, rfc, curp, nss, fecha in valores_por_empleado:
        await insertar_nomina(
            db,
            empresa_id=eid,
            uuid=uuid_cfdi,
            rfc_receptor=rfc,
            curp=curp,
            nss=nss,
            fecha_pago=fecha,
            fecha_final_pago=fecha,
            banco="002",
            cuenta_bancaria="123456789012",
        )

    p = _p()
    assert p.enmascarar_datos_personales is True, "el default del informe es enmascarar"
    resultado = await b10.consultar(db, eid, p)
    claves = _claves(resultado)
    # Premisa: si el fixture no dispara estas validaciones, la prueba no comprueba sus mensajes.
    for clave_esperada in ("CURP_ESTRUCTURA", "CURP_ENTIDAD", "RFC_CURP_INCONSISTENTE", "NSS_LONGITUD",
                           "NSS_DIGITO_VERIFICADOR", "CURP_DUPLICADA", "NSS_DUPLICADO", "RFC_DUPLICADO",
                           "DATOS_CAMBIANTES"):
        assert clave_esperada in claves, f"el fixture debe disparar {clave_esperada} para auditar su mensaje"

    libro = openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, _ctx(p))))
    # Las cuatro hojas viajan en el mismo archivo, así que se auditan las cuatro.
    assert set(libro.sheetnames) == {"Datos", "Parámetros", "Banderas", "Diccionario"}

    # Nunca el valor en el mensaje de una aserción que falla: sería la misma fuga que denuncia.
    fugas = validadores.fugas_de_datos_personales_en_libro(libro, _valores_personales(valores_por_empleado))
    assert fugas == [], f"datos personales en el libro con enmascaramiento activo: {[f.descripcion for f in fugas]}"


def _ctx(p: b10.Parametros) -> ContextoInforme:
    return ContextoInforme(
        clave=b10.CLAVE,
        nombre=b10.NOMBRE,
        usuario="consulta@test.mx",
        generado_en=datetime(2026, 8, 6, 12, 0),
        parametros=p.model_dump(mode="json"),
        etl_version=1,
    )


def _valores_personales(valores_por_empleado: list[tuple[str, str, str, str, date]]) -> dict[str, list[str]]:
    """CURP y NSS de **todos** los empleados del universo, no los de una fila.

    Es el cierre del primer hueco: comparar celda contra fila dejaba pasar la CURP mal formada de
    *otro* empleado interpolada en un mensaje. Se toman del fixture, que aquí **es** el universo.
    """
    return {
        "CURP": [curp for _uuid, _rfc, curp, _nss, _fecha in valores_por_empleado],
        "NSS": [nss for _uuid, _rfc, _curp, nss, _fecha in valores_por_empleado],
    }


async def test_una_curp_de_otro_empleado_en_una_celda_es_fuga(db: AsyncSession) -> None:
    """**Cierre del hueco A**, comprobado con la fuga que antes se escapaba de las dos redes.

    Una CURP **mal formada de otro** empleado interpolada en la celda de un tercero: invisible para
    el detector por patrón (está mal formada) e invisible para la comparación por valor mientras
    esta fuera por fila (el dato es de otra fila). Con el conjunto del universo completo se ve.

    Se inyecta sobre el libro ya escrito, no sobre el módulo: lo que se está probando es la red,
    no B-10 — que hoy está limpio.
    """
    eid = await _empresa(db)
    curp_ajena = "MALA880326HDF"  # 13 caracteres: no cumple el patrón de una CURP
    valores_por_empleado = [
        ("d0000001-0000-0000-0000-000000000001", "VECJ880326XXX", "VECJ880326HDFLNS09", "12345678901", date(2026, 6, 30)),
        ("d0000002-0000-0000-0000-000000000002", "BBBB880326XXX", curp_ajena, "98765432101", date(2026, 7, 15)),
    ]
    for uuid_cfdi, rfc, curp, nss, fecha in valores_por_empleado:
        await insertar_nomina(db, empresa_id=eid, uuid=uuid_cfdi, rfc_receptor=rfc, curp=curp, nss=nss,
                              fecha_pago=fecha, fecha_final_pago=fecha)

    p = _p()
    resultado = await b10.consultar(db, eid, p)
    libro = openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, _ctx(p))))
    valores = _valores_personales(valores_por_empleado)

    # Premisa: el libro real está limpio, así que la fuga que se detecte es la inyectada.
    assert validadores.fugas_de_datos_personales_en_libro(libro, valores) == []

    # La CURP del segundo empleado, escrita en la fila de **otro** empleado: es el caso exacto que
    # la comparación por fila no podía ver.
    hoja = libro["Datos"]
    titulos = [celda.value for celda in hoja[1]]
    columna_rfc, columna_descripcion = titulos.index("RFC empleado") + 1, titulos.index("Descripción del hallazgo") + 1
    fila_de_otro = next(
        numero for numero in range(2, hoja.max_row + 1) if hoja.cell(row=numero, column=columna_rfc).value == "VECJ880326XXX"
    )
    hoja.cell(row=fila_de_otro, column=columna_descripcion, value=f"Revisar la captura de {curp_ajena} en el sistema de nómina.")

    fugas = validadores.fugas_de_datos_personales_en_libro(libro, valores)
    assert [(f.hoja, f.tipo, f.deteccion) for f in fugas] == [("Datos", "CURP", "valor")], [f.descripcion for f in fugas]
    assert validadores.dato_personal_en_texto(hoja.cell(row=fila_de_otro, column=columna_descripcion).value) is None, (
        "la premisa del hueco A: esta CURP está mal formada, así que el detector por patrón no la ve"
    )


async def test_un_dato_personal_en_la_hoja_banderas_es_fuga(db: AsyncSession) -> None:
    """**Cierre del hueco B**: las hojas `Banderas`, `Parámetros` y `Diccionario` viajan en el
    mismo archivo que `Datos`, así que una bandera cuyo mensaje interpolara una CURP saldría igual
    de la empresa. Es donde el riesgo es más real —los mensajes de bandera son texto libre— y ya
    hubo un incidente exactamente así con los mensajes de B-10.
    """
    eid = await _empresa(db)
    valores_por_empleado = [
        ("d0000003-0000-0000-0000-000000000003", "VECJ880326XXX", "VECJ880326HDFLNS09", "12345678901", date(2026, 6, 30)),
    ]
    for uuid_cfdi, rfc, curp, nss, fecha in valores_por_empleado:
        await insertar_nomina(db, empresa_id=eid, uuid=uuid_cfdi, rfc_receptor=rfc, curp=curp, nss=nss,
                              fecha_pago=fecha, fecha_final_pago=fecha)

    p = _p()
    resultado = await b10.consultar(db, eid, p)
    # B-10 sí emite `VALIDACIONES_EJECUTADAS`, así que la hoja `Banderas` tiene al menos una fila
    # que mutar: la premisa de la prueba no depende de que haya hallazgos.
    assert resultado.banderas, "el fixture debe producir alguna bandera que auditar"
    libro = openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, _ctx(p))))
    valores = _valores_personales(valores_por_empleado)

    assert validadores.fugas_de_datos_personales_en_libro(libro, valores) == []

    hoja = libro["Banderas"]
    # Columna 4 = "Mensaje" (ver `excel._escribir_banderas`); fila 2 = la primera bandera.
    hoja.cell(row=2, column=4, value=f"Corregir la CURP {valores['CURP'][0]} del empleado.")

    fugas = validadores.fugas_de_datos_personales_en_libro(libro, valores)
    # Esta CURP sí está bien formada, así que la ven las dos redes: se reporta por patrón y por
    # valor. Lo que importa es que se reporte, y que se reporte en la hoja `Banderas`.
    assert {(f.hoja, f.columna, f.tipo) for f in fugas} == {("Banderas", "Mensaje", "CURP")}, [f.descripcion for f in fugas]
    assert {f.deteccion for f in fugas} == {"patrón", "valor"}


async def test_sdi_menor_al_salario_diario_implicito(db: AsyncSession) -> None:
    """El SDI declarado es menor que el sueldo diario que se deduce del propio CFDI."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="66666666-6666-6666-6666-666666666666",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        sbc="500.00",
        sdi="100.00",
        dias="15.000",
        percepciones=[("001", "001", "Sueldo", "7500.00", "0.00")],
    )
    assert "SDI_MENOR_SD_IMPLICITO" in _claves(await b10.consultar(db, eid, _p()))


async def test_fecha_inicio_posterior_al_periodo(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="77777777-7777-7777-7777-777777777777",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        fecha_final_pago=date(2026, 6, 30),
        fecha_inicio_rel_laboral=date(2026, 12, 1),
    )
    assert "FECHA_INICIO_POSTERIOR" in _claves(await b10.consultar(db, eid, _p()))


async def test_banco_sin_cuenta_y_cuenta_invalida(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="88888888-8888-8888-8888-888888888881",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        banco="002",
        cuenta_bancaria="",
    )
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="88888888-8888-8888-8888-888888888882",
        rfc_receptor="AAAA880326XXX",
        curp="AAAA880326HDFLNS09",
        nss="12345678903",
        banco="002",
        cuenta_bancaria="123456789012",
    )

    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "BANCO_SIN_CUENTA" in claves
    assert "CUENTA_INVALIDA" in claves


async def test_severidad_minima_filtra(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="99999999-9999-9999-9999-999999999999",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="12345678903",
        puesto="",
        departamento="",
    )

    todas = await b10.consultar(db, eid, _p())
    solo_altas = await b10.consultar(db, eid, _p(severidad_minima="alta"))
    assert "PUESTO_VACIO" in _claves(todas)
    assert "PUESTO_VACIO" not in _claves(solo_altas)


async def test_sin_configuracion_fiscal_las_dos_validaciones_de_sbc_no_se_evaluan(db: AsyncSession) -> None:
    """**El estado de la instalación real hoy**, y la degradación que la tarea 8 tenía que dejar
    bien: sin UMA y sin zona salarial las dos validaciones de SBC no se evalúan, ni siquiera con
    un SBC absurdo que las dispararía las dos.

    Esta prueba comprobaba antes que las dos validaciones **no existían** (fase 2, fuera de
    alcance por falta de configuración). Ahora existen, así que comprueba lo que de verdad
    importa: que su ausencia sigue siendo una decisión declarada —con bandera que dice cuál dato
    falta y adónde ir— y no un silencio. Sin las banderas, un SBC de 999,999 pasaría inadvertido
    y el informe se vería idéntico a uno donde todo está bien."""
    eid = await _empresa(db)
    await _empleado_con_sbc(db, eid, uuid="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", sbc="999999.00")

    resultado = await b10.consultar(db, eid, _p())
    claves = _claves(resultado)
    assert "SBC_SOBRE_TOPE" not in claves
    assert "SBC_BAJO_MINIMO" not in claves
    # Pero no en silencio: una bandera por causa, con ámbito `informe`.
    assert len(_banderas(resultado, "FALTA_UMA")) == 1, [b.clave for b in resultado.banderas]
    assert len(_banderas(resultado, "FALTA_ZONA_SALARIAL")) == 1, [b.clave for b in resultado.banderas]
    assert all(b.severidad == "alta" and b.ambito == "informe" for b in resultado.banderas if b.clave.startswith("FALTA_"))  # type: ignore[attr-defined]


async def test_sbc_sobre_tope_dispara_con_la_uma_confirmada(db: AsyncSession) -> None:
    """`SBC_SOBRE_TOPE`: 25 UMA diarias es el límite superior de cotización del art. 28 de la
    LSS. Con la UMA 2026 (117.31) el tope son 2932.75, y un SBC de 3000 lo excede.

    Severidad **media**, no alta: un SBC sobre el tope no daña al trabajador, sobrestima la
    cuota, y el IMSS la recorta al tope de todas formas."""
    eid = await _empresa(db)
    await _sembrar_param(db, "UMA_DIARIA", _UMA_DIARIA, confirmado=True)
    await _empleado_con_sbc(db, eid, uuid="e0000001-0000-0000-0000-000000000001", sbc="3000.00")

    resultado = await b10.consultar(db, eid, _p())
    titulos = [c.titulo for c in resultado.columnas]
    fila = next(f for f in resultado.filas if f[titulos.index("Validación")] == "SBC_SOBRE_TOPE")
    assert fila[titulos.index("Severidad")] == "media"
    # Y con la UMA confirmada ya no hay nada que reportar por ese lado.
    assert _banderas(resultado, "FALTA_UMA") == []
    assert _banderas(resultado, "UMA_SIN_CONFIRMAR") == []


async def test_sbc_justo_por_debajo_del_tope_no_dispara(db: AsyncSession) -> None:
    """La gemela negativa. Sin ella, una comparación invertida —o el tope calculado con la UMA
    en vez de con 25 UMA— pasaría igual: 25 × 117.31 = 2932.75, así que 2900 está dentro."""
    eid = await _empresa(db)
    await _sembrar_param(db, "UMA_DIARIA", _UMA_DIARIA, confirmado=True)
    await _empleado_con_sbc(db, eid, uuid="e0000002-0000-0000-0000-000000000002", sbc="2900.00")

    assert "SBC_SOBRE_TOPE" not in _claves(await b10.consultar(db, eid, _p()))


async def test_sbc_de_400_bajo_minimo_en_la_zona_frontera(db: AsyncSession) -> None:
    """**La mitad frontera de la pareja que prueba que la zona cambia el resultado.** El mínimo
    2026 de la Zona Libre de la Frontera Norte —donde está Ciudad Juárez— es 440.87, así que un
    SBC de 400 está por debajo. Severidad **alta**: es incumplimiento directo."""
    eid = await _empresa(db)
    await _zona(db, eid, ZonaSalarial.ZLFN)
    await _sembrar_param(db, "SALARIO_MINIMO_ZLFN", _SM_ZLFN, confirmado=True, desde=date(2026, 1, 1))
    await _empleado_con_sbc(db, eid, uuid="e0000003-0000-0000-0000-000000000003", sbc="400.00")

    resultado = await b10.consultar(db, eid, _p())
    titulos = [c.titulo for c in resultado.columnas]
    fila = next(f for f in resultado.filas if f[titulos.index("Validación")] == "SBC_BAJO_MINIMO")
    assert fila[titulos.index("Severidad")] == "alta"
    assert _banderas(resultado, "FALTA_ZONA_SALARIAL") == []
    assert _banderas(resultado, "FALTA_SALARIO_MINIMO") == []


async def test_el_mismo_sbc_de_400_no_dispara_en_la_zona_general(db: AsyncSession) -> None:
    """**La otra mitad de la pareja, y la razón de ser de toda la configuración de zona.** El
    mismo SBC de 400, el mismo empleado y la misma fecha: en la zona general el mínimo 2026 es
    315.04 y 400 lo cumple.

    Sin esta pareja nada protegería a `zona_salarial`: una implementación que ignorara la zona y
    leyera siempre el mínimo de la frontera pasaría la prueba anterior. La diferencia entre las
    dos zonas es de casi un 40%, que es exactamente por lo que el campo nace nulo en vez de
    adivinarse."""
    eid = await _empresa(db)
    await _zona(db, eid, ZonaSalarial.GENERAL)
    # Los dos mínimos capturados y confirmados: la única cosa que distingue las dos pruebas es
    # la zona configurada en la empresa, no qué renglones existen.
    await _sembrar_param(db, "SALARIO_MINIMO_GENERAL", _SM_GENERAL, confirmado=True, desde=date(2026, 1, 1))
    await _sembrar_param(db, "SALARIO_MINIMO_ZLFN", _SM_ZLFN, confirmado=True, desde=date(2026, 1, 1))
    await _empleado_con_sbc(db, eid, uuid="e0000004-0000-0000-0000-000000000004", sbc="400.00")

    resultado = await b10.consultar(db, eid, _p())
    assert "SBC_BAJO_MINIMO" not in _claves(resultado)
    assert _banderas(resultado, "FALTA_ZONA_SALARIAL") == []


async def test_sin_zona_salarial_no_se_evalua_el_minimo_aunque_los_valores_esten_capturados(
    db: AsyncSession,
) -> None:
    """`salario_minimo_de_empresa` devuelve `None` **sin mirar los valores** cuando la zona no
    está configurada: aquí los dos mínimos están capturados y confirmados, y aun así
    `SBC_BAJO_MINIMO` no corre.

    Es la protección contra el falso negativo: si se asumiera "GENERAL", este SBC de 400 —que en
    la frontera incumple— saldría como que cumple, y nadie se enteraría."""
    eid = await _empresa(db)
    await _sembrar_param(db, "SALARIO_MINIMO_GENERAL", _SM_GENERAL, confirmado=True, desde=date(2026, 1, 1))
    await _sembrar_param(db, "SALARIO_MINIMO_ZLFN", _SM_ZLFN, confirmado=True, desde=date(2026, 1, 1))
    await _empleado_con_sbc(db, eid, uuid="e0000005-0000-0000-0000-000000000005", sbc="400.00")

    resultado = await b10.consultar(db, eid, _p())
    assert "SBC_BAJO_MINIMO" not in _claves(resultado)
    faltantes = _banderas(resultado, "FALTA_ZONA_SALARIAL")
    assert len(faltantes) == 1
    # Manda a la pantalla correcta: el hueco es de la empresa, no del catálogo fiscal.
    assert "Configuración › Empresa" in faltantes[0].mensaje  # type: ignore[attr-defined]
    assert _banderas(resultado, "FALTA_SALARIO_MINIMO") == []


async def test_zona_configurada_sin_minimo_confirmado_manda_a_la_otra_pantalla(db: AsyncSession) -> None:
    """La causa gemela: con la zona configurada pero sin mínimo confirmado para esa zona, la
    validación tampoco corre — pero la bandera es otra y manda a Configuración › Fiscal. Mandar a
    la pantalla equivocada es lo que vuelve inútil un aviso de configuración."""
    eid = await _empresa(db)
    await _zona(db, eid, ZonaSalarial.ZLFN)
    await _sembrar_param(db, "SALARIO_MINIMO_ZLFN", _SM_ZLFN, confirmado=False, desde=date(2026, 1, 1))
    await _empleado_con_sbc(db, eid, uuid="e0000006-0000-0000-0000-000000000006", sbc="400.00")

    resultado = await b10.consultar(db, eid, _p())
    assert "SBC_BAJO_MINIMO" not in _claves(resultado)
    assert _banderas(resultado, "FALTA_ZONA_SALARIAL") == []
    faltantes = _banderas(resultado, "FALTA_SALARIO_MINIMO")
    assert len(faltantes) == 1
    assert "Configuración › Fiscal" in faltantes[0].mensaje  # type: ignore[attr-defined]


async def test_uma_propuesta_sin_confirmar_no_calcula_pero_la_bandera_trae_su_fuente(db: AsyncSession) -> None:
    """El tercer estado, el que tiene hoy la instalación real: la UMA está capturada y nadie la
    ha confirmado. No calcula —un valor sin confirmar nunca calcula—, pero la bandera lleva la
    **fuente** de la propuesta, que es lo que convierte "falta la UMA, ve a buscarla" en "está
    propuesta con su liga al boletín, confírmala". Un clic contra una búsqueda."""
    eid = await _empresa(db)
    await _sembrar_param(db, "UMA_DIARIA", _UMA_DIARIA, confirmado=False)
    await _empleado_con_sbc(db, eid, uuid="e0000007-0000-0000-0000-000000000007", sbc="999999.00")

    resultado = await b10.consultar(db, eid, _p())
    assert "SBC_SOBRE_TOPE" not in _claves(resultado)
    assert _banderas(resultado, "FALTA_UMA") == []
    sin_confirmar = _banderas(resultado, "UMA_SIN_CONFIRMAR")
    assert len(sin_confirmar) == 1
    assert _FUENTE in sin_confirmar[0].mensaje  # type: ignore[attr-defined]


async def test_un_sbc_en_cero_no_se_reporta_como_bajo_el_minimo(db: AsyncSession) -> None:
    """Regla de aplicabilidad: con `sbc <= 0` ninguna de las dos validaciones corre.

    Un SBC en cero no está "por debajo del mínimo", es una base **ausente**, y `SBC_CERO` ya
    nombra ese defecto donde lo es (`tipo_regimen='02'`). Sin esta regla cada asimilado a
    salarios —que legítimamente no cotiza— saldría con un hallazgo de severidad alta por corrida,
    y el informe cuyo grano es "una fila = algo que corregir" se llenaría de filas que no se
    corrigen."""
    eid = await _empresa(db)
    await _zona(db, eid, ZonaSalarial.ZLFN)
    await _sembrar_param(db, "SALARIO_MINIMO_ZLFN", _SM_ZLFN, confirmado=True, desde=date(2026, 1, 1))
    await _sembrar_param(db, "UMA_DIARIA", _UMA_DIARIA, confirmado=True)
    await _empleado_con_sbc(db, eid, uuid="e0000008-0000-0000-0000-000000000008", sbc="0.00")

    claves = _claves(await b10.consultar(db, eid, _p()))
    assert "SBC_BAJO_MINIMO" not in claves
    assert "SBC_SOBRE_TOPE" not in claves
    # El defecto real sí se reporta, con la clave que le corresponde.
    assert "SBC_CERO" in claves


async def test_curp_y_nss_se_declaran_sensibles(db: AsyncSession) -> None:
    """B-10.R2: este informe es el que más datos personales expone, así que el
    enmascaramiento importa más aquí que en ninguno."""
    eid = await _empresa(db)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb",
        rfc_receptor="VECJ880326XXX",
        curp="VECJ880326HDFLNS09",
        nss="",
    )
    resultado = await b10.consultar(db, eid, _p())
    por_titulo = {c.titulo: c for c in resultado.columnas}
    assert por_titulo["CURP"].sensible is True
    assert por_titulo["NSS"].sensible is True


async def test_sin_comprobantes_devuelve_aviso(db: AsyncSession) -> None:
    eid = await _empresa(db)
    resultado = await b10.consultar(db, eid, _p(fecha_desde=date(2026, 1, 1), fecha_hasta=date(2026, 1, 31)))
    assert resultado.filas == [] and resultado.aviso is not None
