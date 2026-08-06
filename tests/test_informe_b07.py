"""B-07 · Cartera de préstamos y descuentos recurrentes.

El CFDI no contiene el saldo del préstamo, solo el descuento del periodo, así que el valor
de este informe está en el control de continuidad (B-07.R1): detectar un descuento que se
detuvo sin liquidarse.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.informes import b07_prestamos as b07, excel
from app.informes.base import ContextoInforme
from app.models.enums import EstatusCfdi
from tests import factories
from tests.helpers_nomina import insertar_nomina


def _p(**kw: object) -> b07.Parametros:
    base = {"fecha_desde": date(2026, 1, 1), "fecha_hasta": date(2026, 3, 31)}
    base.update(kw)
    return b07.Parametros(**base)  # type: ignore[arg-type]


async def _empresa(db: AsyncSession) -> int:
    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    return empresa.empresa_id


async def _quincenas(db: AsyncSession, eid: int, importes: list[str | None], *, clave: str = "016",
                     tipo: str = "009", rfc: str = "XAXX010101000") -> None:
    """Una quincena por importe; `None` significa que ese periodo NO trae el descuento."""
    cortes = [(date(2026, 1, 15)), (date(2026, 1, 31)), (date(2026, 2, 15)), (date(2026, 2, 28)),
              (date(2026, 3, 15)), (date(2026, 3, 31))]
    for indice, importe in enumerate(importes):
        deducciones = [("002", "045", "ISR", "100.00")]
        if importe is not None:
            deducciones.append((tipo, clave, "Prestamo infonavit", importe))
        await insertar_nomina(db, empresa_id=eid, uuid=f"1111111{indice}-1111-1111-1111-111111111111",
                              rfc_receptor=rfc, fecha_pago=cortes[indice], fecha_final_pago=cortes[indice],
                              periodicidad="04", deducciones=deducciones)


def _fila(resultado: object, titulo: str, indice: int = 0) -> object:
    titulos = [c.titulo for c in resultado.columnas]  # type: ignore[attr-defined]
    return resultado.filas[indice][titulos.index(titulo)]  # type: ignore[attr-defined]


async def test_una_fila_por_empleado_tipo_y_clave(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", "500.00"])

    resultado = await b07.consultar(db, eid, _p())
    # El ISR también es una deducción recurrente y aparece como su propia fila.
    claves = {(f[0], f[1]) for f in [( _fila(resultado, "Tipo deducción", i), _fila(resultado, "Clave", i))
                                     for i in range(len(resultado.filas))]}
    assert len(resultado.filas) == 2, claves


async def test_agregados_de_la_serie(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", "600.00"])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert len(resultado.filas) == 1
    assert _fila(resultado, "Núm. de descuentos") == 3
    assert _fila(resultado, "Total descontado") == Decimal("1600.00")
    assert _fila(resultado, "Primer descuento") == date(2026, 1, 15)
    assert _fila(resultado, "Último descuento") == date(2026, 2, 15)


async def test_descuento_modal_no_es_el_promedio(db: AsyncSession) -> None:
    """El modal identifica la amortización pactada; un solo ajuste desvía el promedio pero
    no el modal."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", "500.00", "2000.00"])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Descuento modal") == Decimal("500.00")
    assert _fila(resultado, "Descuento promedio") != Decimal("500.00")


async def test_continuidad_continuo(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", "500.00", "500.00", "500.00", "500.00"])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Continuidad") == "CONTINUO"
    assert not [b for b in resultado.banderas if b.clave == "DESCUENTO_INTERRUMPIDO"]


async def test_continuidad_interrumpido_es_el_hallazgo_del_informe(db: AsyncSession) -> None:
    """B-07.R1: un préstamo cuyo descuento se detuvo a media serie y luego volvió."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", None, "500.00", "500.00"])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Continuidad") == "INTERRUMPIDO"
    banderas = [b for b in resultado.banderas if b.clave == "DESCUENTO_INTERRUMPIDO"]
    assert banderas
    # El mensaje lista los huecos, porque es la información accionable.
    assert "2026-01 Q2" in banderas[0].mensaje


async def test_continuidad_concluido(db: AsyncSession) -> None:
    """El descuento terminó antes del final del rango: probablemente se liquidó."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", None, None, None, None])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Continuidad") == "CONCLUIDO"


async def test_amortizacion_modificada(db: AsyncSession) -> None:
    """B-07.R3: lo esperado en un crédito Infonavit al actualizarse el factor."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00", "500.00", "620.00", "620.00", "620.00"])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    banderas = [b for b in resultado.banderas if b.clave == "AMORTIZACION_MODIFICADA"]
    assert banderas
    assert "500" in banderas[0].mensaje and "620" in banderas[0].mensaje


async def test_no_hay_columnas_de_saldo(db: AsyncSession) -> None:
    """B-07.R2: sin el monto original no se estima el saldo, y no se infiere del descuento."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00"])
    resultado = await b07.consultar(db, eid, _p())
    titulos = [c.titulo for c in resultado.columnas]
    assert not any("saldo" in t.lower() or "liquidaci" in t.lower() for t in titulos)


async def test_descripcion_sat_del_tipo_de_deduccion(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00"])
    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Descripción SAT") is not None


async def test_sin_descuentos_devuelve_aviso(db: AsyncSession) -> None:
    eid = await _empresa(db)
    resultado = await b07.consultar(db, eid, _p())
    assert resultado.filas == [] and resultado.aviso is not None


async def test_no_declara_el_parametro_de_enmascaramiento(db: AsyncSession) -> None:
    """**Hallazgo Important de la revisión final.** B-07 declaraba `enmascarar_datos_personales`
    con la descripción "Sin efecto en este informe" y no tiene ninguna columna sensible. No era
    cosmético: `app.api.v1.informes.generar_endpoint` gatea por el **parámetro**, no por si hay algo
    que enmascarar, así que desmarcar esa casilla devolvía `403` a un usuario `CONSULTA` y, a un
    `OPERADOR`, escribía en bitácora un registro de divulgación de datos personales que nunca
    ocurrió. Un asiento de auditoría falso es peor que una casilla inútil.

    Se comprueba también que ninguna columna sea sensible: es la premisa de la que depende que
    quitar el parámetro sea seguro."""
    assert "enmascarar_datos_personales" not in b07.Parametros.model_fields
    assert "enmascarar_datos_personales" not in b07.Parametros.model_json_schema()["properties"]

    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00"])
    resultado = await b07.consultar(db, eid, _p())
    assert [c.titulo for c in resultado.columnas if c.sensible] == []


async def test_el_motor_escribe_el_libro_sin_el_parametro(db: AsyncSession) -> None:
    """El motor lee `ctx.parametros.get("enmascarar_datos_personales", True)` —falla cerrado—, así
    que su ausencia no puede romper la escritura del libro ni cambiar nada: no hay columna sensible
    que enmascarar. Comprobado, no supuesto."""
    eid = await _empresa(db)
    await _quincenas(db, eid, ["500.00", "500.00"])
    p = _p(tipos_deduccion=["009"])
    resultado = await b07.consultar(db, eid, p)
    assert resultado.filas

    ctx = ContextoInforme(
        clave=b07.CLAVE,
        nombre=b07.NOMBRE,
        usuario="consulta@test.mx",
        generado_en=datetime(2026, 8, 6, 12, 0),
        parametros=p.model_dump(mode="json"),
        etl_version=1,
    )
    assert "enmascarar_datos_personales" not in ctx.parametros
    libro = openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))
    assert set(libro.sheetnames) == {"Datos", "Parámetros", "Banderas", "Diccionario"}
    # El RFC del empleado sale completo, como debe: no es una columna sensible en este informe.
    assert libro["Datos"].cell(row=2, column=1).value == "XAXX010101000"


async def test_un_cfdi_cancelado_no_tapa_un_hueco_sin_avisar(db: AsyncSession) -> None:
    """**Hallazgo Important de la revisión final.** Este informe no llamaba a
    `universo_nomina.banderas_de_estatus` y no tiene columna de estatus, así que un CFDI que el SAT
    ya no reconoce podía **tapar** un hueco de la serie: la continuidad salía `CONTINUO`,
    `DESCUENTO_INTERRUMPIDO` no se disparaba y el préstamo se veía al día. El §11 del diseño exige
    que todo comprobante incluido que no sea vigente lleve bandera."""
    eid = await _empresa(db)
    cortes = [date(2026, 1, 15), date(2026, 1, 31), date(2026, 2, 15)]
    for indice, (corte, estatus) in enumerate(zip(cortes, (EstatusCfdi.VIGENTE, EstatusCfdi.CANCELADO, EstatusCfdi.VIGENTE))):
        await insertar_nomina(db, empresa_id=eid, uuid=f"dddddddd-dddd-dddd-dddd-dddddddddd0{indice}",
                              fecha_pago=corte, fecha_final_pago=corte, periodicidad="04",
                              estatus=estatus,
                              deducciones=[("009", "016", "Prestamo infonavit", "500.00")])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"], incluir_cancelados=True))
    assert _fila(resultado, "Núm. de descuentos") == 3  # el cancelado tapa el hueco
    bandera = next(b for b in resultado.banderas if b.clave == "COMPROBANTE_CANCELADO")
    assert bandera.ambito == "uuid:dddddddd-dddd-dddd-dddd-dddddddddd01"
    assert bandera.severidad == "alta"


async def test_no_verificado_lleva_bandera_media(db: AsyncSession) -> None:
    """`no_verificado` es el estado normal de un rango recién descargado (la verificación contra el
    SAT es asíncrona por diseño), así que es la rama que de verdad se ve en producción."""
    eid = await _empresa(db)
    await insertar_nomina(db, empresa_id=eid, uuid="dddddddd-dddd-dddd-dddd-ddddddddddd9",
                          fecha_pago=date(2026, 1, 15), fecha_final_pago=date(2026, 1, 15),
                          periodicidad="04", estatus=EstatusCfdi.NO_VERIFICADO,
                          deducciones=[("009", "016", "Prestamo infonavit", "500.00")])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    bandera = next(b for b in resultado.banderas if b.clave == "ESTATUS_NO_VERIFICADO")
    assert bandera.severidad == "media"


async def test_nombre_empleado_es_del_trabajador(db: AsyncSession) -> None:
    """La columna sale de `comprobante_detalle.nombre_receptor`. Hasta la revisión final el helper
    de pruebas nunca insertaba esa fila, así que `detalle` era `None` y la columna siempre salía
    vacía: nada distinguía el campo correcto del equivocado (ver B-01/B-02)."""
    eid = await _empresa(db)
    await insertar_nomina(db, empresa_id=eid, uuid="dddddddd-dddd-dddd-dddd-dddddddddda1",
                          fecha_pago=date(2026, 1, 15), fecha_final_pago=date(2026, 1, 15),
                          periodicidad="04", nombre_receptor="JUANA INVENTADA DE PRUEBA",
                          deducciones=[("009", "016", "Prestamo infonavit", "500.00")])

    resultado = await b07.consultar(db, eid, _p(tipos_deduccion=["009"]))
    assert _fila(resultado, "Nombre empleado") == "JUANA INVENTADA DE PRUEBA"
