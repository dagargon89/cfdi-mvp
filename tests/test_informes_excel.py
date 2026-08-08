"""Motor de informes: el libro común de cuatro hojas (spec §10) y el enmascaramiento.

Las aserciones se hacen sobre valores de celda, nunca sobre bytes del archivo (spec §13).
"""

from __future__ import annotations

import io
from dataclasses import replace
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook

from app.informes import excel
from app.informes.base import Bandera, Columna, ContextoInforme, EntradaDiccionario, ResultadoInforme


def _resultado_demo() -> ResultadoInforme:
    return ResultadoInforme(
        columnas=[
            Columna(titulo="UUID", tipo="texto"),
            Columna(titulo="Fecha pago", tipo="fecha"),
            Columna(titulo="P¦001¦001¦Sueldo", tipo="monto"),
        ],
        filas=[["11111111-1111-1111-1111-111111111111", datetime(2026, 6, 30).date(), Decimal("8759.700000")]],
        banderas=[Bandera(clave="NETO_NEGATIVO", severidad="alta", ambito="uuid:1111", mensaje="El total es negativo.")],
        diccionario=[
            EntradaDiccionario(
                etiqueta="P¦001¦001¦Sueldo",
                naturaleza="P",
                tipo="001",
                descripcion_sat="Sueldos, Salarios Rayas y Jornales",
                clave_patron="001",
                concepto_canonico="Sueldo",
                descripciones_alternas=["Sueldos"],
                num_comprobantes=8,
                importe_total=Decimal("70077.60"),
            )
        ],
    )


def _contexto() -> ContextoInforme:
    return ContextoInforme(
        clave="B-02",
        nombre="Nómina agrupada por conceptos del patrón",
        usuario="dgarcia@planjuarez.org",
        generado_en=datetime(2026, 8, 5, 11, 30, 0),
        parametros={"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31", "enmascarar_datos_personales": True},
        etl_version=1,
    )


def test_libro_tiene_las_cuatro_hojas() -> None:
    wb = load_workbook(io.BytesIO(excel.escribir_libro(_resultado_demo(), _contexto())))
    assert wb.sheetnames == ["Datos", "Parámetros", "Banderas", "Diccionario"]


def test_hoja_datos_lleva_encabezado_y_valores() -> None:
    wb = load_workbook(io.BytesIO(excel.escribir_libro(_resultado_demo(), _contexto())))
    ws = wb["Datos"]

    assert [c.value for c in ws[1]] == ["UUID", "Fecha pago", "P¦001¦001¦Sueldo"]
    fila = [c.value for c in ws[2]]
    assert fila[0] == "11111111-1111-1111-1111-111111111111"
    # Redondeo único al presentar (R-T4): 6 decimales en BD → 2 en la celda.
    assert float(fila[2]) == 8759.70
    # Encabezado congelado para que la tabla sea navegable.
    assert ws.freeze_panes == "A2"


def test_hoja_parametros_permite_reproducir_la_corrida() -> None:
    """Sin esta hoja un Excel circulando por correo no se puede auditar (spec §10)."""
    wb = load_workbook(io.BytesIO(excel.escribir_libro(_resultado_demo(), _contexto())))
    ws = wb["Parámetros"]
    contenido = {ws.cell(row=f, column=1).value: ws.cell(row=f, column=2).value for f in range(1, ws.max_row + 1)}

    assert contenido["Informe"] == "B-02 · Nómina agrupada por conceptos del patrón"
    assert contenido["Generado por"] == "dgarcia@planjuarez.org"
    assert contenido["Versión del ETL"] == 1
    assert contenido["Filas"] == 1
    assert contenido["fecha_desde"] == "2026-06-01"
    assert contenido["enmascarar_datos_personales"] == "True"


def test_hoja_banderas_y_diccionario() -> None:
    wb = load_workbook(io.BytesIO(excel.escribir_libro(_resultado_demo(), _contexto())))

    banderas = wb["Banderas"]
    assert [c.value for c in banderas[1]] == ["Bandera", "Severidad", "Ámbito", "Mensaje"]
    assert [c.value for c in banderas[2]] == ["NETO_NEGATIVO", "alta", "uuid:1111", "El total es negativo."]

    diccionario = wb["Diccionario"]
    fila = [c.value for c in diccionario[2]]
    assert fila[0] == "P¦001¦001¦Sueldo"
    assert fila[1] == "P"
    assert fila[3] == "Sueldos, Salarios Rayas y Jornales"
    assert fila[6] == "Sueldos"  # descripciones alternas, unidas por '; '


def test_el_importe_del_diccionario_lleva_formato_de_importe() -> None:
    """**Hallazgo de la ronda 2 al auditar los nueve libros reales.** «Importe del periodo» es la
    única celda numérica del libro fuera de `Datos` que lleva un importe, y salía sin
    `number_format`: el valor iba cuantizado a dos decimales pero la celda lo mostraba en `General`
    —`108757.8` en vez de `108,757.80`—, así que la misma cifra se veía distinta en `Datos` y en
    `Diccionario` del mismo libro."""
    wb = load_workbook(io.BytesIO(excel.escribir_libro(_resultado_demo(), _contexto())))
    celda = wb["Diccionario"]["I2"]

    assert wb["Diccionario"]["I1"].value == "Importe del periodo"
    assert celda.number_format == "#,##0.00"
    assert Decimal(str(celda.value)) == Decimal("70077.6")
    # `Núm. comprobantes` sí se queda en `General` a propósito: es un `int` y se muestra exacto,
    # así que no hay divergencia entre lo guardado y lo mostrado que arreglar.
    assert wb["Diccionario"]["H2"].value == 8


def test_informe_vacio_produce_libro_con_aviso_no_una_excepcion() -> None:
    """spec §9: sin filas no es un error."""
    vacio = ResultadoInforme(columnas=[Columna(titulo="UUID", tipo="texto")], filas=[], aviso="Sin comprobantes en el rango solicitado.")
    wb = load_workbook(io.BytesIO(excel.escribir_libro(vacio, _contexto())))

    assert "Datos" in wb.sheetnames
    parametros = wb["Parámetros"]
    contenido = {parametros.cell(row=f, column=1).value: parametros.cell(row=f, column=2).value for f in range(1, parametros.max_row + 1)}
    assert contenido["Filas"] == 0
    assert contenido["Aviso"] == "Sin comprobantes en el rango solicitado."


def test_enmascarar_conserva_los_ultimos_cuatro() -> None:
    assert excel.enmascarar("XXXX800101HCHXXX01") == "****XX01"
    assert excel.enmascarar("12345678901") == "****8901"
    assert excel.enmascarar(None) is None
    assert excel.enmascarar("123") == "****"  # demasiado corto para conservar 4


def test_columna_sensible_se_enmascara_solo_si_el_parametro_esta_activo() -> None:
    """El motor, no el informe, decide si enmascara (ronda de corrección 1): un informe solo
    marca `sensible=True` y entrega el dato en claro."""
    resultado = ResultadoInforme(
        columnas=[
            Columna(titulo="CURP", tipo="texto", sensible=True),
            Columna(titulo="Nombre", tipo="texto"),
        ],
        filas=[["XXXX800101HCHXXX01", "Juan Pérez"]],
    )
    ctx_enmascarado = replace(_contexto(), parametros={"enmascarar_datos_personales": True})
    ctx_en_claro = replace(_contexto(), parametros={"enmascarar_datos_personales": False})

    wb_enmascarado = load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx_enmascarado)))
    fila = [c.value for c in wb_enmascarado["Datos"][2]]
    assert fila[0] == "****XX01"
    assert fila[1] == "Juan Pérez"  # la columna no sensible nunca se toca

    wb_en_claro = load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx_en_claro)))
    fila = [c.value for c in wb_en_claro["Datos"][2]]
    assert fila[0] == "XXXX800101HCHXXX01"
    assert fila[1] == "Juan Pérez"


def test_columna_sensible_se_enmascara_cuando_la_clave_no_viene_en_los_parametros() -> None:
    """**Falla cerrado.** `enmascarar_datos_personales` tiene default `True` en el informe y en
    su JSON Schema, así que un contexto que no traiga la clave —un llamador que no pase por el
    endpoint HTTP, que sí incluye los defaults al hacer `model_dump()`— debe enmascarar igual.
    Con `.get()` sin default, `None` es falsy y el libro salía con el CURP en claro."""
    resultado = ResultadoInforme(
        columnas=[Columna(titulo="CURP", tipo="texto", sensible=True)],
        filas=[["XXXX800101HCHXXX01"]],
    )
    ctx = replace(_contexto(), parametros={"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"})

    wb = load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))
    assert wb["Datos"][2][0].value == "****XX01"


def test_columna_decimal_se_redondea_a_tres_decimales_no_a_dos() -> None:
    """Una columna `decimal` **no** es un `monto`: se redondea a los 3 decimales que su formato
    muestra, no a los 2 de un importe. Esa distinción es la que protegía la versión anterior de
    esta prueba y sigue en pie.

    **Lo que cambió, y por qué la prueba anterior estaba equivocada** (ronda de corrección 2 de la
    tarea 10): aquella afirmaba que un `decimal` "conserva su precisión", es decir que la celda se
    quedaba con `12.345678` mientras su formato mostraba `12.346`. Eso no era el cumplimiento de
    R-T4 sino su incumplimiento: la regla dice que el redondeo ocurre una sola vez **en este
    módulo**, no que este módulo no redondee, y un valor guardado distinto del mostrado es
    precisamente la discrepancia que aparece al exportar a CSV o al sumar la columna en otra
    herramienta. La prueba fue la única de los nueve informes que falló al derivar la escala del
    formato, y su fallo fue el hallazgo, no un ajuste."""
    resultado = ResultadoInforme(columnas=[Columna(titulo="Tasa", tipo="decimal")], filas=[[Decimal("12.345678")]])
    wb = load_workbook(io.BytesIO(excel.escribir_libro(resultado, _contexto())))
    celda = wb["Datos"][2][0]
    assert Decimal(str(celda.value)) == Decimal("12.346")
    assert Decimal(str(celda.value)) != Decimal("12.35"), "un `decimal` no se redondea como un importe"
    assert celda.number_format == "#,##0.000"


def _decimales_declarados(formato: str) -> int:
    """Cuántos decimales muestra un formato de Excel, contados **desde el formato de la celda
    leída del archivo**, no desde una constante del código: si el motor y su prueba leyeran la
    misma constante, la prueba no comprobaría que el formato y el valor concuerdan."""
    _entero, _punto, decimales = formato.partition(".")
    return len(decimales)


def _sin_precision_oculta(hoja: object) -> list[str]:
    """Las celdas numéricas cuyo valor **guardado** lleva más decimales de los que su formato
    **muestra**. Una lista vacía es el invariante que este módulo tiene que cumplir."""
    problemas: list[str] = []
    for fila in hoja.iter_rows(min_row=2):  # type: ignore[attr-defined]
        for celda in fila:
            if celda.value is None or isinstance(celda.value, (str, bool)):
                continue
            if not isinstance(celda.value, (int, float, Decimal)):
                continue  # fechas y horas
            guardado = Decimal(str(celda.value))
            declarados = _decimales_declarados(celda.number_format)
            if -guardado.as_tuple().exponent > declarados:  # type: ignore[operator]
                problemas.append(f"{celda.coordinate}: guarda {guardado} y muestra {declarados} decimales")
    return problemas


def test_lo_guardado_en_la_celda_coincide_con_lo_que_el_formato_muestra() -> None:
    """**El invariante del redondeo, comprobado sobre el resultado y no sobre el mecanismo.**

    Se recorre el libro **ya escrito** y se compara, celda por celda, el valor guardado contra los
    decimales que declara su propio `number_format`. No se compara contra `_ESCALA_POR_TIPO` a
    propósito: mirar la misma constante que usa el motor probaría que el motor se llama a sí mismo,
    no que lo guardado y lo mostrado concuerdan. Es la misma lección que la fuga de datos
    personales de B-10, donde probar el mecanismo dejó pasar el resultado.

    Los valores de entrada son divisiones no terminantes, que es de donde salen en la realidad
    (`importe / dias`, `parte * 100 / total`)."""
    tercios = Decimal(1) / Decimal(3)
    resultado = ResultadoInforme(
        columnas=[
            Columna(titulo="Importe", tipo="monto"),
            Columna(titulo="Días", tipo="decimal"),
            Columna(titulo="% del total", tipo="decimal"),
            Columna(titulo="Núm. de CFDI", tipo="entero"),
            Columna(titulo="Fecha", tipo="fecha"),
            Columna(titulo="Texto", tipo="texto"),
        ],
        filas=[
            [Decimal("8759.7") * tercios, Decimal("15") * tercios, Decimal(100) * tercios, 8, datetime(2026, 6, 30).date(), "x"],
            [Decimal("12.885"), Decimal("12.8885"), Decimal("0.0005"), 0, datetime(2026, 7, 15).date(), "y"],
        ],
    )

    wb = load_workbook(io.BytesIO(excel.escribir_libro(resultado, _contexto())))
    assert _sin_precision_oculta(wb["Datos"]) == []
    # Y el diccionario, que también escribe un importe por esta misma vía.
    assert _sin_precision_oculta(wb["Diccionario"]) == []

    fila = [c.value for c in wb["Datos"][2]]
    assert Decimal(str(fila[0])) == Decimal("2919.90")
    assert Decimal(str(fila[1])) == Decimal("5.000")
    assert Decimal(str(fila[2])) == Decimal("33.333")
    assert fila[3] == 8
    # Las columnas no numéricas siguen intactas: no se cuantiza lo que no es un número.
    assert fila[4] == datetime(2026, 6, 30)
    assert fila[5] == "x"


def test_el_redondeo_es_medio_arriba_en_los_tres_tipos_numericos() -> None:
    """`ROUND_HALF_UP` en todos, no solo en `monto`: el default de Python es `ROUND_HALF_EVEN`,
    que con `12.885` daría `12.88` en vez de `12.89`. Cada valor está justo en el medio del
    dígito que se pierde, que es el único sitio donde los dos modos difieren."""
    resultado = ResultadoInforme(
        columnas=[
            Columna(titulo="Importe", tipo="monto"),
            Columna(titulo="Días", tipo="decimal"),
            Columna(titulo="Conteo", tipo="entero"),
        ],
        filas=[[Decimal("12.885"), Decimal("12.8885"), Decimal("12.5")]],
    )

    wb = load_workbook(io.BytesIO(excel.escribir_libro(resultado, _contexto())))
    fila = [Decimal(str(c.value)) for c in wb["Datos"][2]]
    assert fila == [Decimal("12.89"), Decimal("12.889"), Decimal("13")]


def test_la_escala_de_redondeo_se_deriva_del_formato_de_presentacion() -> None:
    """La escala no se escribe dos veces: sale de `_FORMATO`. Si alguien cambia el formato de un
    tipo para mostrar más decimales, la celda guarda más decimales **sola**, y no hay forma de
    que los dos números se separen."""
    for tipo, escala in excel._ESCALA_POR_TIPO.items():
        declarados = _decimales_declarados(excel._FORMATO[tipo])
        assert -escala.as_tuple().exponent == declarados, f"{tipo}: escala {escala} contra formato declarado"  # type: ignore[operator]
    assert excel._ESCALA_POR_TIPO == {"monto": Decimal("0.01"), "decimal": Decimal("0.001"), "entero": Decimal("1")}
    # Y las columnas no numéricas no tienen escala: nada que cuantizar.
    assert set(excel._ESCALA_POR_TIPO) == {"monto", "decimal", "entero"}
