"""Endpoints de informes (spec §7.2) y el pre-vuelo del ETL (disparador 3, spec §6.3)."""

from __future__ import annotations

import io
import os
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession, async_sessionmaker

from app.core.config import get_settings
from app.models.cfdi_detalle import ComprobanteDetalle
from app.models.enums import RolEmpresa, RolGlobal
from app.worker import tasks as worker_tasks
from tests import factories, fixtures_cfdi


@pytest.fixture(autouse=True)
def _sesion_de_prueba_en_worker(engine: AsyncEngine, monkeypatch: pytest.MonkeyPatch) -> None:
    """`worker_tasks.SessionLocal` apunta a `DATABASE_URL` (placeholder en pruebas): la
    tarea de generación de informes se ejerce aquí llamando su lógica async directo (sin
    `asyncio.run`/Celery), así que basta apuntarlo al mismo engine que usa el fixture `db`
    (testcontainer real) — mismo patrón que `test_worker_comprobantes.py`."""
    monkeypatch.setattr(worker_tasks, "SessionLocal", async_sessionmaker(engine, expire_on_commit=False))


async def test_catalogo_expone_b02_con_su_json_schema(client, db: AsyncSession) -> None:  # type: ignore[no-untyped-def]
    await factories.crear_usuario(db, uid="op", correo="op@test.mx", rol_global=RolGlobal.OPERADOR)

    r = await client.get("/v1/informes", headers={"Authorization": "Bearer op"})
    assert r.status_code == 200, r.text
    claves = {i["clave"]: i for i in r.json()}
    assert "B-02" in claves
    b02 = claves["B-02"]
    assert b02["grupo"] == "B"
    # El frontend arma el formulario desde aquí: los parámetros deben venir descritos.
    propiedades = b02["parametros"]["properties"]
    assert "fecha_desde" in propiedades
    assert propiedades["enmascarar_datos_personales"]["default"] is True


async def test_generar_informe_encola_con_los_parametros(client, db: AsyncSession, monkeypatch) -> None:  # type: ignore[no-untyped-def]
    from app.worker import tasks

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    usuario = await factories.crear_usuario(db, uid="con", correo="con@test.mx", rol_global=RolGlobal.CONSULTA)
    await factories.asignar_permiso(db, usuario, empresa, RolEmpresa.CONSULTA)

    encoladas: list[tuple[int, str, dict, str]] = []

    class _Tarea:
        id = "tarea-informe"

    monkeypatch.setattr(
        tasks.generar_informe,
        "delay",
        lambda empresa_id, clave, parametros, actor: (encoladas.append((empresa_id, clave, parametros, actor)), _Tarea())[1],
    )

    r = await client.post(
        f"/v1/empresas/{empresa.empresa_id}/informes/B-02",
        json={"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"},
        headers={"Authorization": "Bearer con"},
    )
    assert r.status_code == 202, r.text
    assert r.json()["tarea_id"] == "tarea-informe"
    assert encoladas[0][1] == "B-02"
    assert encoladas[0][2]["fecha_desde"] == "2026-06-01"
    assert encoladas[0][3] == "con@test.mx"


async def test_clave_desconocida_da_404(client, db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    usuario = await factories.crear_usuario(db, uid="con", correo="con@test.mx", rol_global=RolGlobal.CONSULTA)
    await factories.asignar_permiso(db, usuario, empresa, RolEmpresa.CONSULTA)

    r = await client.post(
        f"/v1/empresas/{empresa.empresa_id}/informes/Z-99",
        json={"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"},
        headers={"Authorization": "Bearer con"},
    )
    assert r.status_code == 404


async def test_parametros_invalidos_dan_422(client, db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    usuario = await factories.crear_usuario(db, uid="con", correo="con@test.mx", rol_global=RolGlobal.CONSULTA)
    await factories.asignar_permiso(db, usuario, empresa, RolEmpresa.CONSULTA)

    r = await client.post(
        f"/v1/empresas/{empresa.empresa_id}/informes/B-02",
        json={"fecha_desde": "no-es-fecha"},
        headers={"Authorization": "Bearer con"},
    )
    assert r.status_code == 422


async def test_sin_enmascarar_exige_operador_y_deja_bitacora(client, db: AsyncSession, monkeypatch) -> None:  # type: ignore[no-untyped-def]
    """spec §8."""
    from sqlalchemy import select

    from app.models.bitacora import Bitacora
    from app.worker import tasks

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    consulta = await factories.crear_usuario(db, uid="con", correo="con@test.mx", rol_global=RolGlobal.CONSULTA)
    await factories.asignar_permiso(db, consulta, empresa, RolEmpresa.CONSULTA)
    operador = await factories.crear_usuario(db, uid="op", correo="op@test.mx", rol_global=RolGlobal.OPERADOR)
    await factories.asignar_permiso(db, operador, empresa, RolEmpresa.OPERADOR)

    class _Tarea:
        id = "t"

    monkeypatch.setattr(tasks.generar_informe, "delay", lambda *a, **k: _Tarea())

    cuerpo = {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31", "enmascarar_datos_personales": False}

    r = await client.post(f"/v1/empresas/{empresa.empresa_id}/informes/B-02", json=cuerpo, headers={"Authorization": "Bearer con"})
    assert r.status_code == 403

    r = await client.post(f"/v1/empresas/{empresa.empresa_id}/informes/B-02", json=cuerpo, headers={"Authorization": "Bearer op"})
    assert r.status_code == 202

    registros = list((await db.execute(select(Bitacora).where(Bitacora.accion == "generar_informe"))).scalars().all())
    assert registros, "la generación sin enmascarar debe quedar en bitácora"
    assert registros[-1].detalle["enmascarar_datos_personales"] is False


async def test_b07_no_gatea_ni_deja_bitacora_falsa(client, db: AsyncSession, monkeypatch) -> None:  # type: ignore[no-untyped-def]
    """**Hallazgo Important de la revisión final.** El endpoint gatea por el **parámetro**, no por
    si el informe tiene algo que enmascarar. B-07 declaraba `enmascarar_datos_personales` con la
    nota "Sin efecto en este informe" y ninguna columna sensible, así que desmarcar esa casilla
    devolvía `403` a un usuario `CONSULTA` y, a un `OPERADOR`, escribía en bitácora un registro de
    divulgación de datos personales **que nunca ocurrió**. Un asiento de auditoría falso es peor que
    una casilla inútil, así que el parámetro se quitó de B-07.

    Se comprueba lo que importa de punta a punta: (1) el catálogo ya no lo publica, así que el
    frontend no puede pintar la casilla; (2) un `CONSULTA` genera el informe con `202`; (3) no queda
    ningún registro de divulgación en bitácora."""
    from sqlalchemy import select

    from app.models.bitacora import Bitacora
    from app.worker import tasks

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    consulta = await factories.crear_usuario(db, uid="con", correo="con@test.mx", rol_global=RolGlobal.CONSULTA)
    await factories.asignar_permiso(db, consulta, empresa, RolEmpresa.CONSULTA)

    class _Tarea:
        id = "t"

    monkeypatch.setattr(tasks.generar_informe, "delay", lambda *a, **k: _Tarea())

    r = await client.get("/v1/informes", headers={"Authorization": "Bearer con"})
    b07 = {i["clave"]: i for i in r.json()}["B-07"]
    assert "enmascarar_datos_personales" not in b07["parametros"]["properties"]

    r = await client.post(
        f"/v1/empresas/{empresa.empresa_id}/informes/B-07",
        json={"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"},
        headers={"Authorization": "Bearer con"},
    )
    assert r.status_code == 202, r.text

    registros = list((await db.execute(select(Bitacora).where(Bitacora.accion == "generar_informe"))).scalars().all())
    assert registros == [], "B-07 no divulga datos personales: no puede dejar un asiento que diga que sí"


async def test_tarea_genera_libro_y_normaliza_lo_pendiente(db: AsyncSession) -> None:
    """Pre-vuelo: el comprobante existe en `comprobantes` pero nunca se normalizó. La
    tarea lo normaliza antes de consultar, así que el informe NO sale vacío."""
    from app.worker.tasks import _generar_informe_async

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    carpeta = os.path.join(get_settings().storage_root, str(empresa.empresa_id), "comprobantes")
    os.makedirs(carpeta, exist_ok=True)
    with open(os.path.join(carpeta, "nomina.xml"), "wb") as f:
        f.write(fixtures_cfdi.cfdi_nomina())
    await factories.crear_comprobante(
        db,
        empresa_id=empresa.empresa_id,
        uuid="77777777-7777-7777-7777-777777777777",
        # El universo del grupo B son los `N` EMITIDOS por la empresa (§11 del diseño), así que
        # el RFC emisor tiene que ser el de la empresa, como en el XML del fixture.
        rfc_emisor="CHL960913IX9",
        tipo_comprobante="N",
        xml_path=os.path.join(str(empresa.empresa_id), "comprobantes", "nomina.xml"),
    )

    resultado = await _generar_informe_async(
        empresa.empresa_id,
        "B-02",
        {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"},
        "op@test.mx",
    )

    assert resultado["filas"] == 1
    ruta = os.path.join(get_settings().storage_root, resultado["ruta"])
    assert os.path.isfile(ruta)

    with open(ruta, "rb") as f:
        wb = load_workbook(io.BytesIO(f.read()))
    assert wb.sheetnames == ["Datos", "Parámetros", "Banderas", "Diccionario"]
    datos = wb["Datos"]
    encabezados = [c.value for c in datos[1]]
    assert "UUID" in encabezados
    assert any("Sueldo" in (t or "") for t in encabezados)


async def _nomina_en_disco(db: AsyncSession, empresa_id: int, uuid: str, xml: bytes, tipo: str = "N") -> int:
    """Deja un XML en el resguardo y registra su comprobante sin normalizar, para que lo
    normalice el pre-vuelo de la tarea."""
    carpeta = os.path.join(get_settings().storage_root, str(empresa_id), "comprobantes")
    os.makedirs(carpeta, exist_ok=True)
    nombre = f"{uuid}.xml"
    with open(os.path.join(carpeta, nombre), "wb") as f:
        f.write(xml)
    comprobante = await factories.crear_comprobante(
        db,
        empresa_id=empresa_id,
        uuid=uuid,
        rfc_emisor="CHL960913IX9",
        tipo_comprobante=tipo,
        total=Decimal("8168.60"),
        xml_path=os.path.join(str(empresa_id), "comprobantes", nombre),
    )
    return comprobante.comprobante_id


def _celda(ws, titulo: str, fila: int = 2):  # type: ignore[no-untyped-def]
    encabezados = [c.value for c in ws[1]]
    assert titulo in encabezados, f"falta la columna {titulo!r}; hay {encabezados}"
    return ws.cell(row=fila, column=encabezados.index(titulo) + 1).value


def _libro(resultado: dict) -> object:  # type: ignore[type-arg]
    with open(os.path.join(get_settings().storage_root, resultado["ruta"]), "rb") as f:
        return load_workbook(io.BytesIO(f.read()))


async def test_el_libro_se_enmascara_aunque_el_llamador_omita_el_parametro(db: AsyncSession) -> None:
    """**Enmascaramiento que falla CERRADO.** La tarea recibía el dict crudo del cliente y lo
    pasaba tal cual al `ContextoInforme`, mientras `escribir_libro` decidía con
    `.get("enmascarar_datos_personales")` sin default: si la clave no venía, `None` es falsy y
    el libro salía con CURP y NSS **en claro**, aunque el default declarado y publicado en el
    JSON Schema del informe sea `True`.

    No había fuga por HTTP (el endpoint hace `model_dump()`, que incluye defaults) ni por la UI,
    pero esta ruta ya estaba ejercitada en el repo y cualquier llamador no-HTTP futuro heredaba
    el fallo. Aquí se llama a la tarea **sin** la clave, como lo haría ese llamador."""
    from app.worker.tasks import _generar_informe_async

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    await _nomina_en_disco(db, empresa.empresa_id, "77777777-7777-7777-7777-777777777777", fixtures_cfdi.cfdi_nomina())

    resultado = await _generar_informe_async(
        empresa.empresa_id,
        "B-02",
        {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"},  # sin `enmascarar_datos_personales`
        "op@test.mx",
    )
    assert resultado["filas"] == 1

    wb = _libro(resultado)
    datos = wb["Datos"]  # type: ignore[index]
    assert _celda(datos, "CURP") == "****XX01", "el CURP del fixture es XXXX800101HCHXXX01"
    assert _celda(datos, "NSS") == "****8901"

    # Segunda mitad del hallazgo: la hoja `Parámetros` debe registrar los filtros EFECTIVOS,
    # no solo lo que el cliente escribió, o el Excel no es reproducible (§10 del diseño).
    parametros = wb["Parámetros"]  # type: ignore[index]
    contenido = {parametros.cell(row=f, column=1).value: parametros.cell(row=f, column=2).value for f in range(1, parametros.max_row + 1)}
    assert contenido["enmascarar_datos_personales"] == "True"
    assert contenido["tipo_nomina"] == "AMBOS"
    assert contenido["incluir_cancelados"] == "False"


async def test_el_pre_vuelo_solo_normaliza_los_tipos_que_el_informe_declara(db: AsyncSession) -> None:
    """`TIPOS_COMPROBANTE` del informe acota el pre-vuelo. Sin el filtro, la primera generación
    posterior a subir `ETL_VERSION` —el mecanismo diseñado de reproceso— reprocesa el histórico
    completo de la empresa dentro de la tarea del informe, en serie y con un commit por
    comprobante: con 50 000 comprobantes son horas, y la pantalla sondea sin timeout."""
    from app.worker.tasks import _generar_informe_async

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    await _nomina_en_disco(db, empresa.empresa_id, "77777777-7777-7777-7777-777777777777", fixtures_cfdi.cfdi_nomina())
    id_ingreso = await _nomina_en_disco(
        db, empresa.empresa_id, "11111111-1111-1111-1111-111111111111", fixtures_cfdi.cfdi_ingreso(), tipo="I"
    )

    resultado = await _generar_informe_async(
        empresa.empresa_id, "B-02", {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"}, "op@test.mx"
    )
    assert resultado["filas"] == 1

    detalle_ingreso = await db.scalar(select(ComprobanteDetalle).where(ComprobanteDetalle.comprobante_id == id_ingreso))
    assert detalle_ingreso is None, "el CFDI de ingreso no lo necesita B-02: el pre-vuelo no debe tocarlo"


async def test_las_tres_columnas_de_totales_de_percepciones_llegan_a_la_celda(db: AsyncSession) -> None:
    """`Total sueldos`, `Total separación indemnización` y `Total jubilación pensión retiro` son
    columnas de **dinero del informe entregado** y no había una sola aserción sobre un valor no
    nulo suyo en toda la cadena ETL → escritor → informe. Un recibo ordinario deja las dos
    últimas vacías, así que un `None` en la celda pasaba por "no aplica" y nadie notaría que el
    ETL nunca las leyó."""
    from app.worker.tasks import _generar_informe_async

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    xml = fixtures_cfdi.cfdi_nomina(
        percepciones_xml=(
            '<nomina12:Percepcion TipoPercepcion="001" Clave="001" Concepto="Sueldo" ImporteGravado="8000.00" ImporteExento="0.00" />'
            '<nomina12:Percepcion TipoPercepcion="022" Clave="022" Concepto="Prima por antiguedad" ImporteGravado="1000.00" ImporteExento="0.00" />'
            '<nomina12:Percepcion TipoPercepcion="039" Clave="039" Concepto="Jubilacion" ImporteGravado="0.00" ImporteExento="500.00" />'
        ),
        total_sueldos="8000.00",
        total_separacion_indemnizacion="1000.00",
        total_jubilacion_pension_retiro="500.00",
        total_gravado="9000.00",
        total_exento="500.00",
        total_percepciones="9500.00",
        subtotal="9500.00",
        total="8408.90",
    )
    await _nomina_en_disco(db, empresa.empresa_id, "77777777-7777-7777-7777-777777777777", xml)

    resultado = await _generar_informe_async(
        empresa.empresa_id, "B-02", {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"}, "op@test.mx"
    )
    datos = _libro(resultado)["Datos"]  # type: ignore[index]

    assert float(_celda(datos, "Total sueldos")) == 8000.00
    assert float(_celda(datos, "Total separación indemnización")) == 1000.00
    assert float(_celda(datos, "Total jubilación pensión retiro")) == 500.00


async def test_pre_vuelo_sin_pendientes_no_deja_transaccion_de_lectura_abierta(db: AsyncSession, monkeypatch: pytest.MonkeyPatch) -> None:
    """`normalizar_lote` cierra la transacción de lectura por contrato, pero solo si se la llama.
    Con todo normalizado —el caso normal— el `SELECT` de `ids_pendientes` dejaba abierta la
    transacción implícita y la consulta del informe heredaba un snapshot de `REPEATABLE READ`
    anterior al pre-vuelo."""
    from app.informes import b02_conceptos_patron as b02
    from app.worker.tasks import _generar_informe_async

    empresa = await factories.crear_empresa(db, rfc="CHL960913IX9")
    consultar_original = b02.consultar
    en_transaccion: list[bool] = []

    async def _consultar_espia(db_, empresa_id, p):  # type: ignore[no-untyped-def]
        en_transaccion.append(db_.in_transaction())
        return await consultar_original(db_, empresa_id, p)

    monkeypatch.setattr(b02, "consultar", _consultar_espia)

    await _generar_informe_async(empresa.empresa_id, "B-02", {"fecha_desde": "2026-06-01", "fecha_hasta": "2026-07-31"}, "op@test.mx")

    assert en_transaccion == [False]
