"""B-06 · Costo de nómina por centro de costo.

Lo que estas pruebas fijan, en orden de importancia:

1. **El agrupamiento es auditable.** El informe dice con qué resolvió cada CFDI su centro de
   costo (`RESOLUCION_DE_CENTRO_DE_COSTO`) y avisa **una sola vez**, con el conteo, cuando
   tuvo que caer al texto crudo (`DEPARTAMENTO_SIN_MAPEO`). Es la organización la que ejecuta
   recursos etiquetados: un total por centro que nadie puede defender no comprueba nada.
2. **Sin mapeo, dos variantes ortográficas del mismo departamento son dos centros distintos.**
   No es un defecto que estas pruebas toleren: es el problema que `map_departamento` existe
   para resolver, y verlo es lo que justifica que el mapeo exista. Si algún día el informe
   "arregla" esto por su cuenta normalizando el texto, esta prueba tiene que fallar.
3. **B-06.R3: cada CFDI al departamento que él declara.** El departamento actual no se
   retropropaga; el gasto se quedó donde se ejecutó.
4. **Cada regla con su gemela negativa.** Una bandera que dispara siempre es peor que no
   tenerla — con una excepción declarada y probada: la de calidad del agrupamiento, que es un
   reporte y no una alerta.

Ninguna prueba de este archivo se dio por buena sin romper a propósito lo que dice proteger y
comprobar que **falla** (ver las tres trampas de la evidencia por mutación anotadas en
`tests/test_cli_configuracion.py`).

**Y una cuarta trampa, encontrada aquí y la más traicionera de las cuatro: la mutación que es
un no-op.** Las otras tres hacen que la comprobación mienta; esta hace que dudes de una prueba
que sí protege. Para "apagar" el reporte de descuadres se escribió
`banderas.extend([] or universo_nomina.banderas_de_totales_descuadrados(...))`, y `[] or X`
evalúa a `X`: el informe siguió emitiendo la bandera, la prueba siguió pasando y la mutación
"sobrevivió" sin haber cambiado nada. Antes de dar por débil una prueba porque su mutación
sobrevive, hay que comprobar que la mutación **hace algo**.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import update
from sqlalchemy.ext.asyncio import AsyncSession

from app.informes import b06_centro_costo as b06
from app.informes import excel, registro
from app.informes.base import ContextoInforme
from app.models.configuracion_fiscal import MapDepartamento
from app.models.enums import EstatusCfdi
from app.models.nomina import NominaReceptor
from tests import factories
from tests.helpers_nomina import insertar_nomina

RFC_EMPRESA = "CHL960913IX9"

_DESDE = date(2026, 1, 1)
_HASTA = date(2026, 12, 31)

# Los dos departamentos reales de la empresa 11, que hoy no tienen ningún mapeo cargado: la
# ruta degradada de estas pruebas es la que corre en vivo.
_EDIFICIOS = "EDIFICIOS"
_SOCIAL = "SOCIAL"


# --------------------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------------------


def _p(**extra: Any) -> Any:
    return b06.Parametros(fecha_desde=_DESDE, fecha_hasta=_HASTA, **extra)


def _columna(resultado: Any, titulo: str) -> int:
    titulos = [c.titulo for c in resultado.columnas]
    assert titulo in titulos, f"falta la columna {titulo!r}; hay {titulos}"
    return titulos.index(titulo)


def _valor(resultado: Any, fila: int, titulo: str) -> Any:
    return resultado.filas[fila][_columna(resultado, titulo)]


def _claves(resultado: Any) -> list[str]:
    return [b.clave for b in resultado.banderas]


def _de_clave(resultado: Any, clave: str) -> list[Any]:
    return [b for b in resultado.banderas if b.clave == clave]


def _centros(resultado: Any) -> list[Any]:
    indice = _columna(resultado, "Centro de costo")
    return [fila[indice] for fila in resultado.filas]


async def _mapear(db: AsyncSession, empresa_id: int, texto: str, centro: str) -> None:
    db.add(MapDepartamento(empresa_id=empresa_id, departamento_texto=texto, centro_costo=centro))
    await db.commit()


async def _borrar_departamento(db: AsyncSession, comprobante_id: int) -> None:
    """Deja `nomina_receptor.departamento` en NULL, que es como llega un CFDI que no trae el
    nodo `Departamento` (es opcional en el complemento). Se hace con un `UPDATE` en vez de
    con un parámetro del helper compartido para no tocar `tests/helpers_nomina.py`, que usan
    todas las pruebas del grupo B."""
    await db.execute(
        update(NominaReceptor).where(NominaReceptor.comprobante_id == comprobante_id).values(departamento=None)
    )
    await db.commit()
    # La sesión de pruebas usa `expire_on_commit=False`, así que el objeto que el helper dejó
    # en el mapa de identidad podría responder con el texto viejo aunque la base ya diga NULL:
    # sin esto la prueba mediría la memoria de SQLAlchemy, no el informe. Se **expulsa** en vez
    # de expirar: expirar dejaría a los objetos ya devueltos (la empresa, por ejemplo)
    # recargándose por IO perezosa fuera del greenlet de SQLAlchemy async.
    db.expunge_all()


# --------------------------------------------------------------------------------------
# 1. El grano
# --------------------------------------------------------------------------------------


async def test_el_grano_es_periodo_y_centro_de_costo(db: AsyncSession) -> None:
    """Dos CFDI del mismo departamento y el mismo periodo son **una** fila; el mismo
    departamento en otro periodo es otra fila."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, rfc, pago in (
        ("11111111-1111-1111-1111-111111111111", "XAXX010101000", date(2026, 6, 30)),
        ("22222222-2222-2222-2222-222222222222", "XEXX010101000", date(2026, 6, 30)),
        ("33333333-3333-3333-3333-333333333333", "XAXX010101000", date(2026, 7, 31)),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=rfc,
            fecha_pago=pago,
            departamento=_EDIFICIOS,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert len(resultado.filas) == 2
    assert [_valor(resultado, i, "Periodo") for i in range(2)] == [6, 7]
    assert [_valor(resultado, i, "Ejercicio") for i in range(2)] == [2026, 2026]
    assert _centros(resultado) == [_EDIFICIOS, _EDIFICIOS]
    # Dos empleados y dos CFDI en junio; uno y uno en julio.
    assert _valor(resultado, 0, "Núm. de empleados") == 2
    assert _valor(resultado, 0, "Núm. de CFDI") == 2
    assert _valor(resultado, 1, "Núm. de empleados") == 1


async def test_dos_cfdi_del_mismo_empleado_son_un_empleado_y_dos_cfdi(db: AsyncSession) -> None:
    """Gemela de la anterior sobre la columna que más fácil se equivoca: contar recibos como
    personas inflaría la plantilla del centro y hundiría el costo promedio."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, tipo in (
        ("11111111-1111-1111-1111-111111111111", "O"),
        ("22222222-2222-2222-2222-222222222222", "E"),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor="XAXX010101000",
            fecha_pago=date(2026, 6, 30),
            tipo_nomina=tipo,
            departamento=_EDIFICIOS,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert len(resultado.filas) == 1
    assert _valor(resultado, 0, "Núm. de empleados") == 1
    assert _valor(resultado, 0, "Núm. de CFDI") == 2
    # Y la nómina extraordinaria SÍ entra: el aguinaldo es costo del centro igual que la
    # quincena ordinaria.
    assert _valor(resultado, 0, "Costo bruto") == Decimal("2000.00")


# --------------------------------------------------------------------------------------
# 2. El desglose de importes
# --------------------------------------------------------------------------------------


async def test_desglose_de_sueldos_prestaciones_y_asimilados(db: AsyncSession) -> None:
    """Sueldos es `001`, asimilados es `046` y prestaciones es todo lo demás. Las tres suman
    exactamente el total de percepciones."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[
            ("001", "001", "Sueldo", "8000.00", "0.00"),
            ("002", "002", "Aguinaldo", "3000.00", "1000.00"),
            ("019", "019", "Horas extra", "250.00", "250.00"),
            ("046", "046", "Asimilado", "500.00", "0.00"),
        ],
        total_percepciones="13000.00",
        otros_pagos=[("002", "035", "Subsidio", "120.00", "120.00")],
        total_otros_pagos="120.00",
        deducciones=[("002", "045", "I.S.R.", "900.00"), ("001", "052", "I.M.S.S.", "300.00")],
        total_deducciones="1200.00",
        total="11920.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Sueldos") == Decimal("8000.00")
    # Aguinaldo (4000) + horas extra (500): lo que no es 001 ni 046.
    assert _valor(resultado, 0, "Prestaciones") == Decimal("4500.00")
    assert _valor(resultado, 0, "Asimilados") == Decimal("500.00")
    assert _valor(resultado, 0, "Total percepciones") == Decimal("13000.00")
    assert _valor(resultado, 0, "Otros pagos") == Decimal("120.00")
    # El subsidio es un DESGLOSE de «Otros pagos», no un sumando aparte: sigue dentro del costo
    # bruto (13000 + 120), que es lo que ata con `nomina.total_percepciones + total_otros_pagos`.
    assert _valor(resultado, 0, "Subsidio para el empleo entregado") == Decimal("120.00")
    assert _valor(resultado, 0, "Costo bruto") == Decimal("13120.00")
    assert _valor(resultado, 0, "ISR retenido") == Decimal("900.00")
    assert _valor(resultado, 0, "IMSS obrero retenido") == Decimal("300.00")
    assert _valor(resultado, 0, "Neto pagado") == Decimal("11920.00")
    assert _valor(resultado, 0, "Días pagados") == Decimal("15.000")
    # No hay descuadre que reportar: los encabezados coinciden con la suma de los nodos.
    assert "TOTALES_DESCUADRADOS" not in _claves(resultado)


async def test_los_asimilados_no_se_cuentan_como_prestaciones(db: AsyncSession) -> None:
    """Gemela negativa del desglose: `046` es otro régimen de contratación, no una prestación
    laboral. Sumarlo en «Prestaciones» daría un costo de prestaciones falso."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("046", "046", "Asimilado", "5000.00", "0.00")],
        total_percepciones="5000.00",
        total="5000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Asimilados") == Decimal("5000.00")
    assert _valor(resultado, 0, "Prestaciones") == Decimal("0")
    assert _valor(resultado, 0, "Sueldos") == Decimal("0")


async def test_las_celdas_de_importe_sin_dato_son_cero_no_vacio(db: AsyncSession) -> None:
    """R-T7: un nulo en una celda de importe rompe cualquier suma de la hoja."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        deducciones=[],
        total_deducciones="0.00",
        otros_pagos=[],
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    for titulo in (
        "Prestaciones",
        "Asimilados",
        "Otros pagos",
        "Subsidio para el empleo entregado",
        "ISR retenido",
        "IMSS obrero retenido",
    ):
        assert _valor(resultado, 0, titulo) == Decimal("0"), titulo
        assert _valor(resultado, 0, titulo) is not None, titulo


async def test_un_otro_pago_que_no_es_subsidio_no_llena_la_columna_del_subsidio(db: AsyncSession) -> None:
    """Gemela negativa del desglose del subsidio: los viáticos son «Otros pagos» y no subsidio.

    Si la columna copiara «Otros pagos», el lector restaría dinero que el patrón sí desembolsó y
    no recupera — el error espejo del que la columna existe para evitar.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        otros_pagos=[("999", "099", "Ajuste al neto", "50.00", "")],
        total_otros_pagos="50.00",
        total="1050.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Otros pagos") == Decimal("50.00")
    assert _valor(resultado, 0, "Subsidio para el empleo entregado") == Decimal("0")
    assert _valor(resultado, 0, "Costo bruto") == Decimal("1050.00")


async def test_el_subsidio_se_desglosa_sin_salir_del_costo_bruto(db: AsyncSession) -> None:
    """El subsidio entregado es dinero federal recuperable contra el ISR, no costo propio — pero
    «Costo bruto» **no se redefine**: tiene que atar con `total_percepciones + total_otros_pagos`
    y con lo que B-01 y B-02 reportan del mismo periodo. La salida es aditiva: se publica aparte
    para que el lector reste, y el costo bruto sigue significando lo de la ficha.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        otros_pagos=[("002", "035", "Subs al Empleo mes", "80.00", "80.00"), ("999", "099", "Ajuste", "20.00", "")],
        total_otros_pagos="100.00",
        total="1100.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Otros pagos") == Decimal("100.00")
    assert _valor(resultado, 0, "Subsidio para el empleo entregado") == Decimal("80.00")
    # Ata con el encabezado del complemento: 1000 + 100. El subsidio NO se resta.
    assert _valor(resultado, 0, "Costo bruto") == Decimal("1100.00")
    assert "TOTALES_DESCUADRADOS" not in _claves(resultado)
    # Y la columna va entre «Otros pagos» y «Costo bruto», donde se lee como su desglose.
    titulos = [c.titulo for c in resultado.columnas]
    assert titulos.index("Otros pagos") + 1 == titulos.index("Subsidio para el empleo entregado")
    assert titulos.index("Subsidio para el empleo entregado") + 1 == titulos.index("Costo bruto")


async def test_el_costo_promedio_por_empleado_divide_entre_empleados_no_entre_cfdi(db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, rfc in (
        ("11111111-1111-1111-1111-111111111111", "XAXX010101000"),
        ("22222222-2222-2222-2222-222222222222", "XAXX010101000"),
        ("33333333-3333-3333-3333-333333333333", "XEXX010101000"),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=rfc,
            fecha_pago=date(2026, 6, 30),
            departamento=_EDIFICIOS,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Costo bruto") == Decimal("3000.00")
    assert _valor(resultado, 0, "Núm. de CFDI") == 3
    assert _valor(resultado, 0, "Núm. de empleados") == 2
    assert _valor(resultado, 0, "Costo promedio por empleado") == Decimal("1500.00")


async def test_el_porcentaje_del_periodo_suma_cien_por_periodo(db: AsyncSession) -> None:
    """Tres centros en junio y uno en julio: cada periodo suma 100 por su cuenta."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    casos = (
        ("11111111-1111-1111-1111-111111111111", "AAA", date(2026, 6, 30), "2500.00"),
        ("22222222-2222-2222-2222-222222222222", "BBB", date(2026, 6, 30), "2500.00"),
        ("33333333-3333-3333-3333-333333333333", "CCC", date(2026, 6, 30), "5000.00"),
        ("44444444-4444-4444-4444-444444444444", "AAA", date(2026, 7, 31), "700.00"),
    )
    for uuid_cfdi, departamento, pago, importe in casos:
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            fecha_pago=pago,
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", importe, "0.00")],
            total_percepciones=importe,
            total=importe,
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    indice = _columna(resultado, "% del total del periodo")
    de_junio = [fila[indice] for fila in resultado.filas if fila[_columna(resultado, "Periodo")] == 6]
    de_julio = [fila[indice] for fila in resultado.filas if fila[_columna(resultado, "Periodo")] == 7]
    assert sorted(de_junio) == [Decimal("25"), Decimal("25"), Decimal("50")]
    assert sum(de_junio) == Decimal("100")
    # El de julio es 100 aunque su importe sea mucho menor: el denominador es su propio
    # periodo, no el informe entero.
    assert de_julio == [Decimal("100")]


# --------------------------------------------------------------------------------------
# 3. B-06.R1 — la cascada de dos niveles
# --------------------------------------------------------------------------------------


async def test_con_mapeo_cargado_agrupa_por_centro_de_costo(db: AsyncSession) -> None:
    """Nivel 1: dos departamentos distintos que mapean al mismo centro son **una** fila."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, "CC-OPERACION")
    await _mapear(db, empresa.empresa_id, _SOCIAL, "CC-OPERACION")
    for uuid_cfdi, departamento in (
        ("11111111-1111-1111-1111-111111111111", _EDIFICIOS),
        ("22222222-2222-2222-2222-222222222222", _SOCIAL),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor="XAXX010101000" if departamento == _EDIFICIOS else "XEXX010101000",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _centros(resultado) == ["CC-OPERACION"]
    assert _valor(resultado, 0, "Costo bruto") == Decimal("2000.00")
    # Gemela negativa de la alerta: todo resolvió por el mapeo, así que no hay aviso.
    assert "DEPARTAMENTO_SIN_MAPEO" not in _claves(resultado)


async def test_sin_mapeo_agrupa_por_el_texto_crudo_con_una_sola_bandera_agregada(db: AsyncSession) -> None:
    """Nivel 2, el caso que corre en vivo (la empresa real no tiene ningún mapeo cargado).

    La bandera es **una**, con el conteo de CFDI afectados y los textos: una por fila
    sepultaría los hallazgos de la hoja `Banderas`.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", _EDIFICIOS),
            ("22222222-2222-2222-2222-222222222222", _EDIFICIOS),
            ("33333333-3333-3333-3333-333333333333", _SOCIAL),
        )
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert sorted(_centros(resultado)) == [_EDIFICIOS, _SOCIAL]
    banderas = _de_clave(resultado, "DEPARTAMENTO_SIN_MAPEO")
    assert len(banderas) == 1, "la bandera es agregada, no una por fila ni una por CFDI"
    assert banderas[0].ambito == "informe"
    # El conteo es de CFDI afectados, y los textos sin mapear van citados para poder
    # capturarlos sin abrir la base.
    assert "3 CFDI" in banderas[0].mensaje
    assert _EDIFICIOS in banderas[0].mensaje and _SOCIAL in banderas[0].mensaje


async def test_dos_variantes_ortograficas_sin_mapeo_caen_en_grupos_distintos(db: AsyncSession) -> None:
    """El problema que `map_departamento` existe para resolver, visible.

    El informe **no** normaliza el texto por su cuenta: hacerlo escondería el problema a
    medias (la pantalla de configuración seguiría pidiendo mapear las dos variantes por
    separado) y quitaría la razón de ser del mapeo. Si alguien agrega un `strip()`, un
    `.upper()` o un plegado de acentos a la resolución, esta prueba tiene que fallar.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", "EDIFICIOS"),
            ("22222222-2222-2222-2222-222222222222", "Edificios "),
        )
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert sorted(_centros(resultado)) == ["EDIFICIOS", "Edificios "]
    assert len(resultado.filas) == 2
    assert len(_de_clave(resultado, "DEPARTAMENTO_SIN_MAPEO")) == 1


async def test_el_mapeo_une_las_dos_variantes_ortograficas(db: AsyncSession) -> None:
    """Gemela positiva de la anterior: con el mapeo capturado sobre los dos textos, las dos
    variantes caen en el mismo centro. Es lo que demuestra que el mapeo sirve.

    Las variantes de esta prueba difieren en más que mayúsculas y espacios a propósito, y no
    por comodidad: la PK de `map_departamento` vive en una tabla `utf8mb4_unicode_ci`, que es
    insensible a mayúsculas y con relleno de espacios, así que **no se pueden capturar dos
    renglones que solo difieran en eso** — el segundo choca con el primero. Ver el residuo
    declarado en el docstring de `app/informes/b06_centro_costo.py`.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, "EDIFICIOS", "CC-EDIF")
    await _mapear(db, empresa.empresa_id, "EDIFICIO", "CC-EDIF")
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", "EDIFICIOS"),
            ("22222222-2222-2222-2222-222222222222", "EDIFICIO"),
        )
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _centros(resultado) == ["CC-EDIF"]
    assert _valor(resultado, 0, "Costo bruto") == Decimal("2000.00")
    assert "DEPARTAMENTO_SIN_MAPEO" not in _claves(resultado)


async def test_los_dos_niveles_se_funden_cuando_resuelven_al_mismo_centro(db: AsyncSession) -> None:
    """El nivel de resolución **no** forma parte de la llave de agrupamiento: "EDIF" mapeado
    a "EDIFICIOS" y el texto crudo "EDIFICIOS" son el mismo centro de costo."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, "EDIF", "EDIFICIOS")
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", "EDIF"),
            ("22222222-2222-2222-2222-222222222222", "EDIFICIOS"),
        )
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _centros(resultado) == ["EDIFICIOS"]
    assert _valor(resultado, 0, "Núm. de CFDI") == 2
    # Y aun así se avisa del que cayó al texto crudo: el grupo cuadra, la configuración no.
    assert len(_de_clave(resultado, "DEPARTAMENTO_SIN_MAPEO")) == 1


# --------------------------------------------------------------------------------------
# 3-bis. El otro lado del agrupamiento: el VALOR configurado del centro de costo
# --------------------------------------------------------------------------------------


async def test_dos_centros_de_costo_que_solo_difieren_en_forma_se_reportan_sin_unificarse(
    db: AsyncSession,
) -> None:
    """`map_departamento.centro_costo` es texto libre sin catálogo.

    Sin esta bandera, mapear `EDIF → "EDIFICIOS"` y `EDIFICIO → "Edificios"` parte el gasto en
    dos centros, **los dos contados como nivel 1**, y el censo diría "2 se resolvieron con el
    mapeo" con el informe viéndose impecable. Es la asimetría que el módulo no cubría: el lado
    del texto del departamento estaba instrumentado y el del valor configurado no.

    Y **no se unifican**: declarar que dos centros son el mismo es una decisión contable, no
    tipográfica. El informe lo hace visible, no lo arregla por su cuenta.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, "EDIF", "EDIFICIOS")
    await _mapear(db, empresa.empresa_id, "EDIFICIO", "Edificios")
    # Y la variante con acento, que es la que un `upper()` a secas no atrapa.
    await _mapear(db, empresa.empresa_id, "GEST", "OPERACIÓN")
    await _mapear(db, empresa.empresa_id, "GESTION", "OPERACION")
    for indice, departamento in enumerate(("EDIF", "EDIFICIO", "GEST", "GESTION")):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=f"{indice + 1}1111111-1111-1111-1111-111111111111",
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    # Cuatro centros distintos, no dos: el gasto sale partido, y eso es lo que la bandera dice.
    assert sorted(_centros(resultado)) == ["EDIFICIOS", "Edificios", "OPERACION", "OPERACIÓN"]
    banderas = _de_clave(resultado, "CENTRO_DE_COSTO_AMBIGUO")
    assert len(banderas) == 2, "una por grupo de variantes que colisionan, no una por renglón"
    assert all(b.severidad == "media" and b.ambito == "informe" for b in banderas)
    mensajes = " ".join(b.mensaje for b in banderas)
    assert "'EDIFICIOS'" in mensajes and "'Edificios'" in mensajes
    assert "'OPERACIÓN'" in mensajes and "'OPERACION'" in mensajes
    # Nombra los departamentos afectados: es lo que hay que ir a corregir.
    assert "'EDIF'" in mensajes and "'GESTION'" in mensajes
    # El censo sigue diciendo la verdad de la cascada: los cuatro resolvieron por el mapeo.
    assert "4 se resolvieron con el mapeo" in _de_clave(resultado, "RESOLUCION_DE_CENTRO_DE_COSTO")[0].mensaje


async def test_centros_de_costo_bien_capturados_no_disparan_ninguna_bandera_de_configuracion(
    db: AsyncSession,
) -> None:
    """Gemela negativa de las dos anteriores: con dos centros de verdad distintos y ninguno en
    blanco, no hay nada que reportar. Una bandera que dispara siempre no la lee nadie."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, "CC-EDIFICIOS")
    await _mapear(db, empresa.empresa_id, _SOCIAL, "CC-SOCIAL")
    for indice, departamento in enumerate((_EDIFICIOS, _SOCIAL)):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=f"{indice + 1}1111111-1111-1111-1111-111111111111",
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert sorted(_centros(resultado)) == ["CC-EDIFICIOS", "CC-SOCIAL"]
    assert "CENTRO_DE_COSTO_AMBIGUO" not in _claves(resultado)
    assert "CENTRO_DE_COSTO_EN_BLANCO" not in _claves(resultado)


async def test_un_centro_de_costo_configurado_en_blanco_se_reporta(db: AsyncSession) -> None:
    """`min_length=1` en el esquema deja pasar un espacio: el costo sale en una fila con la
    columna «Centro de costo» en blanco y contado como agrupamiento configurado y revisado."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, " ")
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    banderas = _de_clave(resultado, "CENTRO_DE_COSTO_EN_BLANCO")
    assert len(banderas) == 1
    assert banderas[0].severidad == "media"
    assert repr(_EDIFICIOS) in banderas[0].mensaje
    # Un solo centro en blanco NO es además "ambiguo": el diagnóstico útil es que está vacío.
    assert "CENTRO_DE_COSTO_AMBIGUO" not in _claves(resultado)


async def test_la_ambiguedad_se_detecta_desde_el_mapa_aunque_no_haya_cobrado_nadie(db: AsyncSession) -> None:
    """Una configuración ambigua es un defecto de configuración aunque en este rango no haya
    cobrado nadie de esos departamentos: se arregla una vez y sirve para todos los rangos. Es la
    misma razón por la que la bandera no cuelga de ninguna fila impresa."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, "OTRO_A", "MANTENIMIENTO")
    await _mapear(db, empresa.empresa_id, "OTRO_B", "Mantenimiento")
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _centros(resultado) == [_EDIFICIOS], "ninguna fila usa los centros ambiguos"
    assert len(_de_clave(resultado, "CENTRO_DE_COSTO_AMBIGUO")) == 1


async def test_un_cfdi_sin_departamento_se_agrupa_aparte_con_su_propia_bandera(db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", _EDIFICIOS),
            ("22222222-2222-2222-2222-222222222222", "SE VA A NULO"),
            ("33333333-3333-3333-3333-333333333333", "   "),
        )
    ):
        cid = await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )
        if departamento == "SE VA A NULO":
            # El CFDI que no trae el nodo `Departamento`: NULL en la base, no cadena vacía.
            await _borrar_departamento(db, cid)

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert sorted(_centros(resultado)) == sorted([_EDIFICIOS, b06.SIN_DEPARTAMENTO])
    banderas = _de_clave(resultado, "DEPARTAMENTO_AUSENTE")
    assert len(banderas) == 1
    assert "2 CFDI" in banderas[0].mensaje
    # Un departamento en blanco NO es un texto sin mapear: no hay nada que capturar.
    sin_mapeo = _de_clave(resultado, "DEPARTAMENTO_SIN_MAPEO")
    assert len(sin_mapeo) == 1
    assert "1 CFDI" in sin_mapeo[0].mensaje


async def test_sin_cfdi_sin_departamento_no_hay_bandera_de_departamento_ausente(db: AsyncSession) -> None:
    """Gemela negativa: la bandera no dispara cuando todos los CFDI traen departamento."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert "DEPARTAMENTO_AUSENTE" not in _claves(resultado)


async def test_un_centro_llamado_como_el_rotulo_de_ausentes_no_se_funde_con_ellos(db: AsyncSession) -> None:
    """La guarda de `SIN_DEPARTAMENTO`: es un rótulo de presentación, no una llave.

    Rebuscado —hay que capturar un centro de costo llamado exactamente «(sin departamento)»—
    pero si se fundiera, el grupo mezclaría costo atribuido con costo no atribuible mientras el
    censo los cuenta en niveles distintos, y el rótulo dejaría de significar lo que dice.
    """
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, b06.SIN_DEPARTAMENTO)
    for indice, (uuid_cfdi, departamento) in enumerate(
        (
            ("11111111-1111-1111-1111-111111111111", _EDIFICIOS),
            ("22222222-2222-2222-2222-222222222222", "SE VA A NULO"),
        )
    ):
        cid = await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", f"{(indice + 1) * 1000}.00", "0.00")],
            total_percepciones=f"{(indice + 1) * 1000}.00",
            total=f"{(indice + 1) * 1000}.00",
        )
        if departamento == "SE VA A NULO":
            await _borrar_departamento(db, cid)

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    # Dos filas con la misma etiqueta y distinto origen, no una fila con el costo mezclado.
    assert _centros(resultado) == [b06.SIN_DEPARTAMENTO, b06.SIN_DEPARTAMENTO]
    assert sorted(_valor(resultado, i, "Costo bruto") for i in range(2)) == [
        Decimal("1000.00"),
        Decimal("2000.00"),
    ]
    mensaje = _de_clave(resultado, "RESOLUCION_DE_CENTRO_DE_COSTO")[0].mensaje
    assert "1 se resolvieron con el mapeo" in mensaje
    assert "1 no traen departamento" in mensaje


# --------------------------------------------------------------------------------------
# 4. El reporte de calidad del agrupamiento
# --------------------------------------------------------------------------------------


async def test_el_informe_reporta_cuantos_cfdi_resolvio_cada_nivel(db: AsyncSession) -> None:
    """Lo que la ficha pide explícitamente: sin este conteo, un total por centro de costo no
    dice qué parte salió de configuración revisada y qué parte de un texto libre."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, "CC-EDIF")
    casos = (
        ("11111111-1111-1111-1111-111111111111", _EDIFICIOS),
        ("22222222-2222-2222-2222-222222222222", _EDIFICIOS),
        ("33333333-3333-3333-3333-333333333333", _SOCIAL),
        ("44444444-4444-4444-4444-444444444444", "SE VA A NULO"),
    )
    for indice, (uuid_cfdi, departamento) in enumerate(casos):
        cid = await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=f"XAXX01010100{indice}",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )
        if departamento == "SE VA A NULO":
            await _borrar_departamento(db, cid)

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    banderas = _de_clave(resultado, "RESOLUCION_DE_CENTRO_DE_COSTO")
    assert len(banderas) == 1
    mensaje = banderas[0].mensaje
    assert banderas[0].severidad == "baja", "es un reporte, no una alerta"
    assert banderas[0].ambito == "informe"
    assert "sobre 4 CFDI" in mensaje
    assert "2 se resolvieron con el mapeo" in mensaje
    assert "1 con el texto crudo" in mensaje
    assert "1 no traen departamento" in mensaje


async def test_el_reporte_de_calidad_se_emite_aunque_todo_resuelva_por_el_mapeo(db: AsyncSession) -> None:
    """La excepción declarada a "una bandera que dispara siempre no sirve": esta es un
    **reporte** de auditabilidad, no una alerta, y por eso lleva clave y severidad propias.
    Si solo se emitiera al degradar, el informe limpio no podría defenderse."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await _mapear(db, empresa.empresa_id, _EDIFICIOS, "CC-EDIF")
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    banderas = _de_clave(resultado, "RESOLUCION_DE_CENTRO_DE_COSTO")
    assert len(banderas) == 1
    assert "1 se resolvieron con el mapeo" in banderas[0].mensaje
    assert "0 con el texto crudo" in banderas[0].mensaje
    assert "La empresa tiene `map_departamento` cargado." in banderas[0].mensaje
    # Y la alerta sí calla, que es lo que la separa del reporte.
    assert "DEPARTAMENTO_SIN_MAPEO" not in _claves(resultado)


async def test_el_reporte_de_calidad_dice_que_la_empresa_no_tiene_mapeo(db: AsyncSession) -> None:
    """El estado real de la instalación: sin un solo renglón en `map_departamento`, el nivel 1
    no puede resolver nada y el reporte lo dice en vez de dejar un cero sin explicar."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    mensaje = _de_clave(resultado, "RESOLUCION_DE_CENTRO_DE_COSTO")[0].mensaje
    assert "no tiene ningún renglón en `map_departamento`" in mensaje


async def test_el_mapeo_de_otra_empresa_no_resuelve_el_centro_de_esta(db: AsyncSession) -> None:
    """`map_departamento` cuelga de `empresa_id`: el mismo texto puede ser otro centro en
    otra organización. Sin el filtro, el informe agruparía con la configuración ajena."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    otra = await factories.crear_empresa(db, nombre="Otra", rfc="EKU9003173C9")
    await _mapear(db, otra.empresa_id, _EDIFICIOS, "CC-DE-LA-OTRA")
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _centros(resultado) == [_EDIFICIOS]
    assert len(_de_clave(resultado, "DEPARTAMENTO_SIN_MAPEO")) == 1


# --------------------------------------------------------------------------------------
# 5. B-06.R3 — el departamento no se retropropaga
# --------------------------------------------------------------------------------------


async def test_un_empleado_que_cambia_de_area_se_asigna_al_departamento_de_cada_cfdi(db: AsyncSession) -> None:
    """B-06.R3. Retropropagar el departamento actual reescribiría la historia del gasto, que
    es justo lo que un financiador no puede aceptar."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, pago, departamento, importe in (
        ("11111111-1111-1111-1111-111111111111", date(2026, 6, 30), _EDIFICIOS, "1000.00"),
        ("22222222-2222-2222-2222-222222222222", date(2026, 7, 31), _SOCIAL, "2000.00"),
        ("33333333-3333-3333-3333-333333333333", date(2026, 8, 31), _SOCIAL, "3000.00"),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor="XAXX010101000",
            fecha_pago=pago,
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", importe, "0.00")],
            total_percepciones=importe,
            total=importe,
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    por_periodo = {
        (_valor(resultado, i, "Periodo"), _valor(resultado, i, "Centro de costo")): _valor(resultado, i, "Costo bruto")
        for i in range(len(resultado.filas))
    }
    assert por_periodo == {
        (6, _EDIFICIOS): Decimal("1000.00"),
        (7, _SOCIAL): Decimal("2000.00"),
        (8, _SOCIAL): Decimal("3000.00"),
    }
    # Lo que la retropropagación habría producido: junio bajo SOCIAL, el área actual.
    assert (6, _SOCIAL) not in por_periodo


async def test_un_empleado_en_dos_centros_del_mismo_periodo_suma_en_los_dos(db: AsyncSession) -> None:
    """El caso extremo de B-06.R3: dos CFDI del mismo trabajador y del mismo periodo con
    departamentos distintos van cada uno a su centro, no al mayor ni al último."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, departamento, importe in (
        ("11111111-1111-1111-1111-111111111111", _EDIFICIOS, "1000.00"),
        ("22222222-2222-2222-2222-222222222222", _SOCIAL, "400.00"),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor="XAXX010101000",
            fecha_pago=date(2026, 6, 30),
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", importe, "0.00")],
            total_percepciones=importe,
            total=importe,
        )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert len(resultado.filas) == 2
    assert _centros(resultado) == [_EDIFICIOS, _SOCIAL]
    assert _valor(resultado, 0, "Costo bruto") == Decimal("1000.00")
    assert _valor(resultado, 1, "Costo bruto") == Decimal("400.00")
    assert _valor(resultado, 0, "Núm. de empleados") == 1
    assert _valor(resultado, 1, "Núm. de empleados") == 1


# --------------------------------------------------------------------------------------
# 6. `detalle_empleado`
# --------------------------------------------------------------------------------------


async def test_detalle_empleado_cambia_el_grano_y_agrega_la_identidad(db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    for uuid_cfdi, rfc, nombre, num in (
        ("11111111-1111-1111-1111-111111111111", "XAXX010101000", "JUANA INVENTADA", "001"),
        ("22222222-2222-2222-2222-222222222222", "XEXX010101000", "PEDRO INVENTADO", "002"),
    ):
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=rfc,
            num_empleado=num,
            nombre_receptor=nombre,
            fecha_pago=date(2026, 6, 30),
            departamento=_EDIFICIOS,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    agregado = await b06.consultar(db, empresa.empresa_id, _p())
    detallado = await b06.consultar(db, empresa.empresa_id, _p(detalle_empleado=True))

    assert len(agregado.filas) == 1
    assert "RFC empleado" not in [c.titulo for c in agregado.columnas]

    assert len(detallado.filas) == 2
    assert [_valor(detallado, i, "RFC empleado") for i in range(2)] == ["XAXX010101000", "XEXX010101000"]
    assert _valor(detallado, 0, "Nombre empleado") == "JUANA INVENTADA"
    assert _valor(detallado, 0, "Núm. empleado") == "001"
    assert _valor(detallado, 0, "Centro de costo") == _EDIFICIOS
    assert _valor(detallado, 0, "Núm. de empleados") == 1
    # El denominador del porcentaje sigue siendo el periodo, así que sigue sumando 100.
    indice = _columna(detallado, "% del total del periodo")
    assert sum(fila[indice] for fila in detallado.filas) == Decimal("100")


async def test_el_detalle_no_publica_curp_ni_nss(db: AsyncSession) -> None:
    """La fuga real de B-10, cerrada aquí por construcción: B-06 no lee los datos personales
    del receptor, así que no pueden aparecer ni en una celda ni en el mensaje de una bandera."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    curp = "XXXX800101HCHXXX99"
    nss = "99988877766"
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        curp=curp,
        nss=nss,
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p(detalle_empleado=True))

    titulos = [c.titulo for c in resultado.columnas]
    assert "CURP" not in titulos and "NSS" not in titulos
    celdas = [str(celda) for fila in resultado.filas for celda in fila]
    assert curp not in celdas and nss not in celdas
    mensajes = " ".join(b.mensaje + b.ambito for b in resultado.banderas)
    assert curp not in mensajes and nss not in mensajes


# --------------------------------------------------------------------------------------
# 7. Universo, avisos y orden
# --------------------------------------------------------------------------------------


async def test_sin_comprobantes_devuelve_aviso_y_columnas(db: AsyncSession) -> None:
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert resultado.filas == []
    assert resultado.aviso == "Sin CFDI de nómina en el rango solicitado."
    # Nunca cero columnas: un libro sin encabezados es indistinguible de un informe roto.
    assert [c.titulo for c in resultado.columnas][:3] == ["Ejercicio", "Periodo", "Centro de costo"]


async def test_empresa_inexistente_no_revienta(db: AsyncSession) -> None:
    resultado = await b06.consultar(db, 987654, _p())
    assert resultado.aviso == "La empresa no existe."
    assert resultado.filas == []


async def test_el_cancelado_se_excluye_por_defecto_y_entra_con_bandera_si_se_pide(db: AsyncSession) -> None:
    """R-T1 con la divergencia declarada del grupo B: el cancelado no suma salvo que se pida,
    y cuando se pide lleva bandera para que su importe no entre mudo en el costo del centro."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        estatus=EstatusCfdi.VIGENTE,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="22222222-2222-2222-2222-222222222222",
        rfc_receptor="XEXX010101000",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        estatus=EstatusCfdi.CANCELADO,
        percepciones=[("001", "001", "Sueldo", "7000.00", "0.00")],
        total_percepciones="7000.00",
        total="7000.00",
    )

    sin_cancelados = await b06.consultar(db, empresa.empresa_id, _p())
    con_cancelados = await b06.consultar(db, empresa.empresa_id, _p(incluir_cancelados=True))

    assert _valor(sin_cancelados, 0, "Costo bruto") == Decimal("1000.00")
    assert "COMPROBANTE_CANCELADO" not in _claves(sin_cancelados)
    assert _valor(con_cancelados, 0, "Costo bruto") == Decimal("8000.00")
    assert "COMPROBANTE_CANCELADO" in _claves(con_cancelados)


async def test_el_encabezado_descuadrado_deja_bandera_sin_alterar_los_importes(db: AsyncSession) -> None:
    """Las columnas salen de los **nodos**, así que un encabezado que diga otra cosa no cambia
    el costo del centro: se reporta y ya. Sin la bandera, B-06 y B-02 darían cifras distintas
    del mismo periodo sin una sola señal."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="9999.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert _valor(resultado, 0, "Total percepciones") == Decimal("1000.00")
    descuadres = _de_clave(resultado, "TOTALES_DESCUADRADOS")
    assert len(descuadres) == 1
    assert descuadres[0].ambito == "uuid:11111111-1111-1111-1111-111111111111"
    assert "total_percepciones" in descuadres[0].mensaje


async def test_el_cfdi_de_nomina_sin_normalizar_no_desaparece_en_silencio(db: AsyncSession) -> None:
    """§9 del diseño: un tipo `N` sin fila en `nomina` no puede salir en la hoja Datos, pero
    tiene que salir en Banderas — si no, el centro de costo sale corto sin rastro."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await factories.crear_comprobante(
        db,
        empresa_id=empresa.empresa_id,
        uuid="99999999-9999-9999-9999-999999999999",
        rfc_emisor=RFC_EMPRESA,
        tipo_comprobante="N",
        fecha_emision=None,
    )
    await db.commit()

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert "SIN_NORMALIZAR" in _claves(resultado)


async def test_un_universo_sano_no_reporta_ningun_cfdi_perdido(db: AsyncSession) -> None:
    """Gemela negativa de la anterior. Sin ella, una consulta que reportara *todos* los CFDI de
    nómina como perdidos pasaría igual de verde, y la hoja `Banderas` que el patrón usa para
    saber qué le falta diría que le falta todo."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert "SIN_NORMALIZAR" not in _claves(resultado)
    assert "COMPLEMENTO_AUSENTE" not in _claves(resultado)


# --------------------------------------------------------------------------------------
# 7-bis. Las notas: lo que tiene que llegar al libro, no solo al docstring
# --------------------------------------------------------------------------------------


def _contexto() -> ContextoInforme:
    return ContextoInforme(
        clave=b06.CLAVE,
        nombre=b06.NOMBRE,
        usuario="dgarcia@planjuarez.org",
        generado_en=datetime(2026, 8, 7, 9, 0, 0),
        parametros={"fecha_desde": str(_DESDE), "fecha_hasta": str(_HASTA), "enmascarar_datos_personales": True},
        etl_version=1,
    )


async def test_las_notas_llegan_a_la_hoja_parametros_del_libro(db: AsyncSession) -> None:
    """No basta con que `consultar` las devuelva: quien recibe el Excel por correo no lee el
    código ni el objeto de resultado. Se genera el libro de verdad y se lee la hoja."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    await insertar_nomina(
        db,
        empresa_id=empresa.empresa_id,
        uuid="11111111-1111-1111-1111-111111111111",
        fecha_pago=date(2026, 6, 30),
        departamento=_EDIFICIOS,
        percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
        total_percepciones="1000.00",
        total="1000.00",
    )
    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    libro = load_workbook(io.BytesIO(excel.escribir_libro(resultado, _contexto())))
    hoja = libro["Parámetros"]
    texto = "\n".join(str(c.value) for fila in hoja.iter_rows() for c in fila if c.value is not None)

    # Duda 4: no sumar «Núm. de empleados» entre centros (B-06.R3).
    assert "no la sumes" in texto
    assert "exceder la plantilla real" in texto
    # Duda 3: periodo = mes, y un mes con dos fechas de pago agrega dos quincenas. Hoy mes y
    # quincena coinciden por accidente del calendario, y deja de ser así con la próxima descarga.
    assert "es el MES de la fecha de pago" in texto
    assert "dos quincenas" in texto
    # El rango que parte un mes: porcentajes sobre un mes truncado, indistinguibles.
    assert "mes truncado" in texto
    # Y por qué el subsidio está en su propia columna sin salir del costo bruto.
    assert "subsidio para el empleo" in texto.lower()
    assert "recupera contra el ISR" in texto


async def test_las_notas_viajan_aunque_no_haya_filas(db: AsyncSession) -> None:
    """Un libro sin filas circula igual, y una nota califica cómo leer la columna, no esta
    corrida: no puede depender de que haya datos (mismo criterio que el rótulo de B-08)."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)

    resultado = await b06.consultar(db, empresa.empresa_id, _p())

    assert resultado.filas == []
    assert resultado.notas, "las notas no pueden depender de que haya filas"
    assert any("es el MES de la fecha de pago" in nota for nota in resultado.notas)
    assert any("no la sumes" in nota for nota in resultado.notas)


async def test_el_orden_de_las_filas_es_determinista(db: AsyncSession) -> None:
    """Por `(ejercicio, periodo, centro de costo)`, y con `detalle_empleado` además por RFC.
    Dos corridas del mismo informe tienen que dar el mismo archivo."""
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    casos = (
        ("11111111-1111-1111-1111-111111111111", date(2026, 7, 31), _SOCIAL, "XEXX010101000"),
        ("22222222-2222-2222-2222-222222222222", date(2026, 6, 30), _SOCIAL, "XAXX010101000"),
        ("33333333-3333-3333-3333-333333333333", date(2026, 6, 30), _EDIFICIOS, "XEXX010101000"),
        ("44444444-4444-4444-4444-444444444444", date(2026, 6, 30), _EDIFICIOS, "XAXX010101000"),
    )
    for uuid_cfdi, pago, departamento, rfc in casos:
        await insertar_nomina(
            db,
            empresa_id=empresa.empresa_id,
            uuid=uuid_cfdi,
            rfc_receptor=rfc,
            fecha_pago=pago,
            departamento=departamento,
            percepciones=[("001", "001", "Sueldo", "1000.00", "0.00")],
            total_percepciones="1000.00",
            total="1000.00",
        )

    primera = await b06.consultar(db, empresa.empresa_id, _p())
    segunda = await b06.consultar(db, empresa.empresa_id, _p())
    detallada = await b06.consultar(db, empresa.empresa_id, _p(detalle_empleado=True))

    assert primera.filas == segunda.filas
    assert [(_valor(primera, i, "Periodo"), _valor(primera, i, "Centro de costo")) for i in range(3)] == [
        (6, _EDIFICIOS),
        (6, _SOCIAL),
        (7, _SOCIAL),
    ]
    assert [
        (_valor(detallada, i, "Periodo"), _valor(detallada, i, "Centro de costo"), _valor(detallada, i, "RFC empleado"))
        for i in range(4)
    ] == [
        (6, _EDIFICIOS, "XAXX010101000"),
        (6, _EDIFICIOS, "XEXX010101000"),
        (6, _SOCIAL, "XAXX010101000"),
        (7, _SOCIAL, "XEXX010101000"),
    ]


# --------------------------------------------------------------------------------------
# 8. Registro
# --------------------------------------------------------------------------------------


def test_b06_esta_registrado_y_declara_su_tipo_de_comprobante() -> None:
    """Sin `TIPOS_COMPROBANTE` el pre-vuelo del ETL reprocesa el histórico completo de la
    empresa dentro de la tarea del informe."""
    definicion = registro.obtener("B-06")
    assert definicion is b06
    assert b06.TIPOS_COMPROBANTE == ("N",)
    assert b06.GRUPO == "B"
    esquema = definicion.Parametros.model_json_schema()["properties"]
    assert set(esquema) == {
        "fecha_desde",
        "fecha_hasta",
        "detalle_empleado",
        "incluir_cancelados",
        "enmascarar_datos_personales",
    }
    # La columna 15 (costo patronal estimado) queda fuera de alcance: no es derivable del
    # CFDI, así que no se declara ni la columna ni un parámetro para ella.
    titulos = [c.titulo for c in b06._columnas(detalle_empleado=True)]
    assert not any("patronal" in titulo.lower() for titulo in titulos)


def test_la_resolucion_del_centro_de_costo_no_normaliza_el_texto() -> None:
    """Prueba unitaria de la cascada, sin base: es el único sitio donde se decide el
    agrupamiento y su contrato tiene que poder leerse de un vistazo."""
    mapa = {"EDIFICIOS": "CC-EDIF"}

    assert b06.resolver_centro_de_costo("EDIFICIOS", mapa) == b06._CentroResuelto(
        "CC-EDIF", b06.NivelDeResolucion.MAPEO
    )
    assert b06.resolver_centro_de_costo("edificios", mapa) == b06._CentroResuelto(
        "edificios", b06.NivelDeResolucion.TEXTO_CRUDO
    )
    assert b06.resolver_centro_de_costo(" EDIFICIOS", mapa) == b06._CentroResuelto(
        " EDIFICIOS", b06.NivelDeResolucion.TEXTO_CRUDO
    )
    assert b06.resolver_centro_de_costo(None, mapa) == b06._CentroResuelto(
        b06.SIN_DEPARTAMENTO, b06.NivelDeResolucion.AUSENTE
    )
    assert b06.resolver_centro_de_costo("   ", mapa) == b06._CentroResuelto(
        b06.SIN_DEPARTAMENTO, b06.NivelDeResolucion.AUSENTE
    )
