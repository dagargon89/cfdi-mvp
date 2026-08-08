"""B-08 · Provisión de pasivo laboral.

Lo que estas pruebas fijan, en orden de importancia — y el orden es el del daño, no el del
código, porque la cifra de este informe puede acabar reconocida en los estados financieros:

1. **El informe no se genera con datos que no alcanzan.** Sin la clasificación completa de
   percepciones no se puede distinguir "no se pagó aguinaldo" de "sí se pagó y no sé cuál
   concepto es"; sin días de aguinaldo configurados no hay devengado que calcular. En los dos
   casos el aviso dice **qué** falta, y en el segundo el informe **no supone 15 días**: el
   mínimo legal es un piso y suponerlo subestimaría la provisión.
2. **La completitud se mide solo sobre percepciones.** Una deducción sin clasificar NO bloquea
   el informe. Con el criterio contrario, la empresa real —tres percepciones, siete
   deducciones, dos departamentos— no podría generarlo nunca.
3. **La provisión nunca sale negativa** y **la fila declara de qué fuente salió el salario
   diario base**, porque la vía del SDI sobreestima y nadie puede saberlo mirando la cifra.
4. **Cada regla con su gemela negativa.** Una bandera que dispara siempre no protege nada.

Lo que agregó la ronda de corrección 1
----------------------------------------
5. **El promedio de B-08.R1 arma numerador y denominador del MISMO conjunto de comprobantes.**
   Era un Critical: un periodo elegido que aportaba días sin importe `001` subestimaba el salario
   un 33% y uno con `num_dias_pagados` nulo lo inflaba un 50%, en los dos casos con la fila
   declarando la fuente confiable y sin una sola bandera. Se prueba **en los dos sentidos**, con
   la cifra del defecto como aserción negativa explícita.
6. **El libro dice cuándo la cifra es una proyección** (`COBERTURA_INCOMPLETA_AL_CORTE`) y **por
   dónde empezar a conciliar las bajas** (columna «Último CFDI del ejercicio» y
   `POSIBLE_BAJA_EN_EL_EJERCICIO`), en vez de dejar los dos residuos solo declarados.
7. **Las tres banderas y las tres notas que no tenían prueba** sí la tienen:
   `SIN_SALARIO_DIARIO_BASE`, `FECHA_INICIO_POSTERIOR_AL_CORTE`, el conteo de filas sin provisión
   calculable, y los rótulos del art. 78 LFT y de "lo pagado se identifica por su fecha de pago".

Ninguna prueba de este archivo se dio por buena sin romper a propósito lo que dice proteger y
comprobar que **FALLA** (ver las trampas de la evidencia por mutación anotadas en
`tests/test_cli_configuracion.py`, y la cuarta: una mutación que resulta ser un no-op se ve
igual que una prueba muerta).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import update
from sqlalchemy.ext.asyncio import AsyncSession

from app.informes import b08_pasivo_laboral as b08
from app.informes import excel, registro
from app.informes.base import ContextoInforme
from app.models.configuracion_fiscal import ConfiguracionEmpresa, MapConceptoProvision, TablaVacaciones
from app.models.enums import CategoriaProvision, EstatusCfdi
from app.models.nomina import Nomina, NominaReceptor
from app.services import configuracion_fiscal as cfg
from tests import factories
from tests.helpers_nomina import insertar_nomina

RFC_EMPRESA = "CHL960913IX9"
RFC_EMPLEADO = "XAXX010101000"
RFC_OTRO = "XEXX010101000"

_EJERCICIO = 2026
_CORTE = date(2026, 12, 31)
_INICIO_2020 = date(2020, 1, 1)

# Los tres periodos ordinarios del escenario base, que llegan hasta el corte para que
# `COBERTURA_INCOMPLETA_AL_CORTE` no contamine todas las pruebas (ver `_tres_quincenas`).
_PAGOS_ORDINARIOS: tuple[date, ...] = (date(2026, 10, 31), date(2026, 11, 30), date(2026, 12, 31))

# Tipo de `c_TipoPercepcion` distinto de 001 con el que se construye un periodo que aporta días
# sin sueldo: "Incapacidad" (014), el caso real que subestimaba el salario diario base un 33%.
_TIPO_INCAPACIDAD = "014"

# Progresión del art. 76 LFT (semilla `config/fiscal/tabla_vacaciones.yaml`). Se siembran los
# renglones que las pruebas usan, por el modelo: la carga del YAML ya la prueba
# `tests/test_semillas_fiscales.py` y aquí lo que se ejercita es el informe.
_TABLA_VACACIONES: tuple[tuple[int, int], ...] = ((1, 12), (2, 14), (3, 16), (4, 18), (5, 20), (6, 22), (11, 24))


# --------------------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------------------


async def _empresa(db: AsyncSession) -> int:
    empresa = await factories.crear_empresa(db, rfc=RFC_EMPRESA)
    return empresa.empresa_id


async def _vacaciones(db: AsyncSession) -> None:
    for anios, dias in _TABLA_VACACIONES:
        db.add(TablaVacaciones(anios_antiguedad=anios, dias=dias))
    await db.commit()


async def _config(db: AsyncSession, empresa_id: int, *, dias: int | None = 15, factor: str | None = "0.25") -> None:
    db.add(
        ConfiguracionEmpresa(
            empresa_id=empresa_id,
            dias_aguinaldo=dias,
            factor_prima_vacacional=Decimal(factor) if factor is not None else None,
        )
    )
    await db.commit()


async def _clasificar(db: AsyncSession, empresa_id: int, *conceptos: tuple[str, str, str, CategoriaProvision]) -> None:
    """Captura `map_concepto_provision`. Las tuplas son `(naturaleza, tipo, clave, categoría)`."""
    for naturaleza, tipo, clave, categoria in conceptos:
        db.add(
            MapConceptoProvision(
                empresa_id=empresa_id, naturaleza=naturaleza, tipo=tipo, clave=clave, categoria=categoria
            )
        )
    await db.commit()


def _p(**extra: Any) -> b08.Parametros:
    return b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=_CORTE, **extra)


async def _borrar_sdi(db: AsyncSession, comprobante_id: int) -> None:
    """Deja `nomina_receptor.salario_diario_integrado` en NULL, que es como llega un complemento
    que no lo declara.

    Va por `UPDATE` y no por un parámetro de `tests/helpers_nomina.py` a propósito: ese helper lo
    comparten los nueve informes del grupo B y su firma declara `sdi: str = "607.34"`, que no
    admite `None`. Cambiarla para cubrir un caso de B-08 tocaría las pruebas de B-01, B-02, B-04,
    B-05, B-07 y B-10. Mismo criterio y misma técnica que `test_informe_b06._borrar_departamento`.
    """
    await db.execute(
        update(NominaReceptor)
        .where(NominaReceptor.comprobante_id == comprobante_id)
        .values(salario_diario_integrado=None)
    )
    await db.commit()
    # `expire_on_commit=False` en la sesión de pruebas: sin expulsar los objetos, SQLAlchemy
    # respondería con el valor viejo de su mapa de identidad y la prueba mediría su memoria en vez
    # del informe. Se expulsa en vez de expirar para no disparar IO perezosa fuera del greenlet.
    db.expunge_all()


async def _borrar_dias_pagados(db: AsyncSession, comprobante_id: int) -> None:
    """Deja `nomina.num_dias_pagados` en NULL — es `nullable=True` en el modelo, así que basta un
    ETL parcial. Por `UPDATE` por la misma razón que `_borrar_sdi`."""
    await db.execute(update(Nomina).where(Nomina.comprobante_id == comprobante_id).values(num_dias_pagados=None))
    await db.commit()
    db.expunge_all()


def _columna(resultado: Any, titulo: str) -> int:
    titulos = [c.titulo for c in resultado.columnas]
    assert titulo in titulos, f"falta la columna {titulo!r}; hay {titulos}"
    return titulos.index(titulo)


def _valor(resultado: Any, titulo: str, fila: int = 0) -> Any:
    return resultado.filas[fila][_columna(resultado, titulo)]


def _claves(resultado: Any) -> list[str]:
    return [b.clave for b in resultado.banderas]


def _de_clave(resultado: Any, clave: str) -> list[Any]:
    return [b for b in resultado.banderas if b.clave == clave]


async def _quincena(
    db: AsyncSession,
    empresa_id: int,
    *,
    uuid: str,
    fecha_pago: date,
    rfc_receptor: str = RFC_EMPLEADO,
    sueldo: str = "7500.00",
    dias: str = "15.000",
    tipo_nomina: str = "O",
    percepciones: list[tuple[str, str, str, str, str]] | None = None,
    inicio: date | None = _INICIO_2020,
    **extra: Any,
) -> int:
    """Una quincena ordinaria de sueldo, que es el insumo de B-08.R1 vía 1."""
    if percepciones is None:
        percepciones = [("001", "001", "Sueldo", sueldo, "0.00")]
    total = sum(Decimal(g) + Decimal(e) for _t, _c, _n, g, e in percepciones)
    return await insertar_nomina(
        db,
        empresa_id=empresa_id,
        uuid=uuid,
        rfc_receptor=rfc_receptor,
        fecha_pago=fecha_pago,
        fecha_final_pago=fecha_pago,
        tipo_nomina=tipo_nomina,
        dias=dias,
        percepciones=percepciones,
        deducciones=[],
        total_percepciones=str(total),
        total_deducciones="0.00",
        total=str(total),
        fecha_inicio_rel_laboral=inicio,
        **extra,
    )


async def _tres_quincenas(
    db: AsyncSession,
    empresa_id: int,
    *,
    prefijo: str = "1111111",
    pagos: tuple[date, ...] = _PAGOS_ORDINARIOS,
    **extra: Any,
) -> list[int]:
    """Tres periodos ordinarios de 15 días y 7500 de sueldo: salario diario base = 500.00
    exacto (22500 / 45), que hace que todas las cifras derivadas sean exactas.

    Los tres pagos llegan **hasta el corte** (octubre, noviembre, diciembre) a propósito: con
    datos que se detuvieran en junio, `COBERTURA_INCOMPLETA_AL_CORTE` dispararía en casi todas
    las pruebas y el escenario base dejaría de ser un escenario limpio contra el que comparar.
    La cobertura incompleta tiene su propia prueba, con sus propias fechas."""
    return [
        await _quincena(db, empresa_id, uuid=f"{prefijo}{i}-1111-1111-1111-111111111111", fecha_pago=pago, **extra)
        for i, pago in enumerate(pagos, start=1)
    ]


async def _escenario_completo(db: AsyncSession) -> int:
    """La empresa lista para generar: tres quincenas, tabla de vacaciones, política laboral y
    la única percepción del universo clasificada."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    return eid


# --------------------------------------------------------------------------------------
# 1. Las dos puertas: cuándo NO se genera
# --------------------------------------------------------------------------------------


async def test_sin_clasificar_no_se_genera_y_el_aviso_dice_cuales_faltan(db: AsyncSession) -> None:
    """La puerta 1. Con una percepción sin categoría es imposible distinguir «no se pagó
    aguinaldo» de «sí se pagó y no sé en cuál concepto viene», y las dos hipótesis dan
    provisiones distintas: el informe no se genera. El aviso NO puede ser genérico — quien lo
    lee tiene que saber a qué renglón ir."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _vacaciones(db)
    await _config(db, eid)

    resultado = await b08.consultar(db, eid, _p())

    assert resultado.filas == []
    assert resultado.aviso is not None
    assert "P/001/001" in resultado.aviso, "el aviso tiene que nombrar el concepto que falta"
    assert "CLASIFICACION_INCOMPLETA" in _claves(resultado)
    assert "P/001/001" in _de_clave(resultado, "CLASIFICACION_INCOMPLETA")[0].mensaje


async def test_clasificacion_completa_con_no_aplica_si_genera(db: AsyncSession) -> None:
    """La gemela negativa de la anterior, y la razón de ser de `NO_APLICA`: cuando **todas**
    las percepciones están clasificadas —incluidas las que no son ninguna de las tres—,
    «aguinaldo pagado = 0» pasa a ser un hecho conocido y el informe sale con la provisión
    igual al devengado completo."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1
    assert "CLASIFICACION_INCOMPLETA" not in _claves(resultado)
    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("0")
    assert _valor(resultado, "Provisión de aguinaldo") == Decimal("7500")


async def test_una_deduccion_sin_clasificar_no_bloquea_el_informe(db: AsyncSession) -> None:
    """**La corrección que decide si el informe sirve.** El aguinaldo no se le descuenta a
    nadie, así que una deducción no puede ser aguinaldo pagado y exigir su clasificación
    dejaría el informe bloqueado para siempre por una razón sin sentido: la empresa real tiene
    tres percepciones frente a siete deducciones y dos departamentos."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    for i, pago in enumerate(_PAGOS_ORDINARIOS, start=1):
        await insertar_nomina(
            db,
            empresa_id=eid,
            uuid=f"1111111{i}-1111-1111-1111-111111111111",
            fecha_pago=pago,
            fecha_final_pago=pago,
            dias="15.000",
            percepciones=[("001", "001", "Sueldo", "7500.00", "0.00")],
            # Siete deducciones sin clasificar, como en la empresa real. Ninguna se puede
            # clasificar como aguinaldo, vacaciones ni prima vacacional.
            deducciones=[(tipo, clave, nombre, "10.00") for tipo, clave, nombre in (
                ("001", "051", "IMSS"),
                ("002", "045", "ISR"),
                ("004", "052", "Fondo de ahorro"),
                ("006", "053", "Descuento"),
                ("007", "054", "Pensión"),
                ("009", "055", "Infonavit"),
                ("010", "056", "Préstamo"),
            )],
            total_percepciones="7500.00",
            total_deducciones="70.00",
            total="7430.00",
            fecha_inicio_rel_laboral=_INICIO_2020,
        )
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1, f"las deducciones bloquearon el informe: {resultado.aviso}"
    assert "CLASIFICACION_INCOMPLETA" not in _claves(resultado)


async def test_sin_dias_de_aguinaldo_no_se_genera_y_no_supone_quince(db: AsyncSession) -> None:
    """La puerta 2, y la parte que más importa: **no hay 15 por omisión**. El mínimo legal del
    art. 87 LFT es un piso y muchas organizaciones dan más; suponerlo subestimaría la
    provisión, que es el error que un auditor no perdona."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _vacaciones(db)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    # `configuracion_empresa` existe pero con los dos campos en NULL, que es como nace.
    await _config(db, eid, dias=None, factor=None)

    resultado = await b08.consultar(db, eid, _p())

    assert resultado.filas == [], "no puede haber filas sin días de aguinaldo configurados"
    assert resultado.aviso is not None
    assert "días de aguinaldo" in resultado.aviso
    assert "factor de prima vacacional" in resultado.aviso
    assert "FALTA_CONFIGURACION_DE_PROVISION" in _claves(resultado)
    # La gemela del "no supone 15": con 15 días y salario 500 el devengado sería 7500. Que no
    # haya filas ya lo impide, pero la aserción explícita es la que muere si alguien mete un
    # default silencioso y a la vez deja que el informe se genere.
    assert not any(Decimal("7500") in fila for fila in resultado.filas)


async def test_sin_configuracion_de_empresa_tampoco_se_genera(db: AsyncSession) -> None:
    """La variante de la anterior que sí ocurre en vivo: `configuracion_empresa` **vacía**, sin
    fila ninguna, que es el estado de la empresa 11 hoy."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _vacaciones(db)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))

    resultado = await b08.consultar(db, eid, _p())

    assert resultado.filas == []
    assert "FALTA_CONFIGURACION_DE_PROVISION" in _claves(resultado)


async def test_las_dos_puertas_se_reportan_juntas(db: AsyncSession) -> None:
    """Sin esto, quien configura tapa un hueco, vuelve a generar y descubre el otro. Es la ruta
    que más se va a ver en vivo: la empresa real tiene las dos cosas pendientes."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _vacaciones(db)

    resultado = await b08.consultar(db, eid, _p())

    assert resultado.filas == []
    assert set(_claves(resultado)) >= {"CLASIFICACION_INCOMPLETA", "FALTA_CONFIGURACION_DE_PROVISION"}


async def test_el_parametro_explicito_gana_sobre_la_configuracion(db: AsyncSession) -> None:
    """La corrida puntual con otro supuesto: «¿cuánto sería la provisión si diéramos 30 días?».
    La configuración dice 15; el parámetro dice 30 y manda."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p(dias_aguinaldo=30))

    assert _valor(resultado, "Días de aguinaldo") == 30
    assert _valor(resultado, "Aguinaldo devengado") == Decimal("15000")


async def test_el_parametro_de_prima_gana_sobre_la_configuracion(db: AsyncSession) -> None:
    eid = await _escenario_completo(db)

    con_config = await b08.consultar(db, eid, _p())
    con_parametro = await b08.consultar(db, eid, _p(factor_prima_vacacional=Decimal("0.50")))

    # 22 días de vacaciones × 500 de salario × factor.
    assert _valor(con_config, "Prima vacacional devengada") == Decimal("2750")
    assert _valor(con_parametro, "Prima vacacional devengada") == Decimal("5500")


# --------------------------------------------------------------------------------------
# 2. B-08.R1: el salario diario base y su fuente
# --------------------------------------------------------------------------------------


async def test_el_salario_sale_de_los_ultimos_tres_periodos_ordinarios(db: AsyncSession) -> None:
    """B-08.R1 vía 1. El cuarto periodo, más antiguo y con otro sueldo, **no** entra: con los
    cuatro el promedio sería 425 y con los tres últimos es 500."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    # El más antiguo, con un sueldo distinto: si entrara, 25500/60 = 425.
    await _quincena(db, eid, uuid="00000000-0000-0000-0000-000000000001", fecha_pago=date(2026, 3, 31), sueldo="3000.00")
    await _tres_quincenas(db, eid)

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Salario diario base") == Decimal("500")
    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_CFDI
    assert _valor(resultado, "Salario diario base") != Decimal("425")


async def test_una_nomina_extraordinaria_no_entra_en_el_salario_base(db: AsyncSession) -> None:
    """La gemela negativa: el aguinaldo se paga en una nómina **extraordinaria**, y meterlo en
    el promedio del salario corriente inflaría justo lo que este informe provisiona.

    **La nómina extraordinaria de este fixture trae también una percepción tipo 001** (un sueldo
    pendiente, como en un finiquito), y no es un adorno: la verificación por mutación de la ronda
    de corrección 1 mostró que sin eso la prueba había dejado de proteger el filtro. Con solo el
    aguinaldo (tipo 002), quitar `tipo_nomina == 'O'` ya no cambiaba el resultado, porque el
    arreglo del Critical **descarta** del promedio cualquier periodo sin importe 001. La defensa
    en profundidad es buena; una prueba que la confunde con la defensa que dice proteger, no.
    Con un tipo 001 de 30 000 en 30 días, incluir la extraordinaria daría 45000/60 = 750."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="44444444-4444-4444-4444-444444444444",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        dias="30.000",
        percepciones=[
            ("002", "AGU", "Aguinaldo", "5000.00", "0.00"),
            ("001", "001", "Sueldo pendiente", "30000.00", "0.00"),
        ],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Salario diario base") == Decimal("500")
    assert _valor(resultado, "Salario diario base") != Decimal("750"), "la extraordinaria entró al promedio"
    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_CFDI
    # Y la otra mitad del criterio: la extraordinaria SÍ cuenta como aguinaldo pagado. El filtro a
    # ordinarias vive dentro de B-08.R1, no en el universo del informe (ver `_ParametrosUniverso`).
    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("5000")
    assert "PERIODO_SIN_SUELDO_EN_EL_PROMEDIO" not in _claves(resultado)


async def test_sin_percepciones_001_cae_al_sdi_y_la_fila_lo_declara(db: AsyncSession) -> None:
    """B-08.R1 vía 2. El SDI ya integra la parte proporcional de aguinaldo y prima, así que
    usarlo **sobreestima** la provisión: la fila tiene que decirlo, porque nadie puede saberlo
    mirando la cifra, y la bandera tiene que avisar del conteo."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "029", "029", CategoriaProvision.NO_APLICA))
    for i, pago in enumerate(_PAGOS_ORDINARIOS, start=1):
        await _quincena(
            db,
            eid,
            uuid=f"5555555{i}-5555-5555-5555-555555555555",
            fecha_pago=pago,
            percepciones=[("029", "029", "Asimilado", "7500.00", "0.00")],
        )

    resultado = await b08.consultar(db, eid, _p())

    # El SDI que siembra `helpers_nomina` es 607.34.
    assert _valor(resultado, "Salario diario base") == Decimal("607.34")
    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_SDI
    assert "SDI" in _valor(resultado, "Fuente del salario diario base")
    assert "SALARIO_DE_ULTIMO_RECURSO" in _claves(resultado)
    assert "sobreestim" in _de_clave(resultado, "SALARIO_DE_ULTIMO_RECURSO")[0].mensaje.lower()


async def test_con_percepciones_001_la_fila_no_declara_el_ultimo_recurso(db: AsyncSession) -> None:
    """Gemela negativa de la anterior: una bandera que dispara siempre no protege nada."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_CFDI
    assert "SALARIO_DE_ULTIMO_RECURSO" not in _claves(resultado)


async def test_sin_sdi_ni_percepciones_001_las_columnas_van_vacias_no_en_cero(db: AsyncSession) -> None:
    """La tercera rama de B-08.R1, que no tenía prueba (ronda de corrección 1): sin ninguna de
    las dos vías el salario no se puede determinar, y el docstring promete «vacías, nunca en
    cero» — un cero afirmaría que a esta persona no se le debe nada.

    El SDI se borra con un `UPDATE` porque `tests/helpers_nomina.py` lo comparten los nueve
    informes del grupo B y su firma no admite `None` (ver `_borrar_sdi`)."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "029", "029", CategoriaProvision.NO_APLICA))
    cids = await _tres_quincenas(
        db, eid, prefijo="7171717", percepciones=[("029", "029", "Asimilado", "7500.00", "0.00")]
    )
    for cid in cids:
        await _borrar_sdi(db, cid)

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1
    assert _valor(resultado, "Salario diario base") is None
    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_AUSENTE
    assert _valor(resultado, "Aguinaldo devengado") is None
    assert _valor(resultado, "Provisión de aguinaldo") is None
    assert _valor(resultado, "Provisión total") is None
    assert "SIN_SALARIO_DIARIO_BASE" in _claves(resultado)
    assert "SALARIO_DE_ULTIMO_RECURSO" not in _claves(resultado)


# --------------------------------------------------------------------------------------
# 2b. El defecto Critical de la ronda 1: el promedio mezclaba dos universos
# --------------------------------------------------------------------------------------


async def test_un_periodo_sin_sueldo_se_descarta_del_promedio_y_deja_bandera(db: AsyncSession) -> None:
    """**Critical de la ronda 1, sentido «subestima».** Un periodo que aporta 15 días al
    denominador y cero pesos al numerador daba 15000/45 = **333.33** (−33% sobre el salario
    real de 500.00) y la fila seguía declarando la fuente confiable, sin una sola bandera.

    Ahora el periodo se descarta de **los dos** lados —15000/30 = 500.00— y deja bandera: un
    promedio sobre menos periodos de los previstos es una cifra que alguien tiene que mirar.
    Dispara con una incapacidad que cubre el periodo completo, que es el caso de este fixture."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", _TIPO_INCAPACIDAD, "INC", CategoriaProvision.NO_APLICA),
    )
    await _tres_quincenas(db, eid, pagos=_PAGOS_ORDINARIOS[:2])
    # Mes completo de incapacidad: 15 días pagados, ninguna percepción tipo 001.
    await _quincena(
        db,
        eid,
        uuid="14141414-1414-1414-1414-141414141414",
        fecha_pago=_PAGOS_ORDINARIOS[2],
        percepciones=[(_TIPO_INCAPACIDAD, "INC", "Incapacidad", "0.00", "7500.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Salario diario base") == Decimal("500")
    assert _valor(resultado, "Salario diario base") != Decimal("15000") / Decimal("45"), "el defecto Critical volvió"
    assert _valor(resultado, "Fuente del salario diario base") == b08.FUENTE_CFDI
    assert "PERIODO_SIN_SUELDO_EN_EL_PROMEDIO" in _claves(resultado)
    mensaje = _de_clave(resultado, "PERIODO_SIN_SUELDO_EN_EL_PROMEDIO")[0].mensaje
    assert "Periodos descartados del promedio: 1" in mensaje
    assert "Empleados afectados: 1" in mensaje


async def test_un_periodo_sin_dias_pagados_se_descarta_del_promedio_y_deja_bandera(db: AsyncSession) -> None:
    """**Critical de la ronda 1, sentido «sobrestima».** Con `num_dias_pagados` nulo —es
    `nullable=True` en el modelo, así que basta un ETL parcial— el importe entraba al numerador
    y `_dec(None) → 0` al denominador: 22500/30 = **750.00** (+50%), con la provisión total
    subiendo de 21 250 a 31 875 y **cero banderas**.

    Ahora ese periodo se descarta de los dos lados: 15000/30 = 500.00, con bandera."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    cids = await _tres_quincenas(db, eid)
    await _borrar_dias_pagados(db, cids[-1])

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Salario diario base") == Decimal("500")
    assert _valor(resultado, "Salario diario base") != Decimal("750"), "el defecto Critical volvió"
    assert _valor(resultado, "Provisión total") == Decimal("21250")
    assert "PERIODO_SIN_DIAS_EN_EL_PROMEDIO" in _claves(resultado)
    assert "Periodos descartados del promedio: 1" in _de_clave(resultado, "PERIODO_SIN_DIAS_EN_EL_PROMEDIO")[0].mensaje


async def test_con_los_tres_periodos_completos_no_se_descarta_ninguno(db: AsyncSession) -> None:
    """Gemela negativa de las dos anteriores: dos banderas que dispararan siempre serían ruido
    y enterrarían los hallazgos de la misma hoja."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert "PERIODO_SIN_SUELDO_EN_EL_PROMEDIO" not in _claves(resultado)
    assert "PERIODO_SIN_DIAS_EN_EL_PROMEDIO" not in _claves(resultado)
    assert _valor(resultado, "Salario diario base") == Decimal("500")


# --------------------------------------------------------------------------------------
# 3. Los devengos y la provisión
# --------------------------------------------------------------------------------------


async def test_el_aguinaldo_devengado_es_proporcional_a_los_dias_trabajados(db: AsyncSession) -> None:
    """Art. 87 LFT: proporcional al tiempo trabajado en el año. Quien entró el 1 de julio no
    devenga el aguinaldo completo, y suponer que sí duplicaría su provisión."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid, inicio=date(2026, 7, 1))

    resultado = await b08.consultar(db, eid, _p())

    # Del 1 de julio al 31 de diciembre: 184 días de los 365 del ejercicio.
    esperado = Decimal("500") * 15 * Decimal(184) / Decimal(365)
    assert _valor(resultado, "Aguinaldo devengado") == esperado
    assert _valor(resultado, "Aguinaldo devengado") != Decimal("7500"), "no puede devengar el año completo"


async def test_el_aguinaldo_pagado_se_resta_del_devengado(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="66666666-6666-6666-6666-666666666666",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", "AGU", "Aguinaldo", "3000.00", "2000.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    # Gravado + exento: quedarse con el gravado subestimaría lo pagado, o sea sobreestimaría
    # la provisión (el aguinaldo trae tramo exento, art. 93 fr. XIV LISR).
    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("5000")
    assert _valor(resultado, "Provisión de aguinaldo") == Decimal("2500")


async def test_la_provision_nunca_es_negativa(db: AsyncSession) -> None:
    """Un pagado mayor que el devengado es un anticipo o un supuesto equivocado, no un activo:
    en negativo restaría de la provisión de los demás empleados al sumar la columna."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="77777777-7777-7777-7777-777777777777",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", "AGU", "Aguinaldo", "20000.00", "0.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Aguinaldo devengado") == Decimal("7500")
    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("20000")
    assert _valor(resultado, "Provisión de aguinaldo") == Decimal("0")
    assert _valor(resultado, "Provisión total") >= Decimal("0")


async def test_los_dias_de_vacaciones_salen_de_la_tabla_por_antiguedad(db: AsyncSession) -> None:
    """Art. 76 LFT. Dos empleados con la misma nómina y distinta antigüedad tienen derechos
    distintos, y el número sale de `tabla_vacaciones`, no del código."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)  # RFC_EMPLEADO, alta en 2020 → 6 años cumplidos → 22 días
    for i, pago in enumerate(_PAGOS_ORDINARIOS, start=1):
        await _quincena(
            db,
            eid,
            uuid=f"8888888{i}-8888-8888-8888-888888888888",
            fecha_pago=pago,
            rfc_receptor=RFC_OTRO,
            inicio=date(2024, 1, 1),  # 2 años cumplidos → 14 días
        )

    resultado = await b08.consultar(db, eid, _p())

    por_rfc = {fila[_columna(resultado, "RFC empleado")]: fila for fila in resultado.filas}
    indice_dias = _columna(resultado, "Días de vacaciones del año (art. 76 LFT)")
    indice_anios = _columna(resultado, "Antigüedad al corte (años cumplidos)")
    assert por_rfc[RFC_EMPLEADO][indice_anios] == 6
    assert por_rfc[RFC_EMPLEADO][indice_dias] == 22
    assert por_rfc[RFC_OTRO][indice_anios] == 2
    assert por_rfc[RFC_OTRO][indice_dias] == 14


async def test_las_vacaciones_pagadas_bajan_los_dias_pendientes(db: AsyncSession) -> None:
    """Los días pendientes son una **estimación** por devengo proporcional: el saldo real lo
    lleva recursos humanos. Lo único observable es el importe pagado, convertido a días con el
    salario diario base."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "001", "019", CategoriaProvision.VACACIONES),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="99999999-9999-9999-9999-999999999999",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("001", "019", "Vacaciones a tiempo", "2000.00", "0.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Vacaciones pagadas en el ejercicio") == Decimal("2000")
    # 22 días devengados − 2000/500 = 4 días pagados.
    assert _valor(resultado, "Días de vacaciones pendientes (estimados)") == Decimal("18")
    assert _valor(resultado, "Provisión de vacaciones") == Decimal("9000")


async def test_sin_vacaciones_pagadas_los_dias_pendientes_son_el_devengo_completo(db: AsyncSession) -> None:
    """Gemela negativa de la anterior."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Vacaciones pagadas en el ejercicio") == Decimal("0")
    assert _valor(resultado, "Días de vacaciones pendientes (estimados)") == Decimal("22")
    assert _valor(resultado, "Provisión de vacaciones") == Decimal("11000")


async def test_la_prima_vacacional_aplica_el_factor_y_la_provision_total_suma_las_tres(db: AsyncSession) -> None:
    """Art. 80 LFT. Y la hoja tiene que cuadrar por suma: un total con un sumando que no
    aparece en ninguna columna no es auditable."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Prima vacacional devengada") == Decimal("2750")  # 22 × 500 × 0.25
    assert _valor(resultado, "Provisión de prima vacacional") == Decimal("2750")
    assert _valor(resultado, "Provisión total") == Decimal("21250")  # 7500 + 11000 + 2750
    assert _valor(resultado, "Provisión total") == (
        _valor(resultado, "Provisión de aguinaldo")
        + _valor(resultado, "Provisión de vacaciones")
        + _valor(resultado, "Provisión de prima vacacional")
    )


async def test_la_prima_pagada_se_resta_de_la_provision_de_prima(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "021", "PV", CategoriaProvision.PRIMA_VACACIONAL),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("021", "PV", "Prima vacacional", "1000.00", "0.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Prima vacacional pagada en el ejercicio") == Decimal("1000")
    assert _valor(resultado, "Provisión de prima vacacional") == Decimal("1750")


async def test_lo_pagado_en_otro_ejercicio_no_se_resta(db: AsyncSession) -> None:
    """El devengo es del ejercicio: restarle un aguinaldo pagado el año anterior dejaría la
    provisión corta, que es exactamente el error que un pasivo mal reconocido comete."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb",
        fecha_pago=date(2025, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", "AGU", "Aguinaldo", "5000.00", "0.00")],
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("0")
    assert _valor(resultado, "Provisión de aguinaldo") == Decimal("7500")


async def test_lo_pagado_despues_del_corte_no_se_resta(db: AsyncSession) -> None:
    """Con corte a junio, el aguinaldo de diciembre todavía no se pagó: restarlo daría un
    pasivo inexistente en los estados financieros de ese cierre parcial."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    # Periodos propios, todos anteriores al corte de junio.
    await _tres_quincenas(db, eid, pagos=(date(2026, 4, 30), date(2026, 5, 31), date(2026, 6, 30)))
    await _quincena(
        db,
        eid,
        uuid="cccccccc-cccc-cccc-cccc-cccccccccccc",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", "AGU", "Aguinaldo", "5000.00", "0.00")],
    )

    resultado = await b08.consultar(db, eid, b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=date(2026, 6, 30)))

    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("0")
    # Del 1 de enero al 30 de junio: 181 de los 365 días del ejercicio.
    assert _valor(resultado, "Aguinaldo devengado") == Decimal("500") * 15 * Decimal(181) / Decimal(365)
    # El corte coincide con el último periodo timbrado, así que no hay tramo proyectado.
    assert "COBERTURA_INCOMPLETA_AL_CORTE" not in _claves(resultado)


# --------------------------------------------------------------------------------------
# 4. Degradaciones: vacío, nunca cero
# --------------------------------------------------------------------------------------


async def test_sin_fecha_de_inicio_las_columnas_van_vacias_no_en_cero(db: AsyncSession) -> None:
    """No se supone «trabajó todo el año»: para alguien contratado a mitad de ejercicio eso
    duplicaría su provisión. Y un cero afirmaría que no se le debe nada."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid, inicio=None)

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1
    assert _valor(resultado, "Antigüedad al corte (años cumplidos)") is None
    assert _valor(resultado, "Aguinaldo devengado") is None
    assert _valor(resultado, "Provisión de aguinaldo") is None
    assert _valor(resultado, "Provisión total") is None
    assert "FALTA_FECHA_INICIO_RELACION_LABORAL" in _claves(resultado)


async def test_con_fecha_de_inicio_no_se_emite_la_bandera(db: AsyncSession) -> None:
    """Gemela negativa."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert "FALTA_FECHA_INICIO_RELACION_LABORAL" not in _claves(resultado)
    assert _valor(resultado, "Provisión total") is not None


async def test_sin_tabla_de_vacaciones_solo_se_vacian_las_columnas_de_vacaciones(db: AsyncSession) -> None:
    """El aguinaldo no depende del art. 76, así que su columna sigue calculando: apagar una
    comprobación por faltar un dato que esa comprobación no usa es el defecto que B-03 pagó
    tres veces."""
    eid = await _empresa(db)
    await _tres_quincenas(db, eid)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Aguinaldo devengado") == Decimal("7500")
    assert _valor(resultado, "Días de vacaciones del año (art. 76 LFT)") is None
    assert _valor(resultado, "Provisión de vacaciones") is None
    assert _valor(resultado, "Provisión total") is None
    assert "FALTA_TABLA_VACACIONES" in _claves(resultado)


async def test_una_fecha_de_inicio_posterior_al_corte_deja_bandera(db: AsyncSession) -> None:
    """No puede ser cierto de alguien a quien ya se le timbró nómina, y era la quinta bandera por
    empleado sin prueba (ronda de corrección 1). Se devengan cero días —no un número negativo— y
    la captura se reporta."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid, pagos=(date(2026, 4, 30), date(2026, 5, 31), date(2026, 6, 30)), inicio=date(2026, 9, 1))

    resultado = await b08.consultar(db, eid, b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=date(2026, 6, 30)))

    assert "FECHA_INICIO_POSTERIOR_AL_CORTE" in _claves(resultado)
    assert _valor(resultado, "Aguinaldo devengado") == Decimal("0")
    assert _valor(resultado, "Provisión total") == Decimal("0")
    assert _valor(resultado, "Antigüedad al corte (años cumplidos)") == 0


async def test_una_fecha_de_inicio_anterior_al_corte_no_deja_bandera(db: AsyncSession) -> None:
    """Gemela negativa."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert "FECHA_INICIO_POSTERIOR_AL_CORTE" not in _claves(resultado)


async def test_una_percepcion_sin_clave_se_reporta_y_no_bloquea(db: AsyncSession) -> None:
    """`map_concepto_provision` lleva la clave en la PK, así que un concepto sin clave no se
    puede clasificar **ni queriendo**: bloquear por él dejaría el informe muerto sin que nadie
    pueda arreglarlo desde la configuración. Se reporta con la misma clave que ya usa B-02."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)
    await insertar_nomina(
        db,
        empresa_id=eid,
        uuid="dddddddd-dddd-dddd-dddd-dddddddddddd",
        fecha_pago=date(2026, 12, 15),
        fecha_final_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", None, "Aguinaldo sin clave", "5000.00", "0.00")],  # type: ignore[list-item]
        deducciones=[],
        total_percepciones="5000.00",
        total_deducciones="0.00",
        total="5000.00",
        fecha_inicio_rel_laboral=_INICIO_2020,
    )

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1
    assert "CLAVE_VACIA" in _claves(resultado)
    assert "5000" in _de_clave(resultado, "CLAVE_VACIA")[0].mensaje


async def test_sin_percepciones_sin_clave_no_se_emite_clave_vacia(db: AsyncSession) -> None:
    """Gemela negativa."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert "CLAVE_VACIA" not in _claves(resultado)


async def test_un_cancelado_no_cuenta_como_aguinaldo_pagado(db: AsyncSession) -> None:
    """R-T1. Un recibo cancelado no representa un pago que subsista: restarlo del devengado
    dejaría la provisión corta por un aguinaldo que nadie cobró."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(
        db,
        eid,
        ("P", "001", "001", CategoriaProvision.NO_APLICA),
        ("P", "002", "AGU", CategoriaProvision.AGUINALDO),
    )
    await _tres_quincenas(db, eid)
    await _quincena(
        db,
        eid,
        uuid="eeeeeeee-eeee-eeee-eeee-eeeeeeeeeeee",
        fecha_pago=date(2026, 12, 15),
        tipo_nomina="E",
        percepciones=[("002", "AGU", "Aguinaldo", "5000.00", "0.00")],
        estatus=EstatusCfdi.CANCELADO,
    )

    resultado = await b08.consultar(db, eid, _p())

    assert _valor(resultado, "Aguinaldo pagado en el ejercicio") == Decimal("0")
    assert _valor(resultado, "Provisión de aguinaldo") == Decimal("7500")


# --------------------------------------------------------------------------------------
# 4b. Cobertura de los datos frente a la fecha de corte
# --------------------------------------------------------------------------------------


async def test_un_corte_muy_por_delante_de_los_datos_deja_bandera(db: AsyncSession) -> None:
    """**Important 1 de la ronda 1, y ya había afectado a la corrida real.** Con CFDI que
    terminan en julio y corte al 31 de diciembre el informe devolvía una cifra anual con cero
    banderas: una proyección de 12 meses sobre 6.5 observados, presentada como devengo observado.
    El devengo legal del art. 87 LFT corre igual, así que la cifra no cambia — lo que faltaba era
    que el libro lo dijera."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid, pagos=(date(2026, 6, 15), date(2026, 6, 30), date(2026, 7, 15)))

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 1
    assert "COBERTURA_INCOMPLETA_AL_CORTE" in _claves(resultado)
    mensaje = _de_clave(resultado, "COBERTURA_INCOMPLETA_AL_CORTE")[0].mensaje
    assert "2026-07-15" in mensaje
    assert "169 días" in mensaje  # del 15 de julio al 31 de diciembre
    assert "PROYECCIÓN" in mensaje
    # La cifra devengada NO se recorta: el devengo legal no depende de lo que se haya descargado.
    assert _valor(resultado, "Aguinaldo devengado") == Decimal("7500")


async def test_un_corte_al_dia_del_ultimo_cfdi_no_deja_bandera_de_cobertura(db: AsyncSession) -> None:
    """Gemela negativa: en un cierre sano el corte cae en el último periodo timbrado."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert "COBERTURA_INCOMPLETA_AL_CORTE" not in _claves(resultado)


async def test_la_columna_de_ultimo_cfdi_y_la_bandera_de_posible_baja(db: AsyncSession) -> None:
    """**Important 2 de la ronda 1.** La nota manda conciliar la plantilla contra recursos
    humanos, y el libro tiene que decir por dónde empezar: el último `fecha_pago` de cada
    empleado ya viene en el universo, no es una heurística. Convierte «concilia 300 personas» en
    «revisa esta»."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)  # RFC_EMPLEADO, cobrando hasta el 31 de diciembre
    # El otro dejó de recibir recibos en marzo, mientras el resto siguió cobrando.
    await _tres_quincenas(
        db,
        eid,
        prefijo="1515151",
        pagos=(date(2026, 1, 31), date(2026, 2, 28), date(2026, 3, 31)),
        rfc_receptor=RFC_OTRO,
    )

    resultado = await b08.consultar(db, eid, _p())

    indice_ultimo = _columna(resultado, "Último CFDI del ejercicio")
    por_rfc = {fila[_columna(resultado, "RFC empleado")]: fila for fila in resultado.filas}
    assert por_rfc[RFC_EMPLEADO][indice_ultimo] == date(2026, 12, 31)
    assert por_rfc[RFC_OTRO][indice_ultimo] == date(2026, 3, 31)
    assert "POSIBLE_BAJA_EN_EL_EJERCICIO" in _claves(resultado)
    bandera = _de_clave(resultado, "POSIBLE_BAJA_EN_EL_EJERCICIO")[0]
    assert "Empleados afectados: 1" in bandera.mensaje
    assert RFC_OTRO in bandera.mensaje
    assert RFC_EMPLEADO not in bandera.mensaje


async def test_una_plantilla_que_cobra_toda_el_ejercicio_no_deja_bandera_de_baja(db: AsyncSession) -> None:
    """Gemela negativa: con los dos empleados cobrando hasta el corte, nadie es sospechoso."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)
    await _tres_quincenas(db, eid, prefijo="1616161", rfc_receptor=RFC_OTRO)

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 2
    assert "POSIBLE_BAJA_EN_EL_EJERCICIO" not in _claves(resultado)


# --------------------------------------------------------------------------------------
# 5. Presentación: el rótulo, lo sensible y el orden
# --------------------------------------------------------------------------------------


async def test_la_hoja_parametros_lleva_el_rotulo_de_estimacion(db: AsyncSession) -> None:
    """**B-08.R3.** Quien recibe el Excel puede ser el contador que va a armar el asiento: el
    rótulo tiene que estar en el libro, no solo en el docstring del módulo."""
    eid = await _escenario_completo(db)
    resultado = await b08.consultar(db, eid, _p())

    ctx = ContextoInforme(
        clave=b08.CLAVE,
        nombre=b08.NOMBRE,
        usuario="dgarcia@planjuarez.org",
        generado_en=datetime(2026, 8, 7, 9, 0, 0),
        parametros={"ejercicio": _EJERCICIO, "fecha_corte": str(_CORTE), "enmascarar_datos_personales": True},
        etl_version=1,
    )
    libro = load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))
    hoja = libro["Parámetros"]
    texto = "\n".join(
        str(celda.value) for fila in hoja.iter_rows() for celda in fila if celda.value is not None
    )

    assert "estimación" in texto.lower()
    assert "actuarial" in texto.lower()
    assert "NIF D-3" in texto
    assert "prima de antigüedad" in texto.lower()
    # Los otros tres residuos que la ronda 1 exigió rotular, no solo documentar.
    assert "artículo 78" in texto, "las vacaciones no arrastran ejercicios anteriores"
    assert "aguinaldo de 2025 pagado en enero de 2026" in texto, "lo pagado se identifica por fecha de pago"
    assert "Último CFDI del ejercicio" in texto, "la nota de bajas tiene que decir por dónde empezar"


async def test_la_celda_de_dias_pendientes_no_guarda_precision_oculta(db: AsyncSession) -> None:
    """**Menor 6 de la ronda 1, cerrado en la ronda 2.** «Días de vacaciones pendientes» sale de
    una división no terminante (`dias_del_año × dias_trabajados / dias_ejercicio`), y hasta la
    ronda 2 el motor solo cuantizaba el tipo `monto`: la celda guardaba `12.88767123287671`
    mientras su formato mostraba tres decimales. Nadie ve la diferencia en Excel y cualquiera la ve
    al exportar a CSV.

    Se comprueba sobre el libro **escrito**, comparando el valor guardado contra los decimales que
    declara el `number_format` de la propia celda — no contra una constante del motor, que probaría
    que el motor se llama a sí mismo."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid, pagos=(date(2026, 4, 30), date(2026, 5, 31), date(2026, 6, 30)))

    # Corte a media año: 181 de 365 días, así que 22 × 181/365 no termina en decimal.
    resultado = await b08.consultar(db, eid, b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=date(2026, 6, 30)))
    crudo = _valor(resultado, "Días de vacaciones pendientes (estimados)")
    assert -crudo.as_tuple().exponent > 3, "el fixture debe producir una división no terminante"

    ctx = ContextoInforme(
        clave=b08.CLAVE,
        nombre=b08.NOMBRE,
        usuario="dgarcia@planjuarez.org",
        generado_en=datetime(2026, 8, 7, 9, 0, 0),
        parametros={"ejercicio": _EJERCICIO, "fecha_corte": "2026-06-30", "enmascarar_datos_personales": True},
        etl_version=1,
    )
    hoja = load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))["Datos"]

    problemas: list[str] = []
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            if not isinstance(celda.value, (int, float, Decimal)) or isinstance(celda.value, bool):
                continue
            _e, _p, decimales = celda.number_format.partition(".")
            if -Decimal(str(celda.value)).as_tuple().exponent > len(decimales):
                problemas.append(f"{celda.coordinate} guarda {celda.value} y muestra {len(decimales)} decimales")
    assert problemas == [], f"precisión oculta en el libro de B-08: {problemas}"

    columna = _columna(resultado, "Días de vacaciones pendientes (estimados)") + 1
    celda_dias = hoja.cell(row=2, column=columna)
    assert Decimal(str(celda_dias.value)) == crudo.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
    assert celda_dias.number_format == "#,##0.000"


async def test_las_filas_sin_provision_calculable_se_cuentan_en_parametros(db: AsyncSession) -> None:
    """**Menor de la ronda 1.** Dejar la celda vacía es lo correcto —un cero afirmaría que no se
    le debe nada a esa persona— pero **una celda vacía suma cero al totalizar en Excel**, así que
    sin este conteo la decisión correcta produce un total corto con apariencia de completo."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)  # calcula
    await _tres_quincenas(db, eid, prefijo="1717171", rfc_receptor=RFC_OTRO, inicio=None)  # no calcula

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.filas) == 2
    aviso = [nota for nota in resultado.notas if "AL TOTALIZAR" in nota]
    assert aviso, f"faltó el conteo de filas no calculables; notas: {resultado.notas}"
    assert "1 de las 2 filas" in aviso[0]
    assert "suma CERO en Excel" in aviso[0]


async def test_sin_filas_no_calculables_no_se_agrega_el_aviso_de_totalizar(db: AsyncSession) -> None:
    """Gemela negativa: una nota que sale siempre no la lee nadie."""
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, _p())

    assert not [nota for nota in resultado.notas if "AL TOTALIZAR" in nota]


async def test_el_rotulo_tambien_viaja_cuando_el_informe_no_se_genera(db: AsyncSession) -> None:
    """La hoja `Parámetros` lleva el rótulo siempre: un libro sin filas también circula."""
    eid = await _empresa(db)
    resultado = await b08.consultar(db, eid, _p())

    assert resultado.notas, "el rótulo de B-08.R3 no puede depender de que haya filas"
    assert any("actuarial" in nota for nota in resultado.notas)


async def test_curp_y_nss_se_declaran_sensibles_y_el_informe_no_enmascara(db: AsyncSession) -> None:
    """El motor enmascara, el informe declara: un informe que llamara a `enmascarar()` por su
    cuenta rompería el contrato de `enmascarar_datos_personales=False`."""
    eid = await _escenario_completo(db)
    resultado = await b08.consultar(db, eid, _p())

    por_titulo = {c.titulo: c for c in resultado.columnas}
    assert por_titulo["CURP"].sensible is True
    assert por_titulo["NSS"].sensible is True
    assert por_titulo["RFC empleado"].sensible is False
    # El valor llega EN CLARO al motor; es él quien decide.
    assert _valor(resultado, "CURP") == "XXXX800101HCHXXX01"
    assert "****" not in str(_valor(resultado, "CURP"))


async def test_ninguna_bandera_interpola_curp_ni_nss(db: AsyncSession) -> None:
    """Fue una fuga real en B-10: la hoja `Banderas` **no** se enmascara, así que un dato
    personal en el mensaje o en el ámbito sale en claro en un Excel que circula por correo."""
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "029", "029", CategoriaProvision.NO_APLICA))
    for i, pago in enumerate(_PAGOS_ORDINARIOS, start=1):
        await _quincena(
            db,
            eid,
            uuid=f"1212121{i}-1212-1212-1212-121212121212",
            fecha_pago=pago,
            percepciones=[("029", "029", "Asimilado", "7500.00", "0.00")],
            inicio=None,
        )

    resultado = await b08.consultar(db, eid, _p())

    assert len(resultado.banderas) >= 2, "sin banderas la prueba no comprobaría nada"
    for bandera in resultado.banderas:
        assert "XXXX800101HCHXXX01" not in bandera.mensaje
        assert "XXXX800101HCHXXX01" not in bandera.ambito
        assert "12345678901" not in bandera.mensaje
        assert "12345678901" not in bandera.ambito


async def test_el_orden_de_las_filas_es_determinista(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    await _tres_quincenas(db, eid)
    for i, pago in enumerate(_PAGOS_ORDINARIOS, start=1):
        await _quincena(
            db, eid, uuid=f"1313131{i}-1313-1313-1313-131313131313", fecha_pago=pago, rfc_receptor=RFC_OTRO
        )

    resultado = await b08.consultar(db, eid, _p())
    otra_vez = await b08.consultar(db, eid, _p())

    rfcs = [fila[_columna(resultado, "RFC empleado")] for fila in resultado.filas]
    assert rfcs == sorted(rfcs)
    assert resultado.filas == otra_vez.filas


async def test_esta_en_el_registro_y_declara_solo_nomina(db: AsyncSession) -> None:
    definicion = registro.obtener("B-08")
    assert definicion.CLAVE == "B-08"
    assert definicion.TIPOS_COMPROBANTE == ("N",)
    assert "B-08" in {entrada["clave"] for entrada in registro.catalogo()}
    # Los dos parámetros nacen nulos: es lo que hace posible resolverlos de la configuración.
    parametros = b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=_CORTE)
    assert parametros.dias_aguinaldo is None
    assert parametros.factor_prima_vacacional is None
    assert parametros.enmascarar_datos_personales is True


async def test_la_resolucion_local_de_vacaciones_coincide_con_el_servicio(db: AsyncSession) -> None:
    """**La red que ata la duplicación.** `_vacaciones_por_antiguedad` trae `tabla_vacaciones`
    entera de una consulta y resuelve en memoria (así el módulo no tiene ningún bucle con
    entrada/salida), lo que duplica la regla del art. 76 que vive en `cfg.dias_de_vacaciones`.
    Duplicar una regla legal es lo que este proyecto evita, así que la duplicación va atada: esta
    prueba compara las dos implementaciones para 0..50 años y falla en cuanto se separen.

    Se siembra la tabla **completa** del repositorio, no el subconjunto de las demás pruebas: la
    regla que se compara es justamente la de los tramos de cinco años del segundo párrafo."""
    await cfg.cargar_desde_yaml(db, Path("config/fiscal/tabla_vacaciones.yaml"))
    antiguedades = set(range(0, 51))

    resuelto = await b08._vacaciones_por_antiguedad(db, antiguedades)

    for anios in sorted(antiguedades):
        # `max(anios, 1)`: con menos de un año cumplido el derecho es proporcional (art. 77) y la
        # base del prorrateo es el renglón del primer año. Es la única traducción que hace B-08.
        esperado = await cfg.dias_de_vacaciones(db, max(anios, 1))
        assert resuelto.dias_por_antiguedad[anios] == esperado, f"discrepancia con {anios} años"
    assert resuelto.dias_por_antiguedad[0] == 12
    assert resuelto.dias_por_antiguedad[6] == 22
    assert resuelto.dias_por_antiguedad[7] == 22  # el tramo 6-10 no tiene renglón por año
    assert resuelto.dias_por_antiguedad[50] == 38


async def test_una_tabla_de_vacaciones_incompleta_no_dice_que_esta_vacia(db: AsyncSession) -> None:
    """**Menor de la ronda 1.** Decirle a alguien que su tabla está vacía cuando tiene renglones
    lo manda a revisar donde no está el problema: son dos huecos con dos arreglos distintos."""
    eid = await _empresa(db)
    await _config(db, eid)
    await _clasificar(db, eid, ("P", "001", "001", CategoriaProvision.NO_APLICA))
    # La tabla existe pero arranca en el año 6: al empleado de 6 años le resuelve y a uno de 2 no.
    db.add(TablaVacaciones(anios_antiguedad=6, dias=22))
    await db.commit()
    await _tres_quincenas(db, eid, inicio=date(2024, 1, 1))  # 2 años cumplidos al corte

    resultado = await b08.consultar(db, eid, _p())

    mensaje = _de_clave(resultado, "FALTA_TABLA_VACACIONES")[0].mensaje
    assert "1 renglón(es)" in mensaje
    assert "el renglón más bajo capturado es el de 6 año(s)" in mensaje
    assert "No hay ningún renglón" not in mensaje


async def test_el_aviso_de_clasificacion_trunca_la_lista_larga(db: AsyncSession) -> None:
    """La rama de `MAX_CONCEPTOS_EN_AVISO` (`; y N más`), que no ejercitaba ninguna prueba. Se
    prueba sobre la función pura: fabricar 20 percepciones distintas en la base costaría 20 CFDI
    para comprobar un `slice`."""
    conceptos = [
        cfg.ConceptoObservado(
            naturaleza="P",
            tipo="001",
            clave=f"{i:03d}",
            concepto=f"Concepto {i}",
            comprobantes=1,
            importe=Decimal("1"),
            categoria=None,
        )
        for i in range(20)
    ]

    texto = b08._lista_de_conceptos(conceptos)

    assert texto.count(";") == b08.MAX_CONCEPTOS_EN_AVISO
    assert f"y {20 - b08.MAX_CONCEPTOS_EN_AVISO} más" in texto
    assert "P/001/014 (Concepto 14)" in texto  # el último que sí se enumera
    assert "Concepto 19" not in texto
    # Y la gemela: con pocos conceptos no se anuncia ningún resto.
    assert "más" not in b08._lista_de_conceptos(conceptos[:2])


async def test_sin_cfdi_del_ejercicio_no_es_un_error(db: AsyncSession) -> None:
    eid = await _empresa(db)
    await _vacaciones(db)
    await _config(db, eid)

    resultado = await b08.consultar(db, eid, _p())

    assert resultado.filas == []
    assert resultado.aviso is not None


async def test_una_fecha_de_corte_anterior_al_ejercicio_no_devenga_nada(db: AsyncSession) -> None:
    eid = await _escenario_completo(db)

    resultado = await b08.consultar(db, eid, b08.Parametros(ejercicio=_EJERCICIO, fecha_corte=date(2025, 6, 30)))

    assert resultado.filas == []
    assert resultado.aviso is not None and "anterior al ejercicio" in resultado.aviso
