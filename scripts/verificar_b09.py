"""Verificación en vivo de B-09 (recálculo de ISR) y de las tres columnas anuales de B-05
(Task 6 del plan, `docs/superpowers/plans/2026-08-11-b09-isr.md`, §7 del diseño en
`docs/superpowers/specs/2026-08-11-b09-isr-design.md`).

Comprueba, contra la base de datos real y el mismo PDF oficial que usa
`scripts/verificar_tarifa_isr.py` —nunca contra dobles—, las ocho afirmaciones del brief de esta
tarea: que sin configuración confirmada B-09 no genera nada y dice qué falta, que confirmar la
tarifa quincenal y las marcas `001`/`005` (pero no el subsidio) deja calcular el ISR determinado
con las columnas del subsidio vacías, que con todo confirmado los 8 recibos reales de la empresa
11 producen 8 filas, que el ISR de un recibo real coincide con un cálculo hecho a mano, que el
libro de Excel abre con sus cuatro hojas y su hoja de parámetros refleja la configuración usada,
que las tres columnas anuales de B-05 aparecen y se calculan (y salen vacías sin marcas), y que la
base queda exactamente como se encontró al terminar.

Por qué llama a las funciones del endpoint directamente, y no por HTTP
------------------------------------------------------------------------
Mismo argumento que `scripts/verificar_tarifa_isr.py`: los endpoints de `app.api.v1.configuracion`
exigen `require_admin` (un ID token real de Firebase contra el proyecto de producción), que este
script no debe fabricar solo para verificar en local. Se le pasa un `admin` fabricado en memoria
(nunca escrito en `usuarios`) con el único atributo que esas funciones leen (`.correo`), y un
`UploadFile` construido sobre los bytes reales de `tests/fixtures/anexo8-2026.pdf`.

Punto de control no negociable: la base tiene que quedar exactamente como se encontró
----------------------------------------------------------------------------------------
Al arrancar, esta base tiene `tarifa_isr` **vacía** (0 filas), los 5 valores de `param_fiscal`
confirmados el 2026-08-07 (`UMA_DIARIA`, `UMA_MENSUAL`, `UMA_ANUAL`, `SALARIO_MINIMO_GENERAL`,
`SALARIO_MINIMO_ZLFN`), y seis tramos más (UMA 2025 y subsidio al empleo) más las 44 marcas de
`catalogo_percepcion_marca`, **ninguno confirmado**. Este script:

1. **Crea** las cabeceras y renglones de `tarifa_isr` que necesita (importando el Anexo 8 real) y
   los **descarta** al terminar — igual que `verificar_tarifa_isr.py`.
2. **Confirma temporalmente** dos tramos de `param_fiscal` que ya estaban sembrados sin confirmar
   (`SUBSIDIO_FACTOR_UMA` del tramo vigente desde 2026-02-01, `SUBSIDIO_TOPE_INGRESO`) y dos marcas
   de `catalogo_percepcion_marca` (`001`, `005`) — y **revierte la confirmación** (no el dato: las
   seis marcas y el valor de cada tramo no se tocan, solo `confirmado_por`/`confirmado_en` vuelven
   a `NULL`) al terminar. Es exactamente lo que promete la nota de cierre del brief: "la base queda
   como estaba — todas las confirmaciones temporales revertidas".

Antes de empezar comprueba que `tarifa_isr` esté vacía y que los tramos/marcas que va a confirmar
sigan sin confirmar, y **se niega a correr** si no es así: solo puede prometer dejar la base como
la encontró si sabe con certeza qué había antes. Los **5 valores ya confirmados** de `param_fiscal`
nunca se leen para escribir, solo se fotografían al empezar y se comparan al final.

**La limpieza va en un `finally`** que cubre desde la primera escritura (la importación del Anexo
8) hasta el final: si algo revienta a mitad de la corrida, la reversión corre de todos modos. Esto
costó una ronda completa en la Task 12 de la entrega anterior (limpieza fuera del bloque
protegido); se comprobó aquí inyectando temporalmente un `raise` justo antes del paso de
reversión y confirmando por SQL que la base queda igual de limpia (ver el reporte de esta tarea).

Uso: `.venv/bin/python scripts/verificar_b09.py` (host, con `DATABASE_URL` apuntando al MySQL
real de `docker compose`). Sale con código 1 si alguna comprobación falla o si la base no estaba
en el estado esperado al empezar.
"""

from __future__ import annotations

import asyncio
import io
import sys
import traceback
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from fastapi import UploadFile
from sqlalchemy import delete, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1 import configuracion
from app.api.v1.schemas import (
    ImportacionTarifasOut,
    MarcaPercepcionConfirmarIn,
    ParamFiscalConfirmarIn,
    TarifaIsrConfirmarIn,
)
from app.db.session import SessionLocal
from app.informes import b05_acumulado_anual as b05
from app.informes import b09_recalculo_isr as b09
from app.informes import excel
from app.informes.base import Bandera, ContextoInforme, ResultadoInforme
from app.models.bitacora import Bitacora
from app.models.configuracion_fiscal import CatalogoPercepcionMarca, ParamFiscal, TarifaIsr, TarifaIsrRenglon
from app.models.enums import PeriodicidadTarifa
from app.services import configuracion_fiscal as cfg
from app.services import normalizacion

ANEXO_2026 = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "anexo8-2026.pdf"

EMPRESA_ID = 11
EJERCICIO = 2026
FECHA_DESDE = date(2026, 1, 1)
FECHA_HASTA = date(2026, 12, 31)

# Los dos tramos de `param_fiscal` que este script confirma temporalmente. `SUBSIDIO_FACTOR_UMA`
# tiene DOS tramos sembrados (enero y desde febrero); los 8 CFDI reales de la empresa 11 son de
# junio/julio de 2026, así que solo el tramo vigente desde el 1 de febrero mueve el resultado —
# confirmar el de enero no haría nada por esta corrida y ampliaría sin necesidad lo que hay que
# revertir.
_VIGENCIA_FACTOR_SUBSIDIO = date(2026, 2, 1)
_VIGENCIA_TOPE_SUBSIDIO = date(2026, 1, 1)

# El recibo elegido para la comprobación 5 (§ más abajo, junto a `_paso_5_comprobacion_manual`):
# JUAN ALONSO GRANADOS, número de empleado 040, quincena pagada el 2026-07-15, 15 días pagados
# (= los días nominales de la quincena, sin prorrateo). Es uno de los 8 CFDI de nómina reales de
# la empresa 11; el cálculo completo a mano vive en el docstring de esa función.
_UUID_COMPROBACION_MANUAL = "37ACD323-37FB-4F6D-98AE-FD82B7F8D879"

# Objeto en memoria, nunca escrito en `usuarios`: ver el docstring del módulo.
ADMIN = SimpleNamespace(correo="verificacion@script")

_HOJAS_ESPERADAS = {"Datos", "Parámetros", "Banderas", "Diccionario"}


def _archivo_anexo() -> UploadFile:
    return UploadFile(io.BytesIO(ANEXO_2026.read_bytes()), filename="anexo8-2026.pdf")


def _indice(resultado: ResultadoInforme) -> dict[str, int]:
    """Título de columna -> posición. Leer por nombre, no por número: un índice literal se
    rompe en silencio en cuanto alguien inserta una columna."""
    return {columna.titulo: i for i, columna in enumerate(resultado.columnas)}


def _bandera_de(uuid: str, banderas: list[Bandera]) -> list[str]:
    return [b.clave for b in banderas if b.ambito == f"uuid:{uuid}"]


# --------------------------------------------------------------------------------------
# Fotografía de lo que este script promete no tocar (más allá de lo que confirma y revierte)
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class _Fotografia:
    tarifa_isr_total: int
    param_fiscal_total: int
    param_fiscal_confirmados: int
    catalogo_percepcion_total: int
    catalogo_percepcion_confirmados: int
    bitacora_total: int


async def _fotografia(db: AsyncSession) -> _Fotografia:
    tarifa_total = (await db.scalar(select(func.count()).select_from(TarifaIsr))) or 0
    param_total = (await db.scalar(select(func.count()).select_from(ParamFiscal))) or 0
    param_confirmados = (
        await db.scalar(select(func.count()).select_from(ParamFiscal).where(ParamFiscal.confirmado_en.is_not(None)))
    ) or 0
    marca_total = (await db.scalar(select(func.count()).select_from(CatalogoPercepcionMarca))) or 0
    marca_confirmadas = (
        await db.scalar(
            select(func.count())
            .select_from(CatalogoPercepcionMarca)
            .where(CatalogoPercepcionMarca.confirmado_en.is_not(None))
        )
    ) or 0
    bitacora_total = (await db.scalar(select(func.count()).select_from(Bitacora))) or 0
    return _Fotografia(
        tarifa_isr_total=int(tarifa_total),
        param_fiscal_total=int(param_total),
        param_fiscal_confirmados=int(param_confirmados),
        catalogo_percepcion_total=int(marca_total),
        catalogo_percepcion_confirmados=int(marca_confirmadas),
        bitacora_total=int(bitacora_total),
    )


async def _precondiciones(db: AsyncSession) -> list[str]:
    """La base tiene que estar exactamente en el estado documentado: `tarifa_isr` vacía, y los
    dos tramos de `param_fiscal` y las dos marcas que este script va a confirmar temporalmente,
    todavía sin confirmar. Solo así puede prometer dejarla como la encontró."""
    problemas: list[str] = []
    tarifas = (await db.scalars(select(TarifaIsr))).all()
    if tarifas:
        problemas.append(
            f"`tarifa_isr` no está vacía ({len(tarifas)} fila(s)): este script no puede prometer dejarla "
            "como la encontró si no sabe con certeza qué había antes."
        )

    for clave, vigencia in (
        ("SUBSIDIO_FACTOR_UMA", _VIGENCIA_FACTOR_SUBSIDIO),
        ("SUBSIDIO_TOPE_INGRESO", _VIGENCIA_TOPE_SUBSIDIO),
    ):
        fila = await db.get(ParamFiscal, (clave, vigencia))
        if fila is None:
            problemas.append(f"No existe el tramo `{clave}` desde {vigencia}: no hay nada que confirmar.")
        elif fila.confirmado_en is not None:
            problemas.append(f"El tramo `{clave}` desde {vigencia} ya está confirmado; este script no lo tocará.")

    for tipo in ("001", "005"):
        marca = await db.get(CatalogoPercepcionMarca, tipo)
        if marca is None:
            problemas.append(f"No existe la marca `{tipo}` en `catalogo_percepcion_marca`.")
        elif marca.confirmado_en is not None:
            problemas.append(f"La marca `{tipo}` ya está confirmada; este script no la tocará.")

    return problemas


# --------------------------------------------------------------------------------------
# 1: sin ninguna configuración confirmada, B-09 no genera nada
# --------------------------------------------------------------------------------------


async def _paso_1_sin_confirmar(db: AsyncSession) -> list[str]:
    print("\n--- 1. B-09 con la configuración fiscal sin confirmar (tarifa, marcas y subsidio) ---")
    fallas: list[str] = []
    resultado = await b09.consultar(db, EMPRESA_ID, b09.Parametros(fecha_desde=FECHA_DESDE, fecha_hasta=FECHA_HASTA))
    print(f"Filas: {len(resultado.filas)}")
    print(f"Aviso: {resultado.aviso}")

    if resultado.filas:
        fallas.append(f"B-09 produjo {len(resultado.filas)} fila(s) sin ninguna tarifa ni marca confirmada")
    if not resultado.aviso:
        fallas.append("B-09 no generó filas pero tampoco trae ningún aviso explicando por qué")
    else:
        # El aviso tiene que decir las dos cosas que faltan (tarifa Y marcas) y dónde cargar cada
        # una: es literalmente el texto de `configuracion_isr.AVISO_SIN_TARIFA` +
        # `AVISO_SIN_MARCAS` (B-09 nunca redacta el suyo, ver su docstring).
        if "Anexo 8" not in resultado.aviso or "Configuración → Fiscal" not in resultado.aviso:
            fallas.append(f"El aviso no dice dónde cargar la tarifa (Anexo 8 / Configuración → Fiscal): {resultado.aviso!r}")
        if "Quincenal" not in resultado.aviso:
            fallas.append(f"El aviso no nombra la periodicidad quincenal, la única que usa la empresa 11: {resultado.aviso!r}")
        if "001" not in resultado.aviso or "005" not in resultado.aviso:
            fallas.append(f"El aviso no nombra los tipos de percepción sin marca confirmada (001, 005): {resultado.aviso!r}")
    return fallas


# --------------------------------------------------------------------------------------
# 2: importar el Anexo 8 real y confirmar la tarifa quincenal + la anual del ejercicio, y las
#    marcas 001/005 (el subsidio se confirma aparte, en el paso 4, para poder ejercer la
#    degradación intermedia del paso 3)
# --------------------------------------------------------------------------------------


async def _confirmar_marca(db: AsyncSession, tipo: str) -> None:
    fila = await db.get(CatalogoPercepcionMarca, tipo)
    assert fila is not None, f"la marca {tipo} desapareció entre la precondición y este paso"
    body = MarcaPercepcionConfirmarIn(
        es_ingreso_ordinario=fila.es_ingreso_ordinario,
        base_exencion=fila.base_exencion,
        factor_exencion=fila.factor_exencion,
        integra_sbc=fila.integra_sbc,
        es_provisionable=fila.es_provisionable,
        sujeto_a_tope_conjunto=fila.sujeto_a_tope_conjunto,
        multiplicador_no_derivable=fila.multiplicador_no_derivable,
        nota_revision_hash=cfg.huella_de_nota(fila.nota_revision),
    )
    await configuracion.confirmar_percepcion(tipo=tipo, body=body, admin=ADMIN, db=db)


async def _confirmar_param(db: AsyncSession, clave: str, vigencia_desde: date) -> None:
    fila = await db.get(ParamFiscal, (clave, vigencia_desde))
    assert fila is not None, f"el tramo {clave}@{vigencia_desde} desapareció entre la precondición y este paso"
    body = ParamFiscalConfirmarIn(vigencia_desde=vigencia_desde, valor=fila.valor)
    await configuracion.confirmar_fiscal(clave=clave, body=body, admin=ADMIN, db=db)


async def _paso_2_importar_y_confirmar_base(db: AsyncSession) -> tuple[ImportacionTarifasOut, list[str]]:
    print("\n--- 2. Importar el Anexo 8 real; confirmar tarifa DIAS_15/EJERCICIO 2026 y marcas 001/005 ---")
    fallas: list[str] = []
    resultado_import = await configuracion.importar_tarifa_isr(archivo=_archivo_anexo(), admin=ADMIN, db=db)
    print(f"Tarifas importadas: {len(resultado_import.tarifas)}")
    if len(resultado_import.tarifas) != 7:
        fallas.append(f"Se importaron {len(resultado_import.tarifas)} tarifas del Anexo 8 y se esperaban 7")

    quincenal = next(
        (t for t in resultado_import.tarifas if t.ejercicio == EJERCICIO and t.periodicidad is PeriodicidadTarifa.DIAS_15),
        None,
    )
    anual = next(
        (t for t in resultado_import.tarifas if t.ejercicio == EJERCICIO and t.periodicidad is PeriodicidadTarifa.EJERCICIO),
        None,
    )
    if quincenal is None or anual is None:
        fallas.append("No se encontraron las tarifas DIAS_15 y/o EJERCICIO de 2026 entre las importadas")
        return resultado_import, fallas

    confirmada_q = await configuracion.confirmar_tarifa_isr(
        ejercicio=EJERCICIO,
        periodicidad=PeriodicidadTarifa.DIAS_15,
        body=TarifaIsrConfirmarIn(huella=quincenal.huella),
        admin=ADMIN,
        db=db,
    )
    confirmada_a = await configuracion.confirmar_tarifa_isr(
        ejercicio=EJERCICIO,
        periodicidad=PeriodicidadTarifa.EJERCICIO,
        body=TarifaIsrConfirmarIn(huella=anual.huella),
        admin=ADMIN,
        db=db,
    )
    print(f"Quincenal 2026 confirmada: {confirmada_q.confirmada} · Anual 2026 confirmada: {confirmada_a.confirmada}")
    if not (confirmada_q.confirmada and confirmada_a.confirmada):
        fallas.append("La tarifa DIAS_15 y/o EJERCICIO de 2026 no quedó confirmada")

    await _confirmar_marca(db, "001")
    await _confirmar_marca(db, "005")
    marca_001 = await db.get(CatalogoPercepcionMarca, "001")
    marca_005 = await db.get(CatalogoPercepcionMarca, "005")
    print(
        f"Marca 001 confirmada: {marca_001.confirmado_en is not None if marca_001 else False} · "
        f"Marca 005 confirmada: {marca_005.confirmado_en is not None if marca_005 else False}"
    )
    if marca_001 is None or marca_001.confirmado_en is None or marca_005 is None or marca_005.confirmado_en is None:
        fallas.append("Las marcas 001 y/o 005 no quedaron confirmadas")

    return resultado_import, fallas


# --------------------------------------------------------------------------------------
# 3: tarifa + marcas confirmadas, sin subsidio -> hay filas, ISR determinado, subsidio vacío
# --------------------------------------------------------------------------------------


async def _paso_3_sin_subsidio(db: AsyncSession) -> tuple[ResultadoInforme, list[str]]:
    print("\n--- 3. B-09 con tarifa y marcas confirmadas, SIN el subsidio al empleo ---")
    fallas: list[str] = []
    resultado = await b09.consultar(db, EMPRESA_ID, b09.Parametros(fecha_desde=FECHA_DESDE, fecha_hasta=FECHA_HASTA))
    print(f"Filas: {len(resultado.filas)} · notas: {resultado.notas}")

    if len(resultado.filas) != 8:
        fallas.append(f"B-09 produjo {len(resultado.filas)} fila(s); se esperaban los 8 CFDI reales de la empresa 11")

    idx = _indice(resultado)
    for fila in resultado.filas:
        uuid = fila[idx["UUID"]]
        if fila[idx["ISR determinado"]] is None:
            fallas.append(f"{uuid}: 'ISR determinado' vino vacío con tarifa y marcas ya confirmadas")
        for columna in ("Subsidio al empleo teórico", "ISR a retener teórico", "Subsidio a entregar teórico"):
            if fila[idx[columna]] is not None:
                fallas.append(f"{uuid}: '{columna}' trae valor ({fila[idx[columna]]}) sin el subsidio confirmado")

    if not any("subsidio" in nota.lower() for nota in resultado.notas):
        fallas.append(f"No hay ninguna nota explicando la ausencia del subsidio: {resultado.notas}")
    return resultado, fallas


# --------------------------------------------------------------------------------------
# 4: confirmar el subsidio al empleo -> con todo confirmado, 8 recibos producen 8 filas
# --------------------------------------------------------------------------------------


async def _paso_4_todo_confirmado(db: AsyncSession) -> tuple[ResultadoInforme, list[str]]:
    print("\n--- 4. Confirmar el subsidio al empleo; B-09 con todo confirmado ---")
    fallas: list[str] = []
    await _confirmar_param(db, "SUBSIDIO_FACTOR_UMA", _VIGENCIA_FACTOR_SUBSIDIO)
    await _confirmar_param(db, "SUBSIDIO_TOPE_INGRESO", _VIGENCIA_TOPE_SUBSIDIO)

    resultado = await b09.consultar(db, EMPRESA_ID, b09.Parametros(fecha_desde=FECHA_DESDE, fecha_hasta=FECHA_HASTA))
    print(f"Filas: {len(resultado.filas)} · notas: {resultado.notas}")

    if len(resultado.filas) != 8:
        fallas.append(f"B-09 produjo {len(resultado.filas)} fila(s) con todo confirmado; se esperaban 8")

    idx = _indice(resultado)
    for fila in resultado.filas:
        uuid = fila[idx["UUID"]]
        if fila[idx["ISR determinado"]] is None:
            fallas.append(f"{uuid}: 'ISR determinado' vino vacío con todo confirmado")
        if fila[idx["Subsidio al empleo teórico"]] is None:
            fallas.append(f"{uuid}: 'Subsidio al empleo teórico' sigue vacío con el subsidio ya confirmado")
        if fila[idx["ISR a retener teórico"]] is None:
            fallas.append(f"{uuid}: 'ISR a retener teórico' sigue vacío con el subsidio ya confirmado")

    if resultado.notas:
        fallas.append(f"Siguen apareciendo notas con el subsidio ya confirmado: {resultado.notas}")
    return resultado, fallas


# --------------------------------------------------------------------------------------
# 5: el ISR de un recibo real, comparado contra un cálculo hecho a mano
# --------------------------------------------------------------------------------------


def _paso_5_comprobacion_manual(resultado: ResultadoInforme) -> list[str]:
    """Comprobación 5 del brief: el cálculo completo, a mano, para el recibo elegido —
    `_UUID_COMPROBACION_MANUAL` (JUAN ALONSO GRANADOS, núm. empleado 040, quincena pagada el
    2026-07-15, 15 días pagados = los días nominales, sin prorrateo).

    **Base gravable** (Σ `importe_gravado` de percepciones ORDINARIAS del recibo). El tipo `005`
    (fondo de ahorro) es 100 % exento en este recibo (`importe_gravado = 0.00`), así que la base
    la pone solo el tipo `001`:

        001 → 8,759.70
        --------------------------------
        Base gravable = 8,759.70

    **Renglón de la tarifa quincenal 2026** (Anexo 8, DIAS_15) que le toca a 8,759.70 — el
    renglón 6 va de 8,651.41 a 17,448.75, con cuota fija 916.20 y tasa 21.36 % (0.2136):

        excedente         = 8,759.70 − 8,651.41       = 108.29
        impuesto marginal = 108.29 × 0.2136           = 23.130744 → 23.13 (redondeo a 2
                                                          decimales, ROUND_HALF_UP)
        ISR determinado   = cuota fija + marginal     = 916.20 + 23.13 = 939.33

    **Subsidio al empleo teórico** (Anexo I.3, modelo de monto fijo vigente desde 2024): el
    subsidio se define en términos mensuales, así que primero se mensualiza la base con el mes de
    30 días de convención (nunca los días naturales del mes):

        mensualizado = 8,759.70 × 30 días / 15 días pagados = 17,519.40

    17,519.40 rebasa el tope de ingreso confirmado (11,492.66) ⇒ no hay derecho a subsidio:

        Subsidio al empleo teórico = 0.00
        ISR a retener teórico      = max(939.33 − 0.00, 0) = 939.33
        Subsidio a entregar teórico = max(0.00 − 939.33, 0) = 0.00

    **Lo que el patrón de verdad timbró en este CFDI** (`nomina_deduccion`/`nomina_otro_pago`):

        ISR retenido en el CFDI     = 939.32
        Subsidio causado en el CFDI = 0.00

        Diferencia de ISR = 939.32 − 939.33 = −0.01

    |−0.01| ≤ 0.02 (`_UMBRAL_COINCIDE` de `b09_recalculo_isr`) ⇒ bandera `COINCIDE`: el centavo de
    diferencia se explica por redondeo entre dos cálculos hechos por separado.
    """
    print("\n--- 5. ISR de un recibo real vs. cálculo hecho a mano ---")
    fallas: list[str] = []
    idx = _indice(resultado)
    fila = next((f for f in resultado.filas if f[idx["UUID"]] == _UUID_COMPROBACION_MANUAL), None)
    if fila is None:
        fallas.append(f"El recibo {_UUID_COMPROBACION_MANUAL} no aparece entre las filas de B-09")
        return fallas

    esperado = {
        "Base gravable": Decimal("8759.70"),
        "Renglón de la tarifa": 6,
        "Límite inferior": Decimal("8651.41"),
        "Excedente": Decimal("108.29"),
        "Tasa sobre excedente (%)": Decimal("21.36"),
        "Impuesto marginal": Decimal("23.13"),
        "Cuota fija": Decimal("916.20"),
        "ISR determinado": Decimal("939.33"),
        "Subsidio al empleo teórico": Decimal("0.00"),
        "ISR a retener teórico": Decimal("939.33"),
        "Subsidio a entregar teórico": Decimal("0.00"),
        "ISR retenido en el CFDI": Decimal("939.32"),
        "Subsidio causado en el CFDI": Decimal("0.00"),
        "Diferencia de ISR": Decimal("-0.01"),
        "Diferencia de subsidio": Decimal("0.00"),
    }
    for columna, valor_esperado in esperado.items():
        obtenido = fila[idx[columna]]
        print(f"  {columna}: obtenido={obtenido} · calculado a mano={valor_esperado}")
        if obtenido != valor_esperado:
            fallas.append(f"{columna}: B-09 calculó {obtenido}, el cálculo a mano da {valor_esperado}")

    claves = _bandera_de(_UUID_COMPROBACION_MANUAL, resultado.banderas)
    print(f"  banderas de este UUID: {claves}")
    if "COINCIDE" not in claves:
        fallas.append(f"El recibo {_UUID_COMPROBACION_MANUAL} no trae la bandera COINCIDE; trae: {claves}")
    return fallas


# --------------------------------------------------------------------------------------
# 6: el libro de Excel abre con sus cuatro hojas y la de parámetros dice qué configuración se usó
# --------------------------------------------------------------------------------------


def _libro(resultado: ResultadoInforme, clave: str, nombre: str, parametros: dict[str, object]) -> openpyxl.Workbook:
    ctx = ContextoInforme(
        clave=clave,
        nombre=nombre,
        usuario="verificacion@script",
        generado_en=datetime.now(timezone.utc).replace(tzinfo=None),
        parametros=parametros,
        etl_version=normalizacion.ETL_VERSION,
    )
    return openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))


def _filas_parametros(libro: openpyxl.Workbook) -> list[tuple[object, ...]]:
    return list(libro["Parámetros"].iter_rows(values_only=True))


def _paso_6_libro_excel(resultado_sin_subsidio: ResultadoInforme, resultado_final: ResultadoInforme) -> list[str]:
    print("\n--- 6. El libro de Excel abre con sus cuatro hojas; la hoja de Parámetros dice qué se usó ---")
    fallas: list[str] = []
    parametros = {"fecha_desde": FECHA_DESDE.isoformat(), "fecha_hasta": FECHA_HASTA.isoformat(), "incluir_cancelados": False}

    libro_sin_subsidio = _libro(resultado_sin_subsidio, b09.CLAVE, b09.NOMBRE, parametros)
    libro_final = _libro(resultado_final, b09.CLAVE, b09.NOMBRE, parametros)

    for etiqueta, libro in (("sin subsidio", libro_sin_subsidio), ("con todo confirmado", libro_final)):
        if set(libro.sheetnames) != _HOJAS_ESPERADAS:
            fallas.append(f"El libro ({etiqueta}) no tiene las cuatro hojas esperadas: {libro.sheetnames}")

    filas_sin_subsidio = _filas_parametros(libro_sin_subsidio)
    notas_sin_subsidio = [f for f in filas_sin_subsidio if f and f[0] == "Nota"]
    print(f"Filas 'Nota' en Parámetros (sin subsidio): {len(notas_sin_subsidio)}")
    if not any("subsidio" in str(f[1]).lower() for f in notas_sin_subsidio):
        fallas.append("La hoja de Parámetros del libro SIN subsidio no trae ninguna nota que mencione el subsidio")

    filas_final = _filas_parametros(libro_final)
    notas_final = [f for f in filas_final if f and f[0] == "Nota"]
    print(f"Filas 'Nota' en Parámetros (todo confirmado): {len(notas_final)}")
    if notas_final:
        fallas.append(f"La hoja de Parámetros del libro con TODO confirmado todavía trae nota(s): {notas_final}")

    return fallas


# --------------------------------------------------------------------------------------
# 7: las tres columnas anuales de B-05, vacías sin marcas/tarifa y calculadas con todo confirmado
# --------------------------------------------------------------------------------------

_COLUMNAS_ANUALES_B05 = ("Gravado ordinario", "ISR anual teórico", "Subsidio anual acreditable", "Diferencia a cargo / favor")


async def _paso_7a_b05_sin_confirmar(db: AsyncSession) -> list[str]:
    print("\n--- 7a. B-05 con marcas y tarifa EJERCICIO sin confirmar: las columnas anuales van vacías ---")
    fallas: list[str] = []
    resultado = await b05.consultar(db, EMPRESA_ID, b05.Parametros(ejercicio=EJERCICIO))
    print(f"Filas: {len(resultado.filas)} · banderas: {[b.clave for b in resultado.banderas]}")

    if len(resultado.filas) != 4:
        fallas.append(f"B-05 produjo {len(resultado.filas)} fila(s); se esperaban 4 (un empleado por RFC)")

    idx = _indice(resultado)
    for fila in resultado.filas:
        for columna in _COLUMNAS_ANUALES_B05:
            if fila[idx[columna]] is not None:
                fallas.append(f"{fila[idx['RFC empleado']]}: '{columna}' trae valor sin marcas/tarifa confirmadas")

    claves = {b.clave for b in resultado.banderas}
    if "MARCA_SIN_CONFIRMAR" not in claves and "FALTA_MARCA" not in claves:
        fallas.append("B-05 no avisó de marcas sin confirmar antes de confirmar nada")
    if "FALTA_TARIFA_EJERCICIO" not in claves:
        fallas.append("B-05 no avisó de la tarifa EJERCICIO sin confirmar antes de confirmar nada")
    return fallas


async def _paso_7b_b05_confirmado(db: AsyncSession) -> list[str]:
    print("\n--- 7b. B-05 con todo confirmado: las columnas anuales aparecen y se calculan ---")
    fallas: list[str] = []
    resultado = await b05.consultar(db, EMPRESA_ID, b05.Parametros(ejercicio=EJERCICIO))
    print(f"Filas: {len(resultado.filas)} · banderas: {[b.clave for b in resultado.banderas]}")

    if len(resultado.filas) != 4:
        fallas.append(f"B-05 produjo {len(resultado.filas)} fila(s) con todo confirmado; se esperaban 4")

    idx = _indice(resultado)
    for fila in resultado.filas:
        for columna in _COLUMNAS_ANUALES_B05:
            if fila[idx[columna]] is None:
                fallas.append(f"{fila[idx['RFC empleado']]}: '{columna}' sigue vacía con todo confirmado")

    claves = {b.clave for b in resultado.banderas}
    if "MARCA_SIN_CONFIRMAR" in claves or "FALTA_MARCA" in claves:
        fallas.append(f"B-05 sigue avisando de marcas sin confirmar con las 44 marcas de la empresa 11 revisadas: {claves}")
    if "FALTA_TARIFA_EJERCICIO" in claves:
        fallas.append("B-05 sigue avisando de la tarifa EJERCICIO sin confirmar, y ya se confirmó")
    return fallas


# --------------------------------------------------------------------------------------
# 8: dejar la base como estaba, y decirlo
# --------------------------------------------------------------------------------------


async def _paso_8_revertir(db: AsyncSession, antes: _Fotografia) -> list[str]:
    print("\n--- 8. Dejar la base como estaba ---")
    fallas: list[str] = []

    ejercicios_creados = {2025, 2026}
    await db.execute(delete(TarifaIsrRenglon).where(TarifaIsrRenglon.ejercicio.in_(ejercicios_creados)))
    borrado = await db.execute(delete(TarifaIsr).where(TarifaIsr.ejercicio.in_(ejercicios_creados)))
    print(f"Descartadas {borrado.rowcount} cabecera(s) de `tarifa_isr` (con sus renglones), creadas por esta corrida.")

    await db.execute(
        update(CatalogoPercepcionMarca)
        .where(CatalogoPercepcionMarca.tipo_percepcion.in_(["001", "005"]))
        .values(confirmado_por=None, confirmado_en=None)
    )
    print("Marcas 001 y 005: confirmación revertida (los seis campos de la marca no se tocaron).")

    for clave, vigencia in (
        ("SUBSIDIO_FACTOR_UMA", _VIGENCIA_FACTOR_SUBSIDIO),
        ("SUBSIDIO_TOPE_INGRESO", _VIGENCIA_TOPE_SUBSIDIO),
    ):
        await db.execute(
            update(ParamFiscal)
            .where(ParamFiscal.clave == clave, ParamFiscal.vigencia_desde == vigencia)
            .values(confirmado_por=None, confirmado_en=None)
        )
    print("Tramos SUBSIDIO_FACTOR_UMA (desde 2026-02-01) y SUBSIDIO_TOPE_INGRESO: confirmación revertida.")

    await db.commit()

    despues = await _fotografia(db)
    if despues.tarifa_isr_total != 0:
        fallas.append(f"`tarifa_isr` sigue con {despues.tarifa_isr_total} fila(s) tras la limpieza; debía quedar vacía")
    else:
        print("`tarifa_isr` queda vacía, igual que al empezar esta corrida.")

    if despues.param_fiscal_total != antes.param_fiscal_total:
        fallas.append(
            f"El número de filas de `param_fiscal` cambió: antes {antes.param_fiscal_total}, ahora "
            f"{despues.param_fiscal_total}. Este script no debía crear ni borrar tramos, solo confirmarlos."
        )
    if despues.param_fiscal_confirmados != antes.param_fiscal_confirmados:
        fallas.append(
            f"El número de tramos CONFIRMADOS de `param_fiscal` cambió: antes {antes.param_fiscal_confirmados}, "
            f"ahora {despues.param_fiscal_confirmados}. Las dos confirmaciones temporales debían quedar revertidas."
        )
    else:
        print(f"`param_fiscal`: {despues.param_fiscal_total} fila(s), {despues.param_fiscal_confirmados} confirmada(s) — igual que al empezar.")

    if despues.catalogo_percepcion_total != antes.catalogo_percepcion_total:
        fallas.append(
            f"El número de filas de `catalogo_percepcion_marca` cambió: antes {antes.catalogo_percepcion_total}, "
            f"ahora {despues.catalogo_percepcion_total}."
        )
    if despues.catalogo_percepcion_confirmados != antes.catalogo_percepcion_confirmados:
        fallas.append(
            f"El número de marcas CONFIRMADAS cambió: antes {antes.catalogo_percepcion_confirmados}, ahora "
            f"{despues.catalogo_percepcion_confirmados}. Las marcas 001/005 debían quedar sin confirmar otra vez."
        )
    else:
        print(
            f"`catalogo_percepcion_marca`: {despues.catalogo_percepcion_total} fila(s), "
            f"{despues.catalogo_percepcion_confirmados} confirmada(s) — igual que al empezar."
        )

    nuevas_bitacora = despues.bitacora_total - antes.bitacora_total
    print(
        f"La bitácora ganó {nuevas_bitacora} fila(s) nueva(s) (importar/confirmar de esta corrida). NO se "
        "borran: `bitacora` es append-only por diseño (regla 8 de CLAUDE.md). Quedan marcadas con "
        "actor='verificacion@script'."
    )
    return fallas


# --------------------------------------------------------------------------------------


async def main() -> int:
    print(f"Verificación en vivo de B-09 y las columnas anuales de B-05 — {ANEXO_2026}")

    async with SessionLocal() as db:
        problemas = await _precondiciones(db)
    if problemas:
        print("\nABORTA: la base no está en el estado que este script necesita para prometer dejarla como la encontró:")
        for problema in problemas:
            print("  -", problema)
        return 1

    async with SessionLocal() as db:
        fotografia_antes = await _fotografia(db)

    fallas: list[str] = []
    resultado_sin_subsidio: ResultadoInforme | None = None
    resultado_final: ResultadoInforme | None = None

    # Todo lo que confirma configuración fiscal (pasos 2-4 y 7b) va en este `try`, y la limpieza
    # del paso 8 va en su `finally` — no después del bloque, sino *dentro* de él. Si algo de aquí
    # lanza una excepción no prevista, el `finally` corre de todos modos y las tarifas/marcas/
    # tramos de esta corrida no quedan a medio confirmar en la base real sin ningún aviso.
    # Comprobado inyectando un `raise` temporal justo antes de la limpieza y confirmando por SQL
    # que la base queda igual de limpia (ver el reporte de esta tarea).
    try:
        async with SessionLocal() as db:
            fallas += await _paso_1_sin_confirmar(db)
            # 7a corre AQUÍ, antes de confirmar nada: es la mitad "antes" de la comprobación 7
            # (columnas anuales de B-05 vacías sin marcas ni tarifa EJERCICIO). Su mitad
            # "después" (7b) corre al final, una vez que el paso 4 confirmó también el subsidio.
            fallas += await _paso_7a_b05_sin_confirmar(db)

            resultado_import, f = await _paso_2_importar_y_confirmar_base(db)
            fallas += f

            resultado_sin_subsidio, f = await _paso_3_sin_subsidio(db)
            fallas += f

            resultado_final, f = await _paso_4_todo_confirmado(db)
            fallas += f

            fallas += _paso_5_comprobacion_manual(resultado_final)
            fallas += _paso_6_libro_excel(resultado_sin_subsidio, resultado_final)

            fallas += await _paso_7b_b05_confirmado(db)
    except Exception as exc:  # noqa: BLE001 - cualquier excepción no prevista es un defecto que reportar, no esconder
        traceback.print_exc()
        fallas.append(f"Excepción no manejada durante la verificación: {type(exc).__name__}: {exc}")
    finally:
        async with SessionLocal() as db:
            fallas += await _paso_8_revertir(db, fotografia_antes)

    if fallas:
        print("\nFALLAS:")
        for falla in fallas:
            print("  -", falla)
        return 1
    print("\nTodas las comprobaciones pasaron, y la base quedó como se encontró.")
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
