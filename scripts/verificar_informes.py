"""Verificación en vivo de los informes de nómina contra los datos reales (spec §13, §14).

Comprueba las 9 identidades de B-00 (fuente: `Hub_CFDI_docs/00-fuentes/especificacion-
informes-cfdi.md`, sección "B-00 · Definiciones comunes al grupo") sobre los CFDI de
nómina ya normalizados, que B-02 produce filas y columnas dinámicas consistentes con esos
mismos datos, y (desde la fase 2) que **los nueve informes del catálogo** —no solo B-02—
corren sin lanzar, no salen vacíos por error, enmascaran toda columna `sensible=True` y **no
dejan ninguna CURP ni ningún NSS de ningún empleado en ninguna celda de ninguna de las cuatro
hojas**.

Lo que la fase 3 agregó, y por qué
-------------------------------------
Hasta la fase 2 este script comprobaba que **ningún informe lanzara**. Eso deja pasar el fallo
que la fase 3 encontró tres veces: un informe que corre, no lanza y **degrada mintiendo** —dice
que falta capturar un valor que ya está capturado, o manda a arreglar un hueco que no es el que
tiene—. Un aviso que apunta a la causa equivocada es peor que ninguno: manda a una persona a
mirar donde no está el problema y se lee como que el sistema no sabe lo que dice.

Así que ahora, además:

- **`_estado_de_configuracion` lee la configuración real** (marcas capturadas y confirmadas,
  tramos de `param_fiscal`, `map_departamento`, la política laboral de la empresa) y de ahí se
  **derivan** las banderas que cada informe *debe* emitir y las que *no puede* emitir. Nada de
  esperar una lista fija de banderas: la lista se calcula del estado, así que la comprobación
  sigue siendo verdadera cuando el dueño del repo confirme sus valores.
- **La degradación se comprueba por su efecto, no solo por la bandera.** Cuando B-03 avisa de
  una marca sin confirmar, sus cuatro columnas dependientes tienen que salir **vacías** en las
  filas de ese tipo; cuando B-05 avisa, «Gravado ordinario» tiene que salir vacía. Una bandera
  con las columnas llenas significaría que el aviso es decorativo.
- **B-08 cuadra por suma, sin tolerancia** (`_verificar_cuadre_de_b08`): «Provisión total» es la
  suma exacta de las tres columnas de provisión visibles de su fila. Es la cifra que puede acabar
  en estados financieros y la hoja tiene que poder auditarse arrastrando la columna.
- **B-08 declara su cobertura** (`_verificar_cobertura_de_b08`): con la fecha de corte al cierre
  del ejercicio sobre un histórico que se queda corto tiene que emitir
  `COBERTURA_INCOMPLETA_AL_CORTE` **con el número real de días sin datos**, calculado aquí desde
  la base y no leído del mensaje; y con la fecha de corte en el último periodo timbrado tiene que
  salir **sin banderas**. Las dos corridas se imprimen con su pasivo total.
- **El camino confirmado se ejercita en una transacción que se revierte**
  (`_verificar_camino_confirmado_con_reversion`). Las 44 marcas de percepción siguen sin
  confirmar a propósito —las confirma el dueño del repo desde la pantalla, es su
  responsabilidad—, así que la única forma honesta de comprobar que la rama que *calcula*
  funciona es confirmarlas en una transacción, mirar el resultado y **revertirla**. La reversión
  se comprueba en una sesión nueva y se imprime: una simulación que dejara la base tocada sería
  peor que no simular.

**Lo que este script no hardcodea, y sigue sin hardcodear.** Ni cuántas filas, ni cuántas
columnas dinámicas, ni el importe del pasivo, ni cuántos días de cobertura faltan: todo eso
cambia en cuanto se descargue más historia del SAT y una expectativa fija se volvería una falla
falsa a la primera descarga. Lo que sí se exige son **identidades** (las nueve de B-00, el cuadre
por suma de B-08) y **correspondencias** (que la bandera emitida sea la de la causa que la base
realmente tiene). Las cifras se imprimen para que quien corre el script las compare contra lo
que espera.

Esa última comprobación se agregó en la revisión final de la fase 2 y es la razón por la que
este script no vio la fuga más grave del catálogo: comprobaba el **mecanismo** (que las columnas
declaradas `sensible=True` salieran enmascaradas) y nunca el **resultado** (que las columnas no
sensibles no trajeran el dato personal dentro de una frase). B-10 interpolaba la CURP y el NSS en
su columna "Descripción del hallazgo" —no sensible, y que no puede serlo sin volverse ilegible—,
así que 6 de 7 filas salían con el dato completo con `enmascarar_datos_personales=True`.

**Ahora compara por patrón Y por valor, sobre las cuatro hojas.** La red vive en
`validadores.fugas_de_datos_personales_en_libro`, compartida con la prueba anti-fuga de
`tests/test_informe_b10.py`. Hasta el cierre de la fase 2 este script solo buscaba por **patrón**
y solo en la hoja `Datos`, con el argumento de que no debía manejar CURP ni NSS reales. Ese
argumento se descartó a propósito: el script **ya** lee la BD real (calcula las identidades de
B-00 sobre los datos de nómina, que incluyen al receptor), así que ya tiene acceso; compararlos en
memoria sin imprimirlos es seguro; y es la **única** forma de atrapar una CURP **mal formada**
interpolada en un mensaje, que por definición no coincide con el patrón de una CURP. Las hojas
`Banderas`, `Parámetros` y `Diccionario` se auditan porque viajan en el mismo archivo: un mensaje
de bandera que interpolara un dato personal salía igual de la empresa, y los mensajes de bandera
son texto libre.

**Contrato duro que eso no cambia:** este script **nunca imprime** una CURP, un NSS ni una cuenta
bancaria. Cuando encuentra una fuga reporta la hoja, la fila, la columna y el tipo de dato —nunca
el valor— porque se corre en una terminal cuyo historial queda guardado.

Las identidades **no se implementan aquí**: viven en `app/informes/identidades_b00.py` y
`tests/test_identidades_b00.py` las corre en cada pasada de la suite sobre XML sintéticos.
Este script es el otro llamador de esa misma implementación, el que la ejerce contra los
CFDI reales. Antes las identidades vivían solo en este archivo, fuera de `testpaths`, así
que nada las mantenía verdes.

Deliberadamente **no hardcodea** cuántas nóminas, filas o columnas dinámicas esperar: eso
quedaría obsoleto en cuanto se descargue más historia del SAT. Lo que sí es fijo son las
identidades contables — un XML timbrado por el SAT las cumple por construcción, así que
si alguna falla aquí es un bug de nuestro ETL, nunca un dato malo.

No imprime CURP, NSS ni cuenta bancaria: son datos personales de personas reales y esto
se corre en una terminal cuyo historial queda guardado.

Uso: `python scripts/verificar_informes.py` (dentro del contenedor `api`, o con el `.venv`
del host apuntando a la misma base). Sale con código 1 si alguna comprobación falla.
"""

from __future__ import annotations

import asyncio
import io
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl

from app.db.session import SessionLocal
from app.informes import b02_conceptos_patron as b02
from app.informes import b08_pasivo_laboral as b08
from app.informes import excel, identidades_b00, registro, validadores
from app.informes.base import Bandera, ContextoInforme, ResultadoInforme
from app.models.comprobante import Comprobante
from app.models.configuracion_fiscal import (
    CatalogoPercepcionMarca,
    ConfiguracionEmpresa,
    MapDepartamento,
    ParamFiscal,
    TablaVacaciones,
)
from app.models.empresa import Empresa
from app.models.enums import EstatusCfdi
from app.models.nomina import Nomina, NominaPercepcion, NominaReceptor
from app.services import configuracion_fiscal as cfg
from app.services import normalizacion

_ESTATUS_CANCELADO = EstatusCfdi.CANCELADO
"""El estatus que todos los informes del grupo excluyen por omisión (R-T1). Se nombra aquí para
que las consultas de este script apliquen **el mismo** filtro y no esperen una bandera por un
tipo de percepción que ningún informe llegó a ver."""

EMPRESA_ID = 11

EJERCICIO = 2026
"""Ejercicio sobre el que se corren los informes. Es el único año con nómina descargada; se
declara con nombre porque lo usan los parámetros mínimos, el estado de la configuración y las
dos corridas de cobertura de B-08, y tenían que ser el mismo."""

_INICIO_DEL_EJERCICIO = date(EJERCICIO, 1, 1)
_CIERRE_DEL_EJERCICIO = date(EJERCICIO, 12, 31)

_HOJAS_ESPERADAS = {"Datos", "Parámetros", "Banderas", "Diccionario"}
_MASCARA_RE = re.compile(r"^\*{4}([^*]{4})?$")  # `enmascarar()`: "****" o "****" + últimos 4

# Grano garantizado no vacío cuando hay nóminas normalizadas en el rango: una fila por
# CFDI (B-01, B-02), por nodo de percepción (B-03), por empleado (B-04, B-05) o por
# (empleado, tipo, clave) de deducción (B-07). B-10 tiene grano de *hallazgo*: cero filas ahí
# significa "sin hallazgos", un resultado legítimo que no debe tratarse como falla — solo se
# avisa.
#
# B-03 entra en la lista aunque sus columnas de topes salgan vacías mientras la configuración
# fiscal esté sin confirmar: lo que degrada son cuatro columnas, no el grano. Un CFDI de
# nómina normalizado siempre trae al menos un nodo de percepción, así que cero filas ahí es
# una falla igual que en B-02.
#
# **B-08 no está aquí a propósito, y no es una omisión:** su grano depende de la configuración.
# Con una percepción sin clasificar o sin la política laboral capturada, cero filas es el
# comportamiento correcto y documentado (sus dos puertas). Así que su expectativa no es fija:
# se **deriva** del estado real de la configuración en `_fallas_de_degradacion`, que exige filas
# cuando las dos puertas están abiertas y exige cero filas más la bandera de la puerta cuando no.
_CLAVES_CON_FILAS_GARANTIZADAS = {"B-01", "B-02", "B-03", "B-04", "B-05", "B-07"}


_MAX_FUGAS_REPORTADAS = 10
"""Cuántas fugas se detallan por informe antes de resumir el resto. Una fuga real suele afectar a
todas las filas de una columna; con cientos de empleados la lista completa sepultaría las demás
fallas del script — el mismo razonamiento que el colapso de `ESTATUS_NO_VERIFICADO`."""


async def _valores_personales_del_universo(db: AsyncSession) -> dict[str, list[str]]:
    """CURP y NSS de **todos** los receptores de nómina de la empresa, para la comprobación por
    valor de `validadores.fugas_de_datos_personales_en_libro`.

    **Se leen a propósito, y no se imprimen nunca.** Es el único modo de atrapar una CURP mal
    formada interpolada en un mensaje: por definición no coincide con el patrón de una CURP, así
    que el detector estructural no puede verla. Y se recogen de **todo** el universo, no de la fila
    que se está auditando, porque una CURP de otro empleado en la celda de un tercero se escapaba
    de las dos redes a la vez. Solo viven en memoria y solo se usan como aguja de una búsqueda de
    subcadena; lo que sale a la consola es el conteo, nunca un valor.
    """
    consulta = (
        select(NominaReceptor.curp, NominaReceptor.nss)
        .join(Comprobante, Comprobante.comprobante_id == NominaReceptor.comprobante_id)
        .where(Comprobante.empresa_id == EMPRESA_ID)
        .distinct()
    )
    curps: set[str] = set()
    nss: set[str] = set()
    for curp, numero_seguridad in (await db.execute(consulta)).all():
        if curp:
            curps.add(str(curp))
        if numero_seguridad:
            nss.add(str(numero_seguridad))
    print(f"Datos personales del universo para la comprobación por valor: {len(curps)} CURP y {len(nss)} NSS (no se imprimen)")
    return {"CURP": sorted(curps), "NSS": sorted(nss)}


async def _verificar_identidades_b00(db: AsyncSession) -> tuple[int, list[str]]:
    """Corre las identidades del módulo compartido y las reporta por consola."""
    v = await identidades_b00.verificar(db, EMPRESA_ID)

    print(f"CFDI de nómina normalizados: {v.comprobantes}")
    if not v.comprobantes:
        print("FALLA: no hay nóminas normalizadas; ¿corrió el reproceso (normalizacion_lote.normalizar_lote)?")
        return 0, ["no hay CFDI de nómina normalizados"]

    # `cotejos` es lo que hace auditable a la propia verificación: sin él, "cero fallas" no
    # distingue entre "todo cuadra" y "no se comprobó nada". Se imprime el número real, no el
    # teórico: un atributo ausente en el XML no se compara y por tanto no cuenta.
    print(f"Cotejos ejecutados: {v.cotejos} (máximo {identidades_b00.COTEJOS_POR_COMPROBANTE_COMPLETO} por CFDI)")
    if v.cotejos < v.comprobantes * identidades_b00.IDENTIDADES_POR_COMPROBANTE:
        print(
            f"AVISO: se esperaban al menos {v.comprobantes * identidades_b00.IDENTIDADES_POR_COMPROBANTE} cotejos. "
            "Algún atributo del complemento no vino en el XML; revisa qué CFDI y por qué."
        )

    return v.comprobantes, v.fallas


async def _verificar_b02(db: AsyncSession, comprobantes_normalizados: int) -> list[str]:
    fallas: list[str] = []

    resultado = await b02.consultar(
        db,
        EMPRESA_ID,
        b02.Parametros(fecha_desde=_INICIO_DEL_EJERCICIO, fecha_hasta=_CIERRE_DEL_EJERCICIO),
    )
    dinamicas = [c.titulo for c in resultado.columnas if b02.SEPARADOR_ETIQUETA in c.titulo]
    print(f"\nB-02: {len(resultado.filas)} filas, {len(resultado.columnas)} columnas totales, {len(dinamicas)} columnas dinámicas, {len(resultado.banderas)} banderas")
    print("Conceptos detectados (naturaleza¦tipo¦clave¦concepto):")
    for titulo in dinamicas:
        print("  ", titulo)

    if len(resultado.filas) != comprobantes_normalizados:
        fallas.append(
            f"B-02 devolvió {len(resultado.filas)} filas para {comprobantes_normalizados} CFDI de nómina "
            f"normalizados en el rango {_INICIO_DEL_EJERCICIO}/{_CIERRE_DEL_EJERCICIO}"
        )
    if not dinamicas:
        fallas.append("B-02 no produjo ninguna columna dinámica; ¿hay datos en nomina_percepcion/deduccion/otro_pago?")
    if not resultado.diccionario:
        fallas.append("B-02 no produjo entradas de Diccionario para las columnas dinámicas detectadas")

    # Caso real de colisión que B-02.R3 debe resolver: `099` como clave de deducción y de
    # otro pago a la vez ("Ajuste al neto" en ambas naturalezas). Si aparece, confirma que
    # el prefijo de naturaleza en la etiqueta evita que se confundan; si no aparece en este
    # rango de datos no es una falla — es informativo.
    claves_por_tipo: dict[str, set[str]] = {"P": set(), "O": set(), "D": set()}
    for titulo in dinamicas:
        partes = titulo.split(b02.SEPARADOR_ETIQUETA)
        if len(partes) >= 3:
            claves_por_tipo.setdefault(partes[0], set()).add(partes[2])
    colisiones = claves_por_tipo["D"] & claves_por_tipo["O"]
    if colisiones:
        print(f"Claves compartidas entre deducción y otro pago (naturaleza distinta, R3 en acción): {sorted(colisiones)}")

    if resultado.banderas:
        print("\nBanderas de B-02:")
        for bandera in resultado.banderas:
            print(f"  [{bandera.severidad}] {bandera.clave} · {bandera.ambito} · {bandera.mensaje}")
        if any(bandera.clave == "TOTALES_DESCUADRADOS" for bandera in resultado.banderas):
            fallas.append(
                "B-02 emitió TOTALES_DESCUADRADOS sobre CFDI reales timbrados por el SAT: es un bug del ETL, no un hallazgo del informe"
            )
    else:
        print("\nB-02 no emitió banderas.")

    return fallas


def _parametros_minimos(clave: str) -> dict[str, object]:
    """Parámetros mínimos —solo lo requerido, todo lo demás con su default declarado,
    incluido `enmascarar_datos_personales=True`— para ejercitar cada informe del catálogo
    contra el histórico completo de la empresa.

    El `else` final asume que un informe se acota por rango de fechas, que es lo que hacen seis
    de los nueve. Los que no, van nombrados arriba. **Un informe nuevo con parámetros requeridos
    distintos revienta aquí en `Parametros(**...)`, y el mensaje se lee como un defecto del
    informe cuando es de esta función** — le pasó a B-08, que exige `ejercicio` y `fecha_corte`,
    y el resultado fue que el único auditor de fugas contra datos reales nunca miró su libro.
    Si agregas un informe, agrégalo aquí.
    """
    if clave == "B-05":
        return {"ejercicio": EJERCICIO}
    if clave == "B-08":
        # `fecha_corte` al cierre del ejercicio: es la fecha a la que un auditor pide la provisión.
        return {"ejercicio": EJERCICIO, "fecha_corte": _CIERRE_DEL_EJERCICIO}
    return {"fecha_desde": _INICIO_DEL_EJERCICIO, "fecha_hasta": _CIERRE_DEL_EJERCICIO}


def _resumen_de_banderas(banderas: list[Bandera]) -> str:
    """`"[alta] MARCA_SIN_CONFIRMAR·tipo:001"` por bandera, en una línea.

    **Se publica la clave, la severidad y el ámbito, nunca el mensaje.** No es economía de
    espacio: los mensajes son texto libre y ya hubo una fuga real por interpolación (B-10 metía
    CURP y NSS en el suyo). Lo que hace falta para juzgar una corrida es *qué* banderas salieron;
    quien necesite el texto abre el libro. Los ámbitos por empleado llevan RFC —decisión explícita
    del módulo, la hoja `Banderas` no se enmascara— y nunca CURP ni NSS.
    """
    if not banderas:
        return "(ninguna)"
    return "  ".join(f"[{b.severidad}] {b.clave}·{b.ambito}" for b in banderas)


def _indice_de_columnas(resultado: ResultadoInforme) -> dict[str, int]:
    """Título de columna → posición, para leer una celda por nombre y no por número.

    Un índice literal en la comprobación se rompe en silencio en cuanto alguien inserta una
    columna: seguiría leyendo una celda, la equivocada, y la comprobación pasaría.
    """
    return {columna.titulo: i for i, columna in enumerate(resultado.columnas)}


def _celda_vacia(valor: object) -> bool:
    """Si una celda cuenta como «sin calcular». `None` y la cadena vacía; el cero **no**, que
    es justo la distinción que los informes de esta familia mantienen a mano (un cero afirma
    que no se debe nada, un vacío dice que no se pudo saber)."""
    return valor is None or valor == ""


async def _verificar_informe_del_catalogo(
    db: AsyncSession, clave: str, comprobantes_normalizados: int, valores_personales: dict[str, list[str]]
) -> tuple[ResultadoInforme | None, list[str]]:
    """Corre un informe del catálogo end-to-end (consulta + libro de Excel) y comprueba
    que no lance, que su grano no salga vacío por error, que ninguna columna declarada
    `sensible=True` se cuele sin enmascarar al archivo que circula por correo, y que ninguna de las
    cuatro hojas lleve la CURP o el NSS de ningún empleado del universo.

    Devuelve el `ResultadoInforme` (o `None` si lanzó) para que las comprobaciones de degradación
    lo reusen: volver a correr los nueve informes para mirarles las banderas duplicaría el trabajo
    y, peor, podría mirar una corrida distinta de la que se auditó.
    """
    definicion = registro.obtener(clave)

    try:
        p = definicion.Parametros(**_parametros_minimos(clave))
        resultado = await definicion.consultar(db, EMPRESA_ID, p)
    except Exception as exc:  # noqa: BLE001 — se reporta cualquier excepción, no se re-lanza
        return None, [f"{clave} lanzó al generarse: {type(exc).__name__}: {exc}"]

    fallas: list[str] = []
    print(f"\n{clave}: {len(resultado.filas)} filas, {len(resultado.columnas)} columnas, {len(resultado.banderas)} banderas, {len(resultado.notas)} notas")
    print(f"  banderas: {_resumen_de_banderas(resultado.banderas)}")
    if resultado.aviso:
        print(f"  aviso: {resultado.aviso}")

    if not resultado.filas:
        mensaje = f"{clave} devolvió 0 filas con {comprobantes_normalizados} CFDI de nómina normalizados en el rango"
        if clave in _CLAVES_CON_FILAS_GARANTIZADAS and comprobantes_normalizados:
            fallas.append(mensaje)
        elif comprobantes_normalizados:
            print(f"  AVISO: {mensaje} (grano de hallazgo: puede ser un resultado legítimo, no se marca como falla)")

    ctx = ContextoInforme(
        clave=definicion.CLAVE,
        nombre=definicion.NOMBRE,
        usuario="verificacion@script",
        generado_en=datetime.now(timezone.utc).replace(tzinfo=None),
        parametros=p.model_dump(mode="json"),
        etl_version=normalizacion.ETL_VERSION,
    )
    libro = openpyxl.load_workbook(io.BytesIO(excel.escribir_libro(resultado, ctx)))

    if set(libro.sheetnames) != _HOJAS_ESPERADAS:
        fallas.append(f"{clave} no tiene las cuatro hojas esperadas: {libro.sheetnames}")
        return resultado, fallas

    ws = libro["Datos"]
    filas_datos = list(ws.iter_rows(min_row=2, values_only=True))

    columnas_sensibles = [i for i, columna in enumerate(resultado.columnas) if columna.sensible]
    if columnas_sensibles:
        for fila in filas_datos:
            for idx in columnas_sensibles:
                valor = fila[idx] if idx < len(fila) else None
                if valor is None or valor == "":
                    continue
                if not _MASCARA_RE.match(str(valor)):
                    fallas.append(
                        f"{clave}: la columna sensible '{resultado.columnas[idx].titulo}' salió sin enmascarar "
                        "(¿escribir_libro no recibió enmascarar_datos_personales=True?)"
                    )
                    break
            if fallas:
                break

    # La comprobación de arriba verifica el **mecanismo** (que las columnas declaradas sensibles
    # salgan enmascaradas); esta verifica el **resultado**, que es lo que nadie comprobaba y por lo
    # que B-10 emitía CURP y NSS completos con el enmascaramiento activado: los interpolaba en el
    # texto de su columna "Descripción del hallazgo", que no es sensible ni puede serlo. Se auditan
    # las CUATRO hojas (viajan en el mismo archivo) y por las dos vías (patrón y valor); ver el
    # docstring del módulo y `validadores.fugas_de_datos_personales_en_libro`. Nunca se imprime el
    # valor encontrado —solo hoja, fila, columna y tipo— porque este script se corre en una terminal
    # cuyo historial queda guardado.
    fugas = validadores.fugas_de_datos_personales_en_libro(libro, valores_personales)
    for fuga in fugas[:_MAX_FUGAS_REPORTADAS]:
        fallas.append(f"{clave}: dato personal en el libro con enmascarar_datos_personales=True — {fuga.descripcion}")
    if len(fugas) > _MAX_FUGAS_REPORTADAS:
        fallas.append(f"{clave}: y {len(fugas) - _MAX_FUGAS_REPORTADAS} fuga(s) más de datos personales en el mismo libro")

    return resultado, fallas


async def _verificar_catalogo(
    db: AsyncSession, comprobantes_normalizados: int, valores_personales: dict[str, list[str]]
) -> tuple[dict[str, ResultadoInforme], list[str]]:
    """Recorre los **nueve** informes del catálogo: el registro es la fuente de verdad de lo que
    existe, así que un informe nuevo entra a esta comprobación en cuanto se registra en
    `app.informes.registro`, sin tocar este script.

    Devuelve los resultados por clave —sin la de los que lanzaron— para que
    `_fallas_de_degradacion` mire **estas mismas** corridas y no unas nuevas.
    """
    fallas: list[str] = []
    resultados: dict[str, ResultadoInforme] = {}
    for clave in sorted(registro.REGISTRO):
        resultado, fallas_informe = await _verificar_informe_del_catalogo(
            db, clave, comprobantes_normalizados, valores_personales
        )
        fallas.extend(fallas_informe)
        if resultado is not None:
            resultados[clave] = resultado
    return resultados, fallas


# --------------------------------------------------------------------------------------
# El estado real de la configuración fiscal: de aquí se derivan las expectativas
# --------------------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class _EstadoConfiguracion:
    """La configuración fiscal tal como está en la base al momento de correr.

    **Es lo que decide qué degradación es la verdadera**, y por eso se lee en vez de escribirse.
    Una lista fija de banderas esperadas se volvería falsa en cuanto el dueño del repo confirme
    sus valores —que es justo lo que tiene que poder hacer sin que esta verificación se rompa—, y
    peor: no distinguiría «avisó de lo que le falta» de «avisó de otra cosa», que es exactamente el
    fallo que la fase 3 encontró tres veces. Con el estado en la mano, «B-03 dice que falta
    capturar la UMA» es comprobablemente falso cuando la UMA está capturada y confirmada.
    """

    marcas_capturadas: int
    marcas_confirmadas: int
    marcas_con_nota: int
    marcas_con_multiplicador_no_derivable: int
    tipos_capturados: frozenset[str]
    tipos_confirmados: frozenset[str]
    tipos_observados: frozenset[str]
    """Tipos de percepción con al menos un nodo en el ejercicio, sin contar los cancelados."""
    params_capturados: frozenset[str]
    params_confirmados: frozenset[str]
    renglones_map_departamento: int
    renglones_tabla_vacaciones: int
    percepciones_sin_clasificar: int
    zona_salarial: str | None
    dias_aguinaldo: int | None
    factor_prima_vacacional: Decimal | None
    ultimo_pago_del_ejercicio: date | None

    @property
    def tipos_sin_confirmar(self) -> frozenset[str]:
        """Capturados y **sin confirmar**: el estado que tiene hoy la instalación real, y el que
        obliga a B-03 y B-05 a avisar «confírmala» en vez de «captúrala»."""
        return self.tipos_capturados - self.tipos_confirmados

    @property
    def tipos_observados_sin_confirmar(self) -> frozenset[str]:
        return self.tipos_observados & self.tipos_sin_confirmar

    @property
    def puertas_de_b08_abiertas(self) -> bool:
        """Las dos puertas de B-08: clasificación completa de percepciones y política laboral
        capturada. Con las dos abiertas el informe **tiene** que generar filas; con cualquiera
        cerrada **tiene** que no generar ninguna y decir cuál falta."""
        return (
            self.percepciones_sin_clasificar == 0
            and self.dias_aguinaldo is not None
            and self.factor_prima_vacacional is not None
        )


async def _estado_de_configuracion(db: AsyncSession) -> _EstadoConfiguracion:
    """Lee y **publica** el estado de la configuración. Se imprime completo porque sin él la
    salida del script no es interpretable: «B-03 emitió MARCA_SIN_CONFIRMAR» solo se puede juzgar
    sabiendo cuántas marcas hay capturadas y cuántas confirmadas.

    Nada de lo que se lee aquí es un dato personal: son marcas de catálogo, importes fiscales
    publicados en el DOF y conteos de renglones de configuración.
    """
    marcas = list((await db.scalars(select(CatalogoPercepcionMarca))).all())
    tipos_capturados = frozenset(str(m.tipo_percepcion) for m in marcas)
    tipos_confirmados = frozenset(str(m.tipo_percepcion) for m in marcas if m.confirmado_en is not None)

    tramos = list((await db.scalars(select(ParamFiscal))).all())
    params_capturados = frozenset(str(t.clave) for t in tramos)
    params_confirmados = frozenset(str(t.clave) for t in tramos if t.confirmado_en is not None)

    # Los tipos de percepción que los informes van a ver. Se excluyen los cancelados, que es el
    # default de todos los informes del grupo (R-T1): incluirlos haría esperar una bandera por un
    # tipo que ninguna fila trae.
    observados = (
        await db.execute(
            select(NominaPercepcion.tipo_percepcion)
            .join(Comprobante, Comprobante.comprobante_id == NominaPercepcion.comprobante_id)
            .join(Nomina, Nomina.comprobante_id == Comprobante.comprobante_id)
            .where(
                Comprobante.empresa_id == EMPRESA_ID,
                Comprobante.tipo_comprobante == "N",
                Comprobante.estatus != _ESTATUS_CANCELADO,
                Nomina.fecha_pago.between(_INICIO_DEL_EJERCICIO, _CIERRE_DEL_EJERCICIO),
            )
            .distinct()
        )
    ).scalars()
    tipos_observados = frozenset(str(t) for t in observados)

    renglones_departamento = (
        await db.execute(select(func.count()).select_from(MapDepartamento).where(MapDepartamento.empresa_id == EMPRESA_ID))
    ).scalar_one()
    renglones_vacaciones = (await db.execute(select(func.count()).select_from(TablaVacaciones))).scalar_one()

    ultimo_pago = (
        await db.execute(
            select(func.max(Nomina.fecha_pago))
            .join(Comprobante, Comprobante.comprobante_id == Nomina.comprobante_id)
            .where(
                Comprobante.empresa_id == EMPRESA_ID,
                Comprobante.tipo_comprobante == "N",
                Comprobante.estatus != _ESTATUS_CANCELADO,
                Nomina.fecha_pago.between(_INICIO_DEL_EJERCICIO, _CIERRE_DEL_EJERCICIO),
            )
        )
    ).scalar()

    empresa = await db.get(Empresa, EMPRESA_ID)
    if empresa is None:
        raise SystemExit(f"La empresa {EMPRESA_ID} no existe en esta base: revisa a qué base apunta el `.env`.")
    sin_clasificar = cfg.percepciones_sin_clasificar(await cfg.observados_de_empresa(db, empresa))
    config = await db.get(ConfiguracionEmpresa, EMPRESA_ID)

    estado = _EstadoConfiguracion(
        marcas_capturadas=len(marcas),
        marcas_confirmadas=len(tipos_confirmados),
        marcas_con_nota=sum(1 for m in marcas if m.nota_revision),
        marcas_con_multiplicador_no_derivable=sum(1 for m in marcas if m.multiplicador_no_derivable),
        tipos_capturados=tipos_capturados,
        tipos_confirmados=tipos_confirmados,
        tipos_observados=tipos_observados,
        params_capturados=params_capturados,
        params_confirmados=params_confirmados,
        renglones_map_departamento=int(renglones_departamento),
        renglones_tabla_vacaciones=int(renglones_vacaciones),
        percepciones_sin_clasificar=len(sin_clasificar),
        zona_salarial=config.zona_salarial.value if config is not None and config.zona_salarial is not None else None,
        dias_aguinaldo=config.dias_aguinaldo if config is not None else None,
        factor_prima_vacacional=config.factor_prima_vacacional if config is not None else None,
        ultimo_pago_del_ejercicio=ultimo_pago,
    )

    print("\n--- Estado de la configuración fiscal (de aquí se derivan las expectativas) ---")
    print(
        f"Marcas de percepción: {estado.marcas_capturadas} capturadas, {estado.marcas_confirmadas} confirmadas, "
        f"{estado.marcas_con_nota} con nota de revisión, "
        f"{estado.marcas_con_multiplicador_no_derivable} con multiplicador no derivable"
    )
    print(f"Tipos de percepción observados en {EJERCICIO}: {sorted(estado.tipos_observados)}")
    print(f"  de ellos, capturados y SIN confirmar: {sorted(estado.tipos_observados_sin_confirmar)}")
    print(f"  de ellos, SIN marca capturada: {sorted(estado.tipos_observados - estado.tipos_capturados)}")
    print(f"Valores de param_fiscal confirmados: {sorted(estado.params_confirmados)}")
    print(f"  capturados sin confirmar: {sorted(estado.params_capturados - estado.params_confirmados)}")
    print(
        f"Empresa {EMPRESA_ID}: zona salarial {estado.zona_salarial}, {estado.dias_aguinaldo} días de aguinaldo, "
        f"factor de prima {estado.factor_prima_vacacional}"
    )
    print(
        f"map_departamento: {estado.renglones_map_departamento} renglón(es) · "
        f"percepciones sin clasificar: {estado.percepciones_sin_clasificar} · "
        f"tabla_vacaciones: {estado.renglones_tabla_vacaciones} renglón(es)"
    )
    print(f"Último CFDI de nómina pagado del ejercicio: {estado.ultimo_pago_del_ejercicio}")
    print(f"Puertas de B-08 abiertas (clasificación completa + política laboral): {estado.puertas_de_b08_abiertas}")
    return estado


# --------------------------------------------------------------------------------------
# La degradación, comprobada por su causa y por su efecto
# --------------------------------------------------------------------------------------


def _claves(banderas: list[Bandera]) -> set[str]:
    return {b.clave for b in banderas}


def _ambitos(banderas: list[Bandera], clave: str) -> set[str]:
    return {b.ambito for b in banderas if b.clave == clave}


def _mensajes(banderas: list[Bandera], clave: str) -> list[str]:
    """Los mensajes de una bandera, **para inspeccionar en memoria y no para imprimir**: es texto
    libre y ya hubo una fuga por interpolación. Lo que se publica de aquí son conclusiones
    derivadas («el aviso no cita los días reales»), nunca el texto."""
    return [b.mensaje for b in banderas if b.clave == clave]


def _columnas_no_vacias(resultado: ResultadoInforme, titulo: str, filas: list[list[object]] | None = None) -> int:
    idx = _indice_de_columnas(resultado)
    candidatas = resultado.filas if filas is None else filas
    return sum(1 for fila in candidatas if not _celda_vacia(fila[idx[titulo]]))


_COLUMNAS_DE_TOPE_DE_B03 = ("Base de exención", "Tope de exención", "Exceso sobre el tope")
"""Las columnas de B-03 que **dependen de la marca** del tipo de la fila, y que por tanto tienen
que salir vacías mientras la marca no esté confirmada.

**«UMA aplicable» no está aquí, y eso es la comprobación, no un descuido.** Sale de
`param_fiscal.UMA_DIARIA`, que hoy **sí** está confirmada, así que se llena con la marca sin
confirmar: la degradación es exactamente del ancho que debe ser —lo que la marca decide queda
vacío, lo que la UMA decide se calcula— y ver «UMA aplicable» llena con las otras tres vacías es
la prueba de que las dos puertas son independientes y ninguna arrastra a la otra."""


def _fallas_de_degradacion(estado: _EstadoConfiguracion, resultados: dict[str, ResultadoInforme]) -> list[str]:
    """Comprueba que **la bandera emitida sea la de la causa que la base realmente tiene**, y que
    lo que la bandera dice que no se calculó **de verdad salga vacío**.

    Las dos mitades hacen falta. Sin la primera, un informe que grita «captura la UMA» con la UMA
    ya capturada y confirmada pasa la verificación y manda a una persona a mirar donde no está el
    problema —fue un defecto real de esta fase—. Sin la segunda, una bandera puede ser decorativa:
    avisa de que algo no se calculó mientras la columna sale llena, y entonces o la bandera sobra o
    la cifra es inventada.
    """
    fallas: list[str] = []

    # --- B-03: la marca sin confirmar, por tipo -------------------------------------------
    b03 = resultados.get("B-03")
    if b03 is not None and b03.filas:
        idx = _indice_de_columnas(b03)
        tipos_en_filas = {str(fila[idx["Tipo percepción"]]) for fila in b03.filas if fila[idx["Tipo percepción"]] is not None}
        # El universo esperado sale de las **filas del propio informe**, no de una consulta
        # aparte: así la igualdad exacta sigue siendo cierta aunque un CFDI se cancele o quede
        # sustituido (B-03 los excluye) y no hay que replicar aquí sus filtros.
        esperadas = {f"tipo:{tipo}" for tipo in tipos_en_filas & estado.tipos_sin_confirmar}
        obtenidas = _ambitos(b03.banderas, "MARCA_SIN_CONFIRMAR")
        if esperadas != obtenidas:
            fallas.append(
                f"B-03: MARCA_SIN_CONFIRMAR salió con ámbitos {sorted(obtenidas)} y la base dice que los tipos "
                f"con marca capturada y sin confirmar presentes en sus filas son {sorted(esperadas)}"
            )
        faltan_marca = tipos_en_filas - estado.tipos_capturados
        if not faltan_marca and "FALTA_MARCA" in _claves(b03.banderas):
            fallas.append(
                "B-03 emitió FALTA_MARCA («cárgala con la semilla») pero todos los tipos de sus filas tienen "
                "marca capturada: el aviso manda a capturar lo que ya está capturado, y lo que falta es el clic "
                "de confirmación"
            )
        # `MULTIPLICADOR_NO_DERIVABLE` afirma en su mensaje que la marca **está confirmada** y que
        # el hueco es del CFDI, no de la configuración. Sobre un tipo sin confirmar sería falso y
        # mandaría a revisar el sistema de nómina en vez de a dar el clic.
        sin_confirmar_con_multiplicador = {
            ambito for ambito in _ambitos(b03.banderas, "MULTIPLICADOR_NO_DERIVABLE")
            if ambito.removeprefix("tipo:") in estado.tipos_sin_confirmar
        }
        if sin_confirmar_con_multiplicador:
            fallas.append(
                f"B-03 emitió MULTIPLICADOR_NO_DERIVABLE en {sorted(sin_confirmar_con_multiplicador)}, cuyo mensaje "
                "afirma que la marca está confirmada, y esos tipos están sin confirmar"
            )
        # El efecto observable: lo que la bandera dice que no se calculó, vacío.
        filas_afectadas = [fila for fila in b03.filas if str(fila[idx["Tipo percepción"]]) in estado.tipos_sin_confirmar]
        for titulo in _COLUMNAS_DE_TOPE_DE_B03:
            llenas = _columnas_no_vacias(b03, titulo, filas_afectadas)
            if llenas:
                fallas.append(
                    f"B-03 avisó de marcas sin confirmar pero la columna dependiente '{titulo}' salió llena en "
                    f"{llenas} de {len(filas_afectadas)} filas de esos tipos: o el aviso es decorativo o la cifra "
                    "se calculó con un valor que nadie confirmó"
                )
        if filas_afectadas:
            print(
                f"\nB-03 · degradación comprobada: {len(filas_afectadas)} de {len(b03.filas)} filas son de tipos sin "
                f"confirmar; sus columnas {list(_COLUMNAS_DE_TOPE_DE_B03)} salen vacías y «UMA aplicable» llena en "
                f"{_columnas_no_vacias(b03, 'UMA aplicable', filas_afectadas)} (la UMA sí está confirmada)"
            )
        # La cara contraria: valores que **sí** están confirmados no pueden generar su bandera.
        for clave_param, banderas_prohibidas in _BANDERAS_QUE_UN_PARAM_CONFIRMADO_DESCARTA.items():
            if clave_param not in estado.params_confirmados:
                continue
            indebidas = banderas_prohibidas & _claves(b03.banderas)
            if indebidas:
                fallas.append(
                    f"B-03 emitió {sorted(indebidas)} y `{clave_param}` está capturado Y confirmado en "
                    "`param_fiscal`: el aviso manda a arreglar algo que ya está resuelto"
                )
        if estado.zona_salarial is not None and "FALTA_ZONA_SALARIAL" in _claves(b03.banderas):
            fallas.append(
                f"B-03 emitió FALTA_ZONA_SALARIAL y la empresa tiene zona salarial {estado.zona_salarial} "
                "configurada"
            )

    # --- B-05: la misma marca, con efecto en una sola columna ------------------------------
    b05 = resultados.get("B-05")
    if b05 is not None and b05.filas:
        claves_b05 = _claves(b05.banderas)
        if estado.tipos_observados_sin_confirmar and "MARCA_SIN_CONFIRMAR" not in claves_b05:
            fallas.append(
                f"B-05 no emitió MARCA_SIN_CONFIRMAR con {len(estado.tipos_observados_sin_confirmar)} tipo(s) "
                "observado(s) con marca capturada y sin confirmar"
            )
        if not (estado.tipos_observados - estado.tipos_capturados) and "FALTA_MARCA" in claves_b05:
            fallas.append(
                "B-05 emitió FALTA_MARCA («cárgalas con la semilla») y todos los tipos observados tienen marca "
                "capturada: la causa real es la confirmación, no la captura"
            )
        llenas = _columnas_no_vacias(b05, "Gravado ordinario")
        if "MARCA_SIN_CONFIRMAR" in claves_b05 and llenas:
            fallas.append(
                f"B-05 avisó de marcas sin confirmar y su mensaje afirma que «Gravado ordinario» salió vacía en "
                f"todas las filas, pero salió llena en {llenas} de {len(b05.filas)}"
            )
        if "MARCA_SIN_CONFIRMAR" in claves_b05:
            print(f"B-05 · degradación comprobada: «Gravado ordinario» vacía en las {len(b05.filas)} filas")

    # --- B-06: el mapeo de departamentos ---------------------------------------------------
    b06 = resultados.get("B-06")
    if b06 is not None and b06.filas:
        claves_b06 = _claves(b06.banderas)
        if estado.renglones_map_departamento == 0:
            if "DEPARTAMENTO_SIN_MAPEO" not in claves_b06:
                fallas.append(
                    "B-06 no emitió DEPARTAMENTO_SIN_MAPEO y la empresa no tiene ningún renglón en "
                    "`map_departamento`: el agrupamiento salió del texto crudo del CFDI sin constancia de eso"
                )
            # Sin un solo renglón de mapeo no puede haber dos valores que colisionen ni uno en
            # blanco: las dos banderas hablan del **valor configurado**, y no hay ninguno.
            imposibles = {"CENTRO_DE_COSTO_AMBIGUO", "CENTRO_DE_COSTO_EN_BLANCO"} & claves_b06
            if imposibles:
                fallas.append(
                    f"B-06 emitió {sorted(imposibles)}, que hablan del centro de costo **configurado**, con "
                    "`map_departamento` vacío: no hay ningún valor configurado del que quejarse"
                )
            print(
                f"B-06 · degradación comprobada: sin renglones en `map_departamento` emitió "
                f"{sorted(claves_b06)} y agrupó por el texto crudo del departamento"
            )
        elif "DEPARTAMENTO_SIN_MAPEO" in claves_b06:
            print(
                f"B-06 · AVISO: hay {estado.renglones_map_departamento} renglón(es) en `map_departamento` y aun "
                "así queda departamento sin mapear; revisa la cadena de la colación (hallazgo A1 del inventario)"
            )

    # --- B-08: las dos puertas -------------------------------------------------------------
    b08 = resultados.get("B-08")
    if b08 is not None:
        claves_b08 = _claves(b08.banderas)
        if estado.puertas_de_b08_abiertas:
            bloqueos = {"CLASIFICACION_INCOMPLETA", "FALTA_CONFIGURACION_DE_PROVISION"} & claves_b08
            if bloqueos:
                fallas.append(
                    f"B-08 emitió {sorted(bloqueos)} y sus dos puertas están abiertas: 0 percepciones sin "
                    f"clasificar, {estado.dias_aguinaldo} días de aguinaldo y factor "
                    f"{estado.factor_prima_vacacional} en la configuración de la empresa"
                )
            if not b08.filas:
                fallas.append("B-08 no generó ninguna fila con sus dos puertas abiertas y CFDI de nómina en el ejercicio")
            if b08.aviso:
                fallas.append(f"B-08 salió con aviso de no generación teniendo las dos puertas abiertas: {b08.aviso}")
        else:
            if estado.percepciones_sin_clasificar and "CLASIFICACION_INCOMPLETA" not in claves_b08:
                fallas.append(
                    f"B-08 no emitió CLASIFICACION_INCOMPLETA con {estado.percepciones_sin_clasificar} "
                    "percepción(es) sin categoría de provisión"
                )
            if (
                estado.dias_aguinaldo is None or estado.factor_prima_vacacional is None
            ) and "FALTA_CONFIGURACION_DE_PROVISION" not in claves_b08:
                fallas.append("B-08 no emitió FALTA_CONFIGURACION_DE_PROVISION sin política laboral capturada")
            if b08.filas:
                fallas.append(f"B-08 generó {len(b08.filas)} filas con una de sus dos puertas cerrada")
        # `FALTA_TABLA_VACACIONES` distingue «la tabla está vacía» de «falta este renglón», y la
        # primera versión afirmaba la primera también cuando era la segunda.
        if estado.renglones_tabla_vacaciones:
            for mensaje in _mensajes(b08.banderas, "FALTA_TABLA_VACACIONES"):
                if "No hay ningún renglón" in mensaje:
                    fallas.append(
                        f"B-08 dice que `tabla_vacaciones` no tiene ningún renglón y tiene "
                        f"{estado.renglones_tabla_vacaciones}: lo que falta es un renglón aplicable, no la tabla"
                    )

    return fallas


_BANDERAS_QUE_UN_PARAM_CONFIRMADO_DESCARTA: dict[str, frozenset[str]] = {
    "UMA_DIARIA": frozenset({"FALTA_UMA", "UMA_SIN_CONFIRMAR"}),
    "UMA_ANUAL": frozenset({"FALTA_UMA_ANUAL", "UMA_ANUAL_SIN_CONFIRMAR"}),
    "SALARIO_MINIMO_ZLFN": frozenset({"FALTA_SALARIO_MINIMO", "SALARIO_MINIMO_SIN_CONFIRMAR"}),
    "SALARIO_MINIMO_GENERAL": frozenset({"FALTA_SALARIO_MINIMO", "SALARIO_MINIMO_SIN_CONFIRMAR"}),
}
"""Qué banderas de B-03 dejan de ser ciertas cuando un valor de `param_fiscal` está confirmado.

`FALTA_X` dice «ve a buscar el valor» y `X_SIN_CONFIRMAR` dice «está ahí, dale el clic»: las dos
son falsas sobre un valor ya confirmado, y las dos mandan a alguien a la pantalla de configuración
a no hacer nada. Es la mitad más valiosa de la comprobación de degradación, porque un aviso de más
no rompe ninguna cifra y por eso nadie lo mira dos veces.

Las dos claves de salario mínimo comparten banderas a propósito: B-03 resuelve **una** de las dos
según la zona de la empresa (`cfg.clave_de_salario_minimo`), y como aquí las dos están confirmadas
la conclusión no depende de cuál. Si algún día solo una lo estuviera, esta tabla haría falsa la
comprobación en la zona no confirmada — el día que eso pase, hay que resolver la clave por la zona.
"""


# --------------------------------------------------------------------------------------
# B-08: el cuadre por suma y la declaración de cobertura
# --------------------------------------------------------------------------------------

_COLUMNAS_DE_PROVISION = ("Provisión de aguinaldo", "Provisión de vacaciones", "Provisión de prima vacacional")
_COLUMNA_DE_PROVISION_TOTAL = "Provisión total"

_MARGEN_DE_COBERTURA_ESPERADO = 31
"""Cuántos días puede la fecha de corte adelantarse al último CFDI observado antes de que B-08
tenga que declarar que el tramo es proyección. Un mes natural.

**Está escrito aquí y no importado de `b08.MARGEN_COBERTURA_DIAS` a propósito, y lo descubrió una
prueba de mutación.** La versión anterior de esta comprobación decía
`if dias_sin_datos > b08.MARGEN_COBERTURA_DIAS: exige la bandera`, que se lee bien y **no protege
nada del umbral**: subir la constante del informe a 9999 apaga la bandera *y* apaga la expectativa
a la vez, así que el informe dejaba de avisar de 169 días de proyección y el script seguía
diciendo «todas las comprobaciones pasaron». Un umbral que se lee del módulo que se está
verificando no es una expectativa, es un eco.

La duplicación va **atada**: si las dos constantes se separan, `_verificar_cobertura_de_b08`
reporta falla. Cambiar el margen del informe es legítimo —es una decisión de diseño— pero tiene
que ser una decisión, y entonces se cambia aquí también. Es el mismo trato que `b08` da a la regla
de resolución de `tabla_vacaciones`, duplicada y sujeta a una prueba que compara las dos copias.
"""


def _verificar_cuadre_de_b08(resultado: ResultadoInforme, etiqueta: str) -> tuple[Decimal, list[str]]:
    """«Provisión total» tiene que ser la suma **exacta** de las tres columnas de provisión de su
    fila, sin tolerancia. Devuelve el total de la columna y las fallas.

    Sin tolerancia y no por purismo: los tres sumandos son columnas **visibles** de la misma hoja,
    así que quien reciba el Excel puede arrastrarlas y comparar. Una diferencia de un centavo
    significa que el total incluye algo que no está en ninguna columna, y entonces la hoja no es
    auditable — que es lo único que separa esta cifra de un número inventado, porque es la que
    puede acabar reconocida en estados financieros. Todo el módulo trabaja en `Decimal` y el único
    redondeo del sistema está al escribir la celda, así que la igualdad exacta es alcanzable.

    Y la otra mitad: si falta cualquiera de los tres sumandos, el total tiene que salir **vacío**,
    nunca parcial. Un total parcial se ve idéntico a uno completo en la celda.
    """
    fallas: list[str] = []
    idx = _indice_de_columnas(resultado)
    total_de_la_columna = Decimal(0)
    vacios = 0
    for numero, fila in enumerate(resultado.filas, start=1):
        partes = [fila[idx[titulo]] for titulo in _COLUMNAS_DE_PROVISION]
        total = fila[idx[_COLUMNA_DE_PROVISION_TOTAL]]
        if any(parte is None for parte in partes):
            vacios += 1
            if total is not None:
                fallas.append(
                    f"B-08 ({etiqueta}) fila {numero}: «Provisión total» trae un importe y alguna de las tres "
                    "columnas de provisión está vacía, así que el total es parcial con apariencia de completo"
                )
            continue
        suma = sum((parte for parte in partes if isinstance(parte, Decimal)), Decimal(0))
        if total is None:
            fallas.append(
                f"B-08 ({etiqueta}) fila {numero}: las tres columnas de provisión traen importe y «Provisión "
                "total» salió vacía"
            )
            continue
        if not isinstance(total, Decimal) or total != suma:
            fallas.append(
                f"B-08 ({etiqueta}) fila {numero}: «Provisión total» = {total} y la suma exacta de las tres "
                f"columnas de provisión visibles es {suma} (diferencia {total - suma if isinstance(total, Decimal) else 'tipo no Decimal'})"
            )
            continue
        total_de_la_columna += total
    print(
        f"B-08 ({etiqueta}): {len(resultado.filas)} filas, {vacios} sin provisión total, "
        f"pasivo total {total_de_la_columna} (cuadre exacto por suma de las tres columnas: "
        f"{'sí' if not fallas else 'NO'})"
    )
    return total_de_la_columna, fallas


async def _verificar_cobertura_de_b08(db: AsyncSession, estado: _EstadoConfiguracion) -> list[str]:
    """Las dos corridas de B-08 que separan **proyección** de **observación**.

    El devengo del art. 87 LFT corre hasta la fecha de corte con independencia de lo que se haya
    descargado, así que un corte al cierre del ejercicio sobre un histórico que se queda en julio
    da una cifra correcta como devengo y **equívoca como observación**: es una proyección
    presentada como pasivo medido. La bandera `COBERTURA_INCOMPLETA_AL_CORTE` existe para decirlo,
    y aquí se comprueba que diga **el número real de días**, calculado desde la base y no leído
    del propio mensaje. La segunda corrida mueve el corte al último periodo timbrado y ahí la
    bandera no puede aparecer: todo lo devengado está observado.
    """
    fallas: list[str] = []
    definicion = registro.obtener("B-08")
    ultimo = estado.ultimo_pago_del_ejercicio
    if ultimo is None:
        print("\nB-08 · cobertura: no hay CFDI de nómina pagado en el ejercicio, no hay nada que comparar.")
        return fallas

    print("\n--- B-08 · proyección contra observación ---")
    dias_sin_datos = (_CIERRE_DEL_EJERCICIO - ultimo).days
    if b08.MARGEN_COBERTURA_DIAS != _MARGEN_DE_COBERTURA_ESPERADO:
        # El cable que hace que la constante duplicada no se pueda mover en silencio: ver
        # `_MARGEN_DE_COBERTURA_ESPERADO`, que existe justamente porque leerla del informe
        # convertía la comprobación del umbral en un eco.
        fallas.append(
            f"B-08 declara un margen de cobertura de {b08.MARGEN_COBERTURA_DIAS} días y esta verificación espera "
            f"{_MARGEN_DE_COBERTURA_ESPERADO} (un mes natural). Si el cambio es deliberado, actualiza "
            "`_MARGEN_DE_COBERTURA_ESPERADO`; hasta entonces no se sabe cuál de los dos números es el bueno"
        )

    proyectado = await definicion.consultar(
        db, EMPRESA_ID, definicion.Parametros(ejercicio=EJERCICIO, fecha_corte=_CIERRE_DEL_EJERCICIO)
    )
    print(f"corte {_CIERRE_DEL_EJERCICIO} (cierre del ejercicio): banderas {_resumen_de_banderas(proyectado.banderas)}")
    total_proyectado, fallas_cuadre = _verificar_cuadre_de_b08(proyectado, f"corte {_CIERRE_DEL_EJERCICIO}")
    fallas.extend(fallas_cuadre)

    claves_proyectado = _claves(proyectado.banderas)
    if dias_sin_datos > _MARGEN_DE_COBERTURA_ESPERADO:
        if "COBERTURA_INCOMPLETA_AL_CORTE" not in claves_proyectado:
            fallas.append(
                f"B-08 con corte al {_CIERRE_DEL_EJERCICIO} no emitió COBERTURA_INCOMPLETA_AL_CORTE y el último "
                f"CFDI observado se pagó el {ultimo}: {dias_sin_datos} días del devengo son proyección y el libro "
                "no lo dice en ninguna parte"
            )
        else:
            # Se comprueba el **número**, no la presencia: una bandera que avisa de la cobertura
            # con los días equivocados es tan inútil como no avisar, y peor porque parece cubierto.
            esperado = f"{dias_sin_datos} días sin datos"
            if not any(esperado in mensaje for mensaje in _mensajes(proyectado.banderas, "COBERTURA_INCOMPLETA_AL_CORTE")):
                fallas.append(
                    f"B-08: COBERTURA_INCOMPLETA_AL_CORTE no cita los {dias_sin_datos} días que van del último "
                    f"CFDI observado ({ultimo}) al corte ({_CIERRE_DEL_EJERCICIO})"
                )
            else:
                print(f"  COBERTURA_INCOMPLETA_AL_CORTE cita los {dias_sin_datos} días reales de proyección.")
    elif "COBERTURA_INCOMPLETA_AL_CORTE" in claves_proyectado:
        fallas.append(
            f"B-08 emitió COBERTURA_INCOMPLETA_AL_CORTE con solo {dias_sin_datos} días entre el último CFDI "
            f"observado y el corte, por debajo del margen de {_MARGEN_DE_COBERTURA_ESPERADO} días"
        )

    observado = await definicion.consultar(
        db, EMPRESA_ID, definicion.Parametros(ejercicio=EJERCICIO, fecha_corte=ultimo)
    )
    print(f"corte {ultimo} (último CFDI timbrado): banderas {_resumen_de_banderas(observado.banderas)}")
    total_observado, fallas_cuadre_observado = _verificar_cuadre_de_b08(observado, f"corte {ultimo}")
    fallas.extend(fallas_cuadre_observado)

    if "COBERTURA_INCOMPLETA_AL_CORTE" in _claves(observado.banderas):
        fallas.append(
            f"B-08 con corte al {ultimo} —el día del último CFDI timbrado— emitió COBERTURA_INCOMPLETA_AL_CORTE: "
            "no hay ni un día del devengo sin datos observados"
        )
    if observado.banderas:
        # No es falla: con más historia pueden aparecer banderas legítimas (una posible baja, un
        # periodo descartado del promedio). Se avisa porque hoy la corrida observada sale limpia y
        # que deje de estarlo es información.
        print(f"  AVISO: la corrida observada trae {len(observado.banderas)} bandera(s); hoy sale sin ninguna.")

    # El devengo es monótono en los días: la corrida al cierre cubre 169 días más y lo pagado que
    # se resta es el mismo (no hay pagos posteriores al último CFDI), así que no puede ser menor.
    if total_proyectado < total_observado:
        fallas.append(
            f"B-08: el pasivo proyectado al {_CIERRE_DEL_EJERCICIO} ({total_proyectado}) es menor que el "
            f"observado al {ultimo} ({total_observado}), y el devengo solo puede crecer con los días"
        )
    print(
        f"Pasivo proyectado al {_CIERRE_DEL_EJERCICIO}: {total_proyectado.quantize(Decimal('0.01'))} · "
        f"observado al {ultimo}: {total_observado.quantize(Decimal('0.01'))} · "
        f"{dias_sin_datos} días de diferencia son proyección"
    )
    return fallas


# --------------------------------------------------------------------------------------
# El camino confirmado, en una transacción que se revierte
# --------------------------------------------------------------------------------------


async def _verificar_camino_confirmado_con_reversion(estado: _EstadoConfiguracion) -> list[str]:
    """Ejercita la rama que **calcula** confirmando las marcas en una transacción y revirtiéndola.

    **Por qué se simula y no se confirma de verdad.** Confirmar una marca de exención es un acto de
    una persona que responde por el criterio fiscal que activa; las 44 marcas de esta base siguen
    sin confirmar **a propósito** y las confirma el dueño del repo desde la pantalla. Pero entonces
    la rama que calcula el tope nunca se ejercita contra datos reales, que es medio informe sin
    probar en vivo. La salida es una transacción: se confirma por el servicio real —con sus guardas
    de «las marcas cambiaron» y «la duda no estaba a la vista»—, se mira el resultado, y se
    revierte.

    **La huella de la nota se calcula de la propia fila**, lo que abre la guarda de `duda_no_vista`.
    Eso simula a alguien que leyó la duda; **no la responde**. Las 39 notas de revisión siguen
    siendo trabajo de una persona y este script no valida ni un criterio fiscal.

    **La reversión se comprueba en una sesión nueva**, no confiando en que el `rollback` hizo lo
    suyo: se vuelve a contar las marcas confirmadas leyendo de la base. Una simulación que dejara
    la base tocada sería peor que no simular — el estado de la configuración es lo que decide el
    resultado de todo lo demás.
    """
    fallas: list[str] = []
    tipos = sorted(estado.tipos_observados_sin_confirmar)
    print("\n--- Camino confirmado, simulado en una transacción que se revierte ---")
    if not tipos:
        print("No hay tipos observados con marca capturada y sin confirmar: no hay nada que simular.")
        return fallas
    print(f"Se confirmarán en la transacción los tipos {tipos} (por el servicio real, con sus guardas).")

    async with SessionLocal() as db:
        try:
            for tipo in tipos:
                fila = await db.get(CatalogoPercepcionMarca, tipo)
                if fila is None:  # imposible: los tipos vienen de esta misma tabla
                    fallas.append(f"No hay marcas capturadas para el tipo {tipo}, y el estado decía que sí")
                    continue
                _, cambio = await cfg.confirmar_marca_percepcion(
                    db,
                    tipo=tipo,
                    marcas=cfg.MarcasQueSeConfirman.de_fila(fila),
                    nota_revision_hash=cfg.huella_de_nota(fila.nota_revision),
                    actor="verificacion@script",
                )
                if not cambio:
                    fallas.append(f"La confirmación simulada del tipo {tipo} no cambió nada: ¿ya estaba confirmada?")

            for clave in ("B-03", "B-05"):
                definicion = registro.obtener(clave)
                resultado = await definicion.consultar(
                    db, EMPRESA_ID, definicion.Parametros(**_parametros_minimos(clave))
                )
                print(
                    f"{clave} con las marcas confirmadas: {len(resultado.filas)} filas, "
                    f"banderas {_resumen_de_banderas(resultado.banderas)}"
                )
                if "MARCA_SIN_CONFIRMAR" in _claves(resultado.banderas):
                    fallas.append(
                        f"{clave} siguió emitiendo MARCA_SIN_CONFIRMAR con las marcas de {tipos} confirmadas: "
                        "el aviso no se apaga al arreglar la causa que él mismo señala"
                    )
                # El otro lado: la columna que estaba vacía tiene que calcularse ahora. Sin esto,
                # «la bandera desapareció» sería compatible con que el informe siga sin calcular.
                titulos = _COLUMNAS_DE_TOPE_DE_B03 if clave == "B-03" else ("Gravado ordinario",)
                for titulo in titulos:
                    llenas = _columnas_no_vacias(resultado, titulo)
                    print(f"  {titulo}: {llenas}/{len(resultado.filas)} filas calculadas")
                    if resultado.filas and not llenas:
                        fallas.append(
                            f"{clave}: con las marcas confirmadas la columna '{titulo}' sigue vacía en las "
                            f"{len(resultado.filas)} filas, así que confirmar no habilitó el cálculo"
                        )
        finally:
            # En el `finally` a propósito: si una comprobación de arriba lanzara, la reversión tiene
            # que pasar igual. Nada de esto se confirmó nunca con `commit`, y salir del `async with`
            # también revertiría; el `rollback` explícito es para que la intención esté escrita.
            await db.rollback()

    async with SessionLocal() as verificacion:
        confirmadas = (
            await verificacion.execute(
                select(func.count())
                .select_from(CatalogoPercepcionMarca)
                .where(CatalogoPercepcionMarca.confirmado_en.is_not(None))
            )
        ).scalar_one()
    if int(confirmadas) != estado.marcas_confirmadas:
        fallas.append(
            f"La simulación dejó la base tocada: había {estado.marcas_confirmadas} marca(s) confirmada(s) antes "
            f"y hay {confirmadas} después. Revisa `map` y `catalogo_percepcion_marca` a mano antes de seguir"
        )
    else:
        print(
            f"Reversión comprobada en una sesión nueva: {confirmadas} marca(s) confirmada(s), las mismas que "
            f"antes de la simulación ({estado.marcas_confirmadas}). La base queda como se encontró."
        )
    return fallas


async def main() -> int:
    async with SessionLocal() as db:
        comprobantes_normalizados, fallas_b00 = await _verificar_identidades_b00(db)
        print(f"Identidades de B-00 por CFDI: {identidades_b00.IDENTIDADES_POR_COMPROBANTE}")

        fallas_b02 = await _verificar_b02(db, comprobantes_normalizados)
        valores_personales = await _valores_personales_del_universo(db)
        estado = await _estado_de_configuracion(db)
        resultados, fallas_catalogo = await _verificar_catalogo(db, comprobantes_normalizados, valores_personales)
        fallas_degradacion = _fallas_de_degradacion(estado, resultados)
        fallas_cobertura = await _verificar_cobertura_de_b08(db, estado)

    # Fuera de la sesión de lectura: la simulación abre la suya para poder revertirla sin tocar
    # la transacción en la que se auditaron los nueve informes.
    fallas_confirmado = await _verificar_camino_confirmado_con_reversion(estado)

    fallas = fallas_b00 + fallas_b02 + fallas_catalogo + fallas_degradacion + fallas_cobertura + fallas_confirmado
    if fallas:
        print("\nFALLAS:")
        for falla in fallas:
            print("  -", falla)
        return 1
    print("\nTodas las comprobaciones pasaron.")
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
