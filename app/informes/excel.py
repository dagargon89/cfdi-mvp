"""Construcción del libro de Excel común a todos los informes (spec §10).

Cuatro hojas siempre: `Datos`, `Parámetros`, `Banderas`, `Diccionario`. `openpyxl` en modo
`write_only` para no cargar el libro completo en memoria — el mismo patrón que
`exportar_excel` en `app/worker/tasks.py`.

`ROUND_HALF_UP` explícito: el redondeo por defecto de Python es `ROUND_HALF_EVEN`, que
para importes fiscales es incorrecto (⌊x⌉₂ del documento fuente es medio arriba).

**El redondeo lo derivan los propios formatos de presentación, y eso cierra un defecto real
(ronda de corrección 2 de la tarea 10).** La regla R-T4 dice que el redondeo ocurre **una sola
vez, aquí** — no que aquí no se redondee. Hasta esta ronda solo se cuantizaba el tipo `monto`,
así que un `decimal` llegaba a la celda con toda la precisión de la división que lo produjo
mientras su formato mostraba tres decimales: «Días de vacaciones pendientes» de B-08 guardaba
`12.88767123287671` y «% del total del periodo» de B-06 guardaba 28 dígitos significativos, los
dos **mostrando** tres decimales. Nadie ve la diferencia en la pantalla de Excel, y cualquiera la
ve al exportar a CSV, al comparar dos libros o al sumar la columna en otra herramienta: es
exactamente la clase de discrepancia entre lo guardado y lo mostrado que un auditor encuentra y
nadie puede explicar.

Por eso la escala de cada tipo numérico **se deriva de `_FORMATO`** (`_ESCALA_POR_TIPO`) en vez de
ser una constante aparte: el número de decimales que se guarda no puede separarse del que se
muestra ni siquiera si alguien cambia un formato, porque son el mismo dato leído una vez. Si un
informe necesita más precisión en una celda, lo que tiene que cambiar es el **formato** de su
tipo, que es la afirmación visible.

**El enmascaramiento de datos personales es responsabilidad de este módulo, no de cada
informe.** Un informe solo marca `Columna(..., sensible=True)` y entrega el valor en claro;
`escribir_libro` aplica `enmascarar()` a esas columnas cuando
`ContextoInforme.parametros["enmascarar_datos_personales"]` es verdadero. Centralizarlo aquí
evita que, con nueve informes por venir, un olvido en uno solo deje un CURP o un NSS
completo en un Excel que por diseño circula por correo.
"""

from __future__ import annotations

import io
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

from app.informes.base import Bandera, ContextoInforme, EntradaDiccionario, ResultadoInforme

_FORMATO = {
    "monto": "#,##0.00",
    "decimal": "#,##0.000",
    "entero": "#,##0",
    "fecha": "yyyy-mm-dd",
    "fecha_hora": "yyyy-mm-dd hh:mm:ss",
    "texto": "@",
}
_NEGRITA = Font(bold=True)

_TIPOS_NUMERICOS: tuple[str, ...] = ("monto", "decimal", "entero")
"""Los tipos de `base.TipoColumna` cuyas celdas llevan un número. `fecha`, `fecha_hora` y
`texto` quedan fuera a propósito: sus formatos no declaran decimales y sus valores nunca son
`Decimal`, así que no hay nada que cuantizar y sondear sus formatos solo abriría la puerta a
interpretar un `mm` como "minutos" o como "mes"."""


def _escala_del_formato(formato: str) -> Decimal:
    """Cuántos decimales **muestra** un formato numérico de Excel, como el exponente que
    `Decimal.quantize` necesita: `#,##0.00` → `0.01`, `#,##0.000` → `0.001`, `#,##0` → `1`.

    Es la única fuente de la escala de redondeo (ver el docstring del módulo). Se cuentan los
    dígitos declarados después del punto decimal del formato; sin punto, el formato muestra un
    entero y la escala es 1.
    """
    _entero, _punto, decimales = formato.partition(".")
    return Decimal(1).scaleb(-len(decimales))


_ESCALA_POR_TIPO: dict[str, Decimal] = {tipo: _escala_del_formato(_FORMATO[tipo]) for tipo in _TIPOS_NUMERICOS}
"""Escala de redondeo de cada tipo numérico, **derivada de su propio formato**. Que sea derivada
y no escrita es el punto: lo guardado y lo mostrado no pueden separarse."""

HOJAS_CON_ENCABEZADO = frozenset({"Datos", "Banderas", "Diccionario"})
"""De las cuatro hojas del libro, las que llevan una fila de **títulos** en la fila 1.

`Parámetros` no está: su primera fila ya es contenido (`["Informe", "B-05 · ..."]`, ver
`_escribir_parametros`), y su bloque de parámetros trae su propio encabezado más abajo. Lo consume
`validadores.fugas_de_datos_personales_en_libro` para nombrar la columna donde encuentra una fuga:
sin esta distinción reportaba una fuga de la hoja `Parámetros` como si estuviera en una columna
llamada "B-05 · Acumulado anual" —el contenido de la fila 1— en vez de decir su posición. La
constante vive aquí, en el módulo que **escribe** el libro, porque es un hecho de su estructura y
no del auditor."""


def enmascarar(valor: str | None) -> str | None:
    """`****` conservando los últimos 4 caracteres (spec §8). Un valor de 4 o menos se
    enmascara por completo: conservar 4 de 4 no enmascararía nada."""
    if valor is None:
        return None
    texto = str(valor)
    return "****" if len(texto) <= 4 else f"****{texto[-4:]}"


def _celda(valor: Any, tipo: str, *, sensible: bool = False) -> Any:
    """Único punto donde se redondea (R-T4) y donde se enmascara un valor sensible.

    **Se cuantiza todo `Decimal` de una columna numérica, no solo los `monto`**, y a la escala
    que su propio formato muestra: guardar más decimales de los que la celda enseña solo crea la
    posibilidad de que el número guardado y el mostrado difieran (ver el docstring del módulo).
    `int` no se toca —ya es exacto y su `entero` no tiene decimales que perder— y un `float` no
    debería llegar nunca aquí: los informes trabajan en `Decimal` de punta a punta.
    """
    if valor is None:
        return None
    if sensible:
        return enmascarar(str(valor))
    if isinstance(valor, Decimal):
        escala = _ESCALA_POR_TIPO.get(tipo)
        if escala is not None:
            return valor.quantize(escala, rounding=ROUND_HALF_UP)
    return valor


def _con_estilo(ws: Any, valor: Any, *, formato: str | None = None, negrita: bool = False) -> Any:
    """En modo `write_only` no existe `ws.cell(...)`: el estilo se aplica al construir la
    celda con `WriteOnlyCell` antes de hacer `append`."""
    celda = WriteOnlyCell(ws, value=valor)
    if formato:
        celda.number_format = formato
    if negrita:
        celda.font = _NEGRITA
    return celda


def _escribir_datos(wb: Workbook, resultado: ResultadoInforme, ctx: ContextoInforme) -> None:
    ws = wb.create_sheet("Datos")
    ws.freeze_panes = "A2"
    # Default `True` explícito: falla CERRADO. Si la clave no viene en el dict —un llamador
    # que no pase por el endpoint HTTP, que sí incluye los defaults al hacer `model_dump()`—,
    # `get()` sin default devolvería `None` y el libro saldría con CURP y NSS en claro, en
    # contra del default declarado y publicado en el JSON Schema del informe.
    enmascarar_activo = bool(ctx.parametros.get("enmascarar_datos_personales", True))

    ws.append([_con_estilo(ws, columna.titulo, negrita=True) for columna in resultado.columnas])

    for fila in resultado.filas:
        ws.append(
            [
                _con_estilo(
                    ws,
                    _celda(valor, columna.tipo, sensible=columna.sensible and enmascarar_activo),
                    formato=_FORMATO.get(columna.tipo),
                )
                for valor, columna in zip(fila, resultado.columnas)
            ]
        )


def _escribir_parametros(wb: Workbook, resultado: ResultadoInforme, ctx: ContextoInforme) -> None:
    ws = wb.create_sheet("Parámetros")
    ws.append(["Informe", f"{ctx.clave} · {ctx.nombre}"])
    ws.append(["Generado por", ctx.usuario])
    ws.append(["Generado el", ctx.generado_en])
    ws.append(["Versión del ETL", ctx.etl_version])
    ws.append(["Filas", len(resultado.filas)])
    if resultado.aviso:
        ws.append(["Aviso", resultado.aviso])
    # Las notas van ARRIBA del bloque de parámetros y no al final de la hoja: son advertencias
    # sobre cómo leer las cifras (B-08.R3: "esto es una estimación con base en CFDI, no un
    # cálculo actuarial") y tienen que verse sin desplazarse por la lista de parámetros.
    for nota in resultado.notas:
        ws.append(["Nota", nota])
    ws.append([])
    ws.append(["Parámetro", "Valor"])
    for clave, valor in ctx.parametros.items():
        ws.append([clave, "" if valor is None else str(valor)])


def _escribir_banderas(wb: Workbook, banderas: list[Bandera]) -> None:
    ws = wb.create_sheet("Banderas")
    ws.append(["Bandera", "Severidad", "Ámbito", "Mensaje"])
    for bandera in banderas:
        ws.append([bandera.clave, bandera.severidad, bandera.ambito, bandera.mensaje])


def _escribir_diccionario(wb: Workbook, entradas: list[EntradaDiccionario]) -> None:
    ws = wb.create_sheet("Diccionario")
    ws.append(
        [
            "Etiqueta",
            "Naturaleza",
            "Tipo SAT",
            "Descripción SAT",
            "Clave del patrón",
            "Concepto canónico",
            "Descripciones alternas",
            "Núm. comprobantes",
            "Importe del periodo",
        ]
    )
    for entrada in entradas:
        ws.append(
            [
                entrada.etiqueta,
                entrada.naturaleza,
                entrada.tipo,
                entrada.descripcion_sat,
                entrada.clave_patron,
                entrada.concepto_canonico,
                "; ".join(entrada.descripciones_alternas),
                entrada.num_comprobantes,
                # **Con su formato de importe**, no en `General` (hallazgo de la ronda 2 al auditar
                # los nueve libros reales): el valor ya salía cuantizado a dos decimales por
                # `_celda`, pero sin `number_format` la celda lo mostraba como `108757.8` en vez de
                # `108,757.80` — la misma columna con la misma cifra se veía distinta en `Datos` y
                # en `Diccionario`. Es la única celda numérica del libro fuera de `Datos` que lleva
                # un importe, y ahora es la única forma en que se puede escribir.
                _con_estilo(ws, _celda(entrada.importe_total, "monto"), formato=_FORMATO["monto"]),
            ]
        )


def escribir_libro(resultado: ResultadoInforme, ctx: ContextoInforme) -> bytes:
    """Devuelve los bytes del `.xlsx`. Las cuatro hojas existen siempre, aunque estén
    vacías: un consumidor automático no debería tener que comprobar si la hoja está."""
    wb = Workbook(write_only=True)
    _escribir_datos(wb, resultado, ctx)
    _escribir_parametros(wb, resultado, ctx)
    _escribir_banderas(wb, resultado.banderas)
    _escribir_diccionario(wb, resultado.diccionario)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
